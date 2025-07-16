#!/usr/bin/env python3
"""
Race Condition 전체 통합 통계 분석기
- 이상현상 탐지 결과를 바탕으로 4가지 규칙별 전체 통합 통계적 집계 분석
- 방/bin 구분 없이 전체 데이터셋에 대한 발생률, 심각도, 분포 분석 제공
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

def load_and_validate_data(preprocessor_file, analysis_file):
    """데이터 로드 및 필수 컬럼 검증"""
    print("📂 데이터 파일 로드 중...")
    
    # 전처리 데이터 로드
    preprocessor_df = pd.read_csv(preprocessor_file)
    print(f"✅ 전처리 데이터 로드 완료: {len(preprocessor_df)}행")
    
    # 이상현상 분석 데이터 로드
    analysis_df = pd.read_csv(analysis_file)
    print(f"✅ 이상현상 분석 데이터 로드 완료: {len(analysis_df)}행")
    
    # 전처리 데이터 필수 컬럼 검증
    preprocessor_required = ['roomNumber', 'bin', 'user_id']
    missing_preprocessor = [col for col in preprocessor_required if col not in preprocessor_df.columns]
    if missing_preprocessor:
        raise ValueError(f"전처리 데이터에서 필수 컬럼 누락: {missing_preprocessor}")
    
    # 이상현상 분석 데이터 필수 컬럼 검증
    analysis_required = ['roomNumber', 'bin', 'anomaly_type', 'lost_update_diff', 
                        'contention_group_size', 'over_capacity_amount', 'curr_sequence_diff']
    missing_analysis = [col for col in analysis_required if col not in analysis_df.columns]
    if missing_analysis:
        raise ValueError(f"이상현상 분석 데이터에서 필수 컬럼 누락: {missing_analysis}")
    
    # 데이터 타입 정리 (NaN/무한대 값 처리)
    preprocessor_df['roomNumber'] = pd.to_numeric(preprocessor_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    preprocessor_df['bin'] = pd.to_numeric(preprocessor_df['bin'], errors='coerce').fillna(0).astype(int)
    
    analysis_df['roomNumber'] = pd.to_numeric(analysis_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    analysis_df['bin'] = pd.to_numeric(analysis_df['bin'], errors='coerce').fillna(0).astype(int)
    
    print("✅ 데이터 검증 및 타입 변환 완료")
    return preprocessor_df, analysis_df

def calculate_total_requests(preprocessor_df):
    """전체 요청수 집계 (모든 방과 bin 통합)"""
    print("📊 전체 요청수 집계 중...")
    
    total_requests = len(preprocessor_df)
    total_rooms = preprocessor_df['roomNumber'].nunique()
    total_bins = len(preprocessor_df.groupby(['roomNumber', 'bin']))
    
    print(f"✅ 집계 완료:")
    print(f"  - 전체 요청수: {total_requests:,}건")
    print(f"  - 전체 방 수: {total_rooms}개")
    print(f"  - 전체 (방×bin) 조합: {total_bins}개")
    
    return {
        'total_requests': total_requests,
        'total_rooms': total_rooms,
        'total_bins': total_bins
    }

def calculate_overall_statistics(filtered_df, value_column, total_info, rule_name, use_absolute=False):
    """전체 통합 통계 계산 함수"""
    print(f"  📈 {rule_name} 통계 계산 중...")
    
    # 기본 정보
    total_requests = total_info['total_requests']
    total_rooms = total_info['total_rooms']
    total_bins = total_info['total_bins']
    
    # 초기 결과 딕셔너리
    result = {
        '분석 구분': rule_name,
        '전체 요청수': total_requests,
        '전체 방 수': total_rooms,
        '전체 (방×bin) 조합수': total_bins,
        '발생 건수': 0,
        '발생률 (%)': 0.0,
        '영향받은 방 수': 0,
        '영향받은 (방×bin) 조합수': 0,
        '방별 평균 발생률 (%)': 0.0,
        'bin별 평균 발생률 (%)': 0.0,
        '총합 값': 0.0,
        '평균 값': np.nan,
        '최소 값': np.nan,
        '최대 값': np.nan,
        '중간 값': np.nan,
        '표준편차 값': 0.0
    }
    
    # 이상현상이 있는 경우에만 실제 통계 계산
    if len(filtered_df) > 0:
        # 해당 값들 추출 (NaN 제거)
        values = filtered_df[value_column].dropna()
        
        if len(values) > 0:
            # 기본 발생 통계
            occurrence_count = len(values)
            occurrence_rate = (occurrence_count / total_requests * 100) if total_requests > 0 else 0
            
            # 영향받은 방과 bin 조합 계산
            affected_rooms = filtered_df['roomNumber'].nunique()
            affected_bins = len(filtered_df.groupby(['roomNumber', 'bin']))
            
            # 방별 발생률 평균 계산
            room_stats = filtered_df.groupby('roomNumber').size().reset_index(name='room_count')
            # 전체 방별 요청수 계산 (전처리 데이터에서)
            if 'preprocessor_df' in globals():
                room_requests = preprocessor_df.groupby('roomNumber').size().reset_index(name='total_room_requests')
                room_stats = room_stats.merge(room_requests, on='roomNumber', how='left')
                room_stats['room_rate'] = (room_stats['room_count'] / room_stats['total_room_requests'] * 100)
                avg_room_rate = room_stats['room_rate'].mean()
            else:
                avg_room_rate = 0.0
            
            # bin별 발생률 평균 계산
            bin_stats = filtered_df.groupby(['roomNumber', 'bin']).size().reset_index(name='bin_count')
            if 'preprocessor_df' in globals():
                bin_requests = preprocessor_df.groupby(['roomNumber', 'bin']).size().reset_index(name='total_bin_requests')
                bin_stats = bin_stats.merge(bin_requests, on=['roomNumber', 'bin'], how='left')
                bin_stats['bin_rate'] = (bin_stats['bin_count'] / bin_stats['total_bin_requests'] * 100)
                avg_bin_rate = bin_stats['bin_rate'].mean()
            else:
                avg_bin_rate = 0.0
            
            # 결과 업데이트
            result['발생 건수'] = int(occurrence_count)
            result['발생률 (%)'] = round(occurrence_rate, 2)
            result['영향받은 방 수'] = int(affected_rooms)
            result['영향받은 (방×bin) 조합수'] = int(affected_bins)
            result['방별 평균 발생률 (%)'] = round(avg_room_rate, 2)
            result['bin별 평균 발생률 (%)'] = round(avg_bin_rate, 2)
            
            # 값 통계 계산 (절댓값 옵션 적용)
            if use_absolute:
                result['총합 값'] = round(values.abs().sum(), 2)
                result['평균 값'] = round(values.abs().mean(), 2)
            else:
                result['총합 값'] = round(values.sum(), 2)
                result['평균 값'] = round(values.mean(), 2)
            
            result['최소 값'] = round(values.min(), 2)
            result['최대 값'] = round(values.max(), 2)
            result['중간 값'] = round(values.median(), 2)
            result['표준편차 값'] = round(values.std(), 4) if len(values) > 1 else 0.0
    
    return result

def analyze_lost_update(analysis_df, total_info):
    """규칙 1: 값 불일치 전체 통합 분석"""
    print("🔍 규칙 1: 값 불일치 (Lost Update) 전체 분석 중...")
    
    # '값 불일치' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].str.contains('값 불일치', na=False)]
    print(f"  - 값 불일치 발생 레코드: {len(filtered_df)}건")
    
    # lost_update_diff 기준 통계 계산
    result = calculate_overall_statistics(filtered_df, 'lost_update_diff', total_info, "규칙 1: 값 불일치 (Lost Update)")
    
    # 특화된 컬럼명 적용
    specialized_mapping = {
        '총합 값': '오차 누적 총합',
        '평균 값': '평균 오차',
        '최소 값': '최소 오차',
        '최대 값': '최대 오차',
        '중간 값': '중간값 오차',
        '표준편차 값': '오차 표준편차'
    }
    
    for old_key, new_key in specialized_mapping.items():
        if old_key in result:
            result[new_key] = result.pop(old_key)
    
    print("✅ 값 불일치 전체 분석 완료")
    return result

def analyze_contention(analysis_df, total_info):
    """규칙 2: 경합 발생 전체 통합 분석"""
    print("🔍 규칙 2: 경합 발생 (Contention) 전체 분석 중...")
    
    # '경합 발생 오류' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].str.contains('경합 발생 오류', na=False)]
    print(f"  - 경합 발생 레코드: {len(filtered_df)}건")
    
    # contention_group_size 기준 통계 계산
    result = calculate_overall_statistics(filtered_df, 'contention_group_size', total_info, "규칙 2: 경합 발생 (Contention)")
    
    # 특화된 컬럼명 적용
    specialized_mapping = {
        '총합 값': '총 경합 스레드 수',
        '평균 값': '평균 경합 스레드 수',
        '최소 값': '최소 경합 그룹 크기',
        '최대 값': '최대 경합 스레드 수',
        '중간 값': '중간값 경합 그룹 크기',
        '표준편차 값': '경합 강도 표준편차'
    }
    
    for old_key, new_key in specialized_mapping.items():
        if old_key in result:
            result[new_key] = result.pop(old_key)
    
    print("✅ 경합 발생 전체 분석 완료")
    return result

def analyze_capacity_exceeded(analysis_df, total_info):
    """규칙 3: 정원 초과 전체 통합 분석"""
    print("🔍 규칙 3: 정원 초과 (Capacity Exceeded) 전체 분석 중...")
    
    # '정원 초과 오류' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].str.contains('정원 초과 오류', na=False)]
    print(f"  - 정원 초과 발생 레코드: {len(filtered_df)}건")
    
    # over_capacity_amount 기준 통계 계산
    result = calculate_overall_statistics(filtered_df, 'over_capacity_amount', total_info, "규칙 3: 정원 초과 (Capacity Exceeded)")
    
    # 특화된 컬럼명 적용
    specialized_mapping = {
        '총합 값': '총 초과 인원',
        '평균 값': '평균 초과 인원',
        '최소 값': '최소 초과 인원',
        '최대 값': '최대 초과 인원',
        '중간 값': '중간값 초과 인원',
        '표준편차 값': '초과 규모 표준편차'
    }
    
    for old_key, new_key in specialized_mapping.items():
        if old_key in result:
            result[new_key] = result.pop(old_key)
    
    print("✅ 정원 초과 전체 분석 완료")
    return result

def analyze_state_transition(analysis_df, total_info):
    """규칙 4: 상태 전이 오류 전체 통합 분석"""
    print("🔍 규칙 4: 상태 전이 오류 (State Transition) 전체 분석 중...")
    
    # '상태 전이 오류' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].str.contains('상태 전이 오류', na=False)]
    print(f"  - 상태 전이 오류 발생 레코드: {len(filtered_df)}건")
    
    # curr_sequence_diff 기준 통계 계산 (절댓값 사용)
    result = calculate_overall_statistics(filtered_df, 'curr_sequence_diff', total_info, "규칙 4: 상태 전이 오류 (State Transition)", use_absolute=True)
    
    # 상태 전이 오류는 기본 컬럼명 유지 (다른 규칙들과 동일하게)
    
    print("✅ 상태 전이 오류 전체 분석 완료")
    return result

def create_individual_dataframes(lost_update_result, contention_result, capacity_result, state_transition_result):
    """4개 규칙 결과를 각각의 DataFrame으로 생성"""
    print("📊 규칙별 개별 DataFrame 생성 중...")
    
    # 각 규칙별 DataFrame 생성
    lost_update_df = pd.DataFrame([lost_update_result])
    contention_df = pd.DataFrame([contention_result])
    capacity_df = pd.DataFrame([capacity_result])
    state_transition_df = pd.DataFrame([state_transition_result])
    
    print("✅ 개별 DataFrame 생성 완료")
    return lost_update_df, contention_df, capacity_df, state_transition_df

def add_dataframe_to_sheet(ws, df, sheet_title):
    """DataFrame을 워크시트에 추가하고 스타일 적용"""
    # 시트 제목 추가
    ws['A1'] = sheet_title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    
    # 빈 행 추가
    start_row = 3
    
    # 컬럼 헤더 추가
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="B7D7F1", end_color="B7D7F1", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 추가
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 분석 구분 컬럼은 강조 처리
            if col_idx == 1:  # 첫 번째 컬럼 (분석 구분)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    
    # 컬럼 너비 조정
    for col_idx, column in enumerate(df.columns, 1):
        max_length = max(len(str(column)), 12)
        if len(df) > 0:
            col_values = [str(val) for val in df.iloc[:, col_idx-1] if pd.notna(val)]
            if col_values:
                max_length = max(max_length, max(len(val) for val in col_values))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 3, 35)

def create_excel_output(lost_update_df, contention_df, capacity_df, state_transition_df, output_file):
    """4개 시트로 구성된 Excel 파일 생성 (각 규칙별 개별 시트)"""
    print("📊 Excel 파일 생성 중...")
    
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 시트 1: Lost Update 전체 분석
    ws1 = wb.create_sheet("Overall_LostUpdate")
    add_dataframe_to_sheet(ws1, lost_update_df, "전체 통합: 규칙 1 - 값 불일치 (Lost Update) 분석")
    
    # 시트 2: Contention 전체 분석
    ws2 = wb.create_sheet("Overall_Contention") 
    add_dataframe_to_sheet(ws2, contention_df, "전체 통합: 규칙 2 - 경합 발생 (Contention) 분석")
    
    # 시트 3: Capacity Exceeded 전체 분석
    ws3 = wb.create_sheet("Overall_Capacity")
    add_dataframe_to_sheet(ws3, capacity_df, "전체 통합: 규칙 3 - 정원 초과 (Capacity Exceeded) 분석")
    
    # 시트 4: State Transition 전체 분석
    ws4 = wb.create_sheet("Overall_StateTransition")
    add_dataframe_to_sheet(ws4, state_transition_df, "전체 통합: 규칙 4 - 상태 전이 오류 (State Transition) 분석")
    
    wb.save(output_file)
    print(f"✅ Excel 파일 저장 완료: {output_file}")
    print("  📋 생성된 시트:")
    print("    - Overall_LostUpdate: 값 불일치 전체 분석")
    print("    - Overall_Contention: 경합 발생 전체 분석") 
    print("    - Overall_Capacity: 정원 초과 전체 분석")
    print("    - Overall_StateTransition: 상태 전이 오류 전체 분석")

def print_summary_statistics(lost_update_df, contention_df, capacity_df, state_transition_df):
    """분석 결과 요약 통계 출력"""
    print("\n" + "="*90)
    print("📈 RACE CONDITION 전체 통합 분석 결과 요약")
    print("="*90)
    
    # 각 DataFrame에서 정보 추출
    dataframes = [
        ("규칙 1: 값 불일치 (Lost Update)", lost_update_df),
        ("규칙 2: 경합 발생 (Contention)", contention_df),
        ("규칙 3: 정원 초과 (Capacity Exceeded)", capacity_df),
        ("규칙 4: 상태 전이 오류 (State Transition)", state_transition_df)
    ]
    
    # 첫 번째 규칙에서 전체 정보 추출
    if len(lost_update_df) > 0:
        first_row = lost_update_df.iloc[0]
        total_requests = first_row['전체 요청수']
        total_rooms = first_row['전체 방 수']
        total_bins = first_row['전체 (방×bin) 조합수']
        
        print(f"전체 분석 대상:")
        print(f"  - 요청 수: {total_requests:,}건")
        print(f"  - 방 수: {total_rooms}개")
        print(f"  - (방×bin) 조합: {total_bins}개")
    
    print(f"\n--- 규칙별 상세 분석 결과 ---")
    
    total_anomaly_requests = 0
    
    for rule_name, df in dataframes:
        if len(df) > 0:
            row = df.iloc[0]
            occurrence_count = row['발생 건수']
            occurrence_rate = row['발생률 (%)']
            affected_rooms = row['영향받은 방 수']
            affected_bins = row['영향받은 (방×bin) 조합수']
            
            print(f"\n{rule_name}:")
            print(f"  - 발생 건수: {occurrence_count:,}건 ({occurrence_rate}%)")
            print(f"  - 영향받은 방: {affected_rooms}개")
            print(f"  - 영향받은 bin: {affected_bins}개")
            
            total_anomaly_requests += occurrence_count
        else:
            print(f"\n{rule_name}: 데이터 없음")
    
    if len(lost_update_df) > 0:
        overall_anomaly_rate = (total_anomaly_requests / total_requests * 100) if total_requests > 0 else 0
        
        print(f"\n--- 전체 요약 ---")
        print(f"전체 이상현상 발생: {total_anomaly_requests:,}건")
        print(f"전체 이상현상 발생률: {overall_anomaly_rate:.2f}%")
        print(f"정상 요청: {total_requests - total_anomaly_requests:,}건 ({100 - overall_anomaly_rate:.2f}%)")

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description="Race Condition 전체 통합 통계 분석기")
    parser.add_argument('preprocessor_csv', help='전처리 결과 CSV 파일')
    parser.add_argument('analysis_csv', help='이상현상 분석 결과 CSV 파일')
    parser.add_argument('output_xlsx', help='전체 통합 분석 Excel 출력 파일')
    parser.add_argument('--rooms', help='분석할 방 번호 (쉼표로 구분)')
    
    args = parser.parse_args()
    
    try:
        print("🚀 Race Condition 전체 통합 분석기 시작...")
        print(f"입력 파일 1: {args.preprocessor_csv}")
        print(f"입력 파일 2: {args.analysis_csv}")
        print(f"출력 파일: {args.output_xlsx}")
        
        # 1. 데이터 로드 및 검증
        preprocessor_df, analysis_df = load_and_validate_data(args.preprocessor_csv, args.analysis_csv)
        
        # 전역 변수로 설정 (통계 계산 함수에서 참조하기 위해)
        globals()['preprocessor_df'] = preprocessor_df
        
        # 2. 방 번호 필터링 (선택사항)
        if args.rooms:
            room_numbers = [int(room.strip()) for room in args.rooms.split(',')]
            preprocessor_df = preprocessor_df[preprocessor_df['roomNumber'].isin(room_numbers)]
            analysis_df = analysis_df[analysis_df['roomNumber'].isin(room_numbers)]
            globals()['preprocessor_df'] = preprocessor_df  # 전역 변수 업데이트
            print(f"🔍 방 번호 {room_numbers}로 필터링 적용")
        
        # 3. 전체 요청 정보 집계
        total_info = calculate_total_requests(preprocessor_df)
        
        # 4. 4가지 규칙별 전체 통합 분석
        lost_update_result = analyze_lost_update(analysis_df, total_info)
        contention_result = analyze_contention(analysis_df, total_info)
        capacity_result = analyze_capacity_exceeded(analysis_df, total_info)
        state_transition_result = analyze_state_transition(analysis_df, total_info)
        
        # 5. 개별 규칙별 DataFrame 생성
        lost_update_df, contention_df, capacity_df, state_transition_df = create_individual_dataframes(
            lost_update_result, contention_result, capacity_result, state_transition_result)
        
        # 6. Excel 파일 생성 (4개 시트로 분리)
        create_excel_output(lost_update_df, contention_df, capacity_df, state_transition_df, args.output_xlsx)
        
        # 7. 요약 통계 출력
        print_summary_statistics(lost_update_df, contention_df, capacity_df, state_transition_df)
        
        print("\n🎉 전체 통합 분석 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()