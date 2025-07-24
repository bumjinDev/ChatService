#!/usr/bin/env python3
"""
Semaphore 방별 통계 분석기
- 세마포어의 3가지 핵심 분석 영역을 방별로 통계 집계
- 분석 1: 순차적 일관성 방별 관찰 (규칙 1+4 통합) - 현상 관찰
- 분석 2: 동시 실행 패턴 방별 관찰 (규칙 2 재정의) - 현상 관찰  
- 분석 3: 정원 초과 방지 방별 검증 (규칙 3 유지) - 성공/실패 판정
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

def load_and_validate_semaphore_data(preprocessor_file, analysis_file):
    """세마포어 데이터 로드 및 필수 컬럼 검증"""
    print("📂 세마포어 데이터 파일 로드 중...")
    
    # 전처리 데이터 로드
    preprocessor_df = pd.read_csv(preprocessor_file)
    print(f"✅ 세마포어 전처리 데이터 로드 완료: {len(preprocessor_df)}행")
    
    # 분석 결과 데이터 로드
    analysis_df = pd.read_csv(analysis_file)
    print(f"✅ 세마포어 분석 결과 데이터 로드 완료: {len(analysis_df)}행")
    
    # 세마포어 전처리 데이터 필수 컬럼 검증 (10개 컬럼)
    preprocessor_required = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 
                           'max_people', 'room_entry_sequence', 'join_result',
                           'true_critical_section_nanoTime_start', 'true_critical_section_nanoTime_end']
    missing_preprocessor = [col for col in preprocessor_required if col not in preprocessor_df.columns]
    if missing_preprocessor:
        raise ValueError(f"세마포어 전처리 데이터에서 필수 컬럼 누락: {missing_preprocessor}")
    
    # 세마포어 분석 결과 데이터 필수 컬럼 검증 (14개 컬럼)
    analysis_required = ['roomNumber', 'bin', 'user_id', 'anomaly_type', 
                        'over_capacity_amount', 'over_capacity_curr', 'over_capacity_max']
    missing_analysis = [col for col in analysis_required if col not in analysis_df.columns]
    if missing_analysis:
        raise ValueError(f"세마포어 분석 결과 데이터에서 필수 컬럼 누락: {missing_analysis}")
    
    # 데이터 타입 정리
    preprocessor_df['roomNumber'] = pd.to_numeric(preprocessor_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    preprocessor_df['bin'] = pd.to_numeric(preprocessor_df['bin'], errors='coerce').fillna(0).astype(int)
    
    analysis_df['roomNumber'] = pd.to_numeric(analysis_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    analysis_df['bin'] = pd.to_numeric(analysis_df['bin'], errors='coerce').fillna(0).astype(int)
    
    print("✅ 세마포어 데이터 검증 및 타입 변환 완료")
    return preprocessor_df, analysis_df

def calculate_total_requests_per_room(preprocessor_df):
    """방별 전체 요청수 집계 (모든 bin 통합)"""
    print("📊 방별 전체 요청수 집계 중...")
    
    total_requests = preprocessor_df.groupby(['roomNumber']).size().reset_index(name='total_requests')
    
    print(f"✅ 집계 완료: {len(total_requests)}개 방")
    return total_requests

def analyze_sequential_consistency_per_room(preprocessor_df, analysis_df, total_requests_df):
    """분석 1: 순차적 일관성 방별 관찰 (규칙 1+4 통합) - 현상 관찰"""
    print("🔍 분석 1: 순차적 일관성 방별 관찰 (규칙 1+4 통합) 중...")
    
    # 전체 방 리스트를 기준으로 결과 생성
    result_stats = total_requests_df.copy()
    
    # 통계 컬럼들 초기화
    result_stats['순차적 일관성 차이 건수'] = 0
    result_stats['순차적 일관성 차이 발생률 (%)'] = 0.0
    result_stats['순차적 처리 일치율 (%)'] = 100.0
    result_stats['평균 차이 크기'] = 0.0
    result_stats['최소 차이'] = np.nan
    result_stats['최대 차이'] = np.nan
    result_stats['중간값 차이'] = np.nan
    result_stats['차이 표준편차'] = 0.0
    
    # 방별로 순차적 일관성 분석
    for room_num in preprocessor_df['roomNumber'].unique():
        room_data = preprocessor_df[preprocessor_df['roomNumber'] == room_num].copy()
        
        if len(room_data) == 0:
            continue
            
        # room_entry_sequence 기반 이상적 순차 상태 계산
        differences = []
        for _, row in room_data.iterrows():
            initial_people = 1  # 테스트 시작 시 방에 1명 존재
            ideal_sequential_state = min(initial_people + row['room_entry_sequence'], row['max_people'])
            
            if row['curr_people'] != ideal_sequential_state:
                diff = abs(row['curr_people'] - ideal_sequential_state)
                differences.append(diff)
        
        # 결과 DataFrame에서 해당 방 찾기
        mask = (result_stats['roomNumber'] == room_num)
        row_idx = result_stats[mask].index
        
        if len(row_idx) == 0:
            continue
            
        row_idx = row_idx[0]
        total_requests = result_stats.loc[row_idx, 'total_requests']
        
        # 통계 계산
        difference_count = len(differences)
        difference_rate = (difference_count / total_requests * 100) if total_requests > 0 else 0
        consistency_rate = ((total_requests - difference_count) / total_requests * 100) if total_requests > 0 else 100
        
        # 결과 업데이트
        result_stats.loc[row_idx, '순차적 일관성 차이 건수'] = difference_count
        result_stats.loc[row_idx, '순차적 일관성 차이 발생률 (%)'] = round(difference_rate, 2)
        result_stats.loc[row_idx, '순차적 처리 일치율 (%)'] = round(consistency_rate, 2)
        
        if len(differences) > 0:
            differences_array = np.array(differences)
            result_stats.loc[row_idx, '평균 차이 크기'] = round(differences_array.mean(), 2)
            result_stats.loc[row_idx, '최소 차이'] = round(differences_array.min(), 2)
            result_stats.loc[row_idx, '최대 차이'] = round(differences_array.max(), 2)
            result_stats.loc[row_idx, '중간값 차이'] = round(np.median(differences_array), 2)
            result_stats.loc[row_idx, '차이 표준편차'] = round(differences_array.std(), 4) if len(differences) > 1 else 0.0
    
    # 컬럼명 정리
    final_columns = {
        'roomNumber': '방 번호',
        'total_requests': '전체 요청수'
    }
    
    result_stats = result_stats.rename(columns=final_columns)
    
    print(f"✅ 순차적 일관성 방별 분석 완료: {len(result_stats)}개 방")
    return result_stats

def calculate_statistics(filtered_df, value_column, total_requests_df, use_absolute=False):
    """방별 통계 계산 함수 - 기존 동시성 측정과 동일한 방식"""
    # 전체 방 리스트를 기준으로 결과 생성
    result_stats = total_requests_df.copy()
    
    # 통계 컬럼들 초기화 (기존 방식과 동일)
    result_stats['occurrence_count'] = 0
    result_stats['occurrence_rate'] = 0.0
    result_stats['sum_value'] = 0.0
    result_stats['avg_value'] = np.nan
    result_stats['min_value'] = np.nan
    result_stats['max_value'] = np.nan
    result_stats['median_value'] = np.nan
    result_stats['std_value'] = 0.0
    
    # 이상현상이 있는 경우에만 실제 통계 계산
    if len(filtered_df) > 0:
        # roomNumber 단위로 그룹화
        grouped = filtered_df.groupby(['roomNumber'])
        
        for room_num, group in grouped:
            # 해당 방의 모든 값들 추출 (NaN 제거)
            values = group[value_column].dropna()
            
            if len(values) == 0:
                continue
            
            # 결과 DataFrame에서 해당 방 찾기
            mask = (result_stats['roomNumber'] == room_num)
            row_idx = result_stats[mask].index
            
            if len(row_idx) == 0:
                continue
                
            row_idx = row_idx[0]
            total_requests = result_stats.loc[row_idx, 'total_requests']
            
            # 통계 계산
            occurrence_count = len(values)
            occurrence_rate = (occurrence_count / total_requests * 100) if total_requests > 0 else 0
            
            # 결과 업데이트 (절댓값 옵션 적용)
            result_stats.loc[row_idx, 'occurrence_count'] = int(occurrence_count)
            result_stats.loc[row_idx, 'occurrence_rate'] = round(occurrence_rate, 2)
            
            if use_absolute:
                result_stats.loc[row_idx, 'sum_value'] = round(values.abs().sum(), 2)
                result_stats.loc[row_idx, 'avg_value'] = round(values.abs().mean(), 2)
            else:
                result_stats.loc[row_idx, 'sum_value'] = round(values.sum(), 2)
                result_stats.loc[row_idx, 'avg_value'] = round(values.mean(), 2)
            
            result_stats.loc[row_idx, 'min_value'] = round(values.min(), 2)
            result_stats.loc[row_idx, 'max_value'] = round(values.max(), 2)
            result_stats.loc[row_idx, 'median_value'] = round(values.median(), 2)
            result_stats.loc[row_idx, 'std_value'] = round(values.std(), 4) if len(values) > 1 else 0.0
    
    # 데이터 타입 정리
    result_stats['roomNumber'] = result_stats['roomNumber'].astype(int)
    result_stats['total_requests'] = result_stats['total_requests'].astype(int)
    result_stats['occurrence_count'] = result_stats['occurrence_count'].fillna(0).astype(int)
    
    return result_stats

def analyze_concurrent_execution_per_room(preprocessor_df, analysis_df, total_requests_df):
    """분석 2: 동시 실행 패턴 방별 관찰 (규칙 2 재정의) - 기존 방식으로 통일"""
    print("🔍 분석 2: 동시 실행 패턴 방별 관찰 (규칙 2 재정의) 중...")
    
    # contention_group_size 데이터가 있는 레코드 필터링 (기존 방식과 동일)
    filtered_df = analysis_df[
        (analysis_df['contention_group_size'].notna()) & 
        (analysis_df['contention_group_size'] > 1)
    ]
    print(f"  - 동시 실행 패턴 관련 레코드: {len(filtered_df)}건")
    
    # contention_group_size 기준 통계 계산 (기존 calculate_statistics 방식 사용)
    stats_df = calculate_statistics(filtered_df, 'contention_group_size', total_requests_df)
    
    # 컬럼명 변경 (기존 동시성 측정과 동일하게)
    column_mapping = {
        'sum_value': '총 경합 스레드 수',
        'avg_value': '평균 경합 스레드 수',
        'min_value': '최소 경합 그룹 크기',
        'max_value': '최대 경합 스레드 수',
        'median_value': '중간값 경합 그룹 크기',
        'std_value': '경합 강도 표준편차'
    }
    
    final_columns = {
        'roomNumber': '방 번호',
        'total_requests': '전체 요청수',
        'occurrence_count': '발생 건수',
        'occurrence_rate': '발생률 (%)',
        **column_mapping
    }
    
    stats_df = stats_df.rename(columns=final_columns)
    
    print(f"✅ 동시 실행 패턴 방별 분석 완료: {len(stats_df)}개 방")
    return stats_df

def analyze_capacity_prevention_per_room(preprocessor_df, analysis_df, total_requests_df):
    """분석 3: 정원 초과 방지 방별 검증 (규칙 3 유지) - 성공/실패 판정"""
    print("🔍 분석 3: 정원 초과 방지 방별 검증 (규칙 3 유지) 중...")
    
    # anomaly_type이 NaN인 경우 빈 문자열로 처리
    analysis_df['anomaly_type'] = analysis_df['anomaly_type'].fillna('')
    
    # 전체 방 리스트를 기준으로 결과 생성
    result_stats = total_requests_df.copy()
    
    # 통계 컬럼들 초기화
    result_stats['정원 초과 발생 건수'] = 0
    result_stats['정원 초과 발생률 (%)'] = 0.0
    result_stats['정원 초과 방지 성공률 (%)'] = 100.0
    result_stats['평균 초과 인원'] = 0.0
    result_stats['최소 초과 인원'] = np.nan
    result_stats['최대 초과 인원'] = np.nan
    result_stats['중간값 초과 인원'] = np.nan
    result_stats['초과 규모 표준편차'] = 0.0
    
    # 정원 초과 오류 필터링
    capacity_exceeded = analysis_df[analysis_df['anomaly_type'].str.contains('정원 초과 오류', na=False)]
    
    if len(capacity_exceeded) > 0:
        # 방별로 그룹화하여 통계 계산
        grouped = capacity_exceeded.groupby(['roomNumber'])
        
        for room_num, group in grouped:
            # 해당 방의 모든 초과 인원 값들 추출
            exceeded_amounts = group['over_capacity_amount'].dropna()
            
            if len(exceeded_amounts) == 0:
                continue
            
            # 결과 DataFrame에서 해당 방 찾기
            mask = (result_stats['roomNumber'] == room_num)
            row_idx = result_stats[mask].index
            
            if len(row_idx) == 0:
                continue
                
            row_idx = row_idx[0]
            total_requests = result_stats.loc[row_idx, 'total_requests']
            
            # 통계 계산
            exceeded_count = len(exceeded_amounts)
            exceeded_rate = (exceeded_count / total_requests * 100) if total_requests > 0 else 0
            prevention_rate = ((total_requests - exceeded_count) / total_requests * 100) if total_requests > 0 else 100
            
            # 결과 업데이트
            result_stats.loc[row_idx, '정원 초과 발생 건수'] = exceeded_count
            result_stats.loc[row_idx, '정원 초과 발생률 (%)'] = round(exceeded_rate, 2)
            result_stats.loc[row_idx, '정원 초과 방지 성공률 (%)'] = round(prevention_rate, 2)
            result_stats.loc[row_idx, '평균 초과 인원'] = round(exceeded_amounts.mean(), 2)
            result_stats.loc[row_idx, '최소 초과 인원'] = round(exceeded_amounts.min(), 2)
            result_stats.loc[row_idx, '최대 초과 인원'] = round(exceeded_amounts.max(), 2)
            result_stats.loc[row_idx, '중간값 초과 인원'] = round(exceeded_amounts.median(), 2)
            result_stats.loc[row_idx, '초과 규모 표준편차'] = round(exceeded_amounts.std(), 4) if len(exceeded_amounts) > 1 else 0.0
    
    # 컬럼명 정리
    final_columns = {
        'roomNumber': '방 번호',
        'total_requests': '전체 요청수'
    }
    
    result_stats = result_stats.rename(columns=final_columns)
    
    print(f"✅ 정원 초과 방지 방별 검증 완료: {len(result_stats)}개 방")
    return result_stats

def add_dataframe_to_sheet(ws, df, sheet_title):
    """DataFrame을 워크시트에 추가 (단순화된 스타일)"""
    # 시트 제목 추가
    ws['A1'] = sheet_title
    ws['A1'].font = Font(size=14, bold=True)
    
    # 시작 행
    start_row = 3
    
    # 컬럼 헤더 추가
    for col_idx, column in enumerate(df.columns, 1):
        # 컬럼명 길이 제한 및 특수문자 제거
        safe_column = str(column).replace('%', 'percent').replace('(', '').replace(')', '')[:30]
        cell = ws.cell(row=start_row, column=col_idx, value=safe_column)
        cell.font = Font(bold=True)
    
    # 데이터 추가
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            # 데이터 값 안전성 처리
            safe_value = value
            if isinstance(value, str):
                safe_value = str(value).replace('%', 'percent')[:100]
            elif pd.isna(value):
                safe_value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
    
    # 컬럼 너비 기본 설정
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

def create_excel_output(sequential_df, concurrent_df, capacity_df, output_file):
    """3개 시트로 구성된 Excel 파일 생성 (안전한 방식)"""
    print("📊 세마포어 방별 Excel 파일 생성 중...")
    
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 시트 1: 순차적 일관성 방별 분석
    ws1 = wb.create_sheet("Room_Sequential_Analysis")
    add_dataframe_to_sheet(ws1, sequential_df, "Room Analysis 1: Sequential Consistency")
    
    # 시트 2: 동시 실행 패턴 방별 분석 (시트명 변경: 기존과 통일)
    ws2 = wb.create_sheet("Room_Contention_Analysis")
    add_dataframe_to_sheet(ws2, concurrent_df, "Room Analysis 2: Contention")
    
    # 시트 3: 정원 초과 방지 방별 검증
    ws3 = wb.create_sheet("Room_Capacity_Analysis")
    add_dataframe_to_sheet(ws3, capacity_df, "Room Analysis 3: Capacity Prevention")
    
    wb.save(output_file)
    print(f"✅ 세마포어 방별 Excel 파일 저장 완료: {output_file}")
    print("  📋 생성된 시트:")
    print("    - Room_Sequential_Analysis: 순차적 일관성 방별 관찰")
    print("    - Room_Contention_Analysis: 동시 실행 패턴 방별 관찰 (기존과 통일)")
    print("    - Room_Capacity_Analysis: 정원 초과 방지 방별 검증")

def print_summary_statistics(sequential_df, concurrent_df, capacity_df, total_requests_df):
    """분석 결과 요약 통계 출력"""
    print("\n" + "="*80)
    print("📈 SEMAPHORE 방별 통계 분석 결과 요약")
    print("="*80)
    
    total_rooms = len(total_requests_df)
    total_requests_sum = total_requests_df['total_requests'].sum()
    
    print(f"전체 분석 대상: {total_rooms}개 방")
    print(f"전체 요청 수: {total_requests_sum:,}건")
    
    print(f"\n--- 세마포어 3가지 분석 결과 ---")
    
    # 분석 1: 순차적 일관성
    if len(sequential_df) > 0:
        sequential_differences = sequential_df['순차적 일관성 차이 건수'].sum()
        print(f"분석 1 (순차적 일관성 관찰): {total_rooms}개 방에서 {sequential_differences}건 차이 관찰")
    
    # 분석 2: 동시 실행 (기존 방식으로 출력)
    if len(concurrent_df) > 0:
        concurrent_executions = concurrent_df['발생 건수'].sum()
        max_concurrent_level = concurrent_df['최대 경합 스레드 수'].max()
        print(f"분석 2 (동시 실행 패턴 관찰): {total_rooms}개 방에서 {concurrent_executions}건 경합 관찰 (최대 {max_concurrent_level}개 스레드)")
    
    # 분석 3: 정원 초과
    if len(capacity_df) > 0:
        capacity_exceeded = capacity_df['정원 초과 발생 건수'].sum()
        affected_rooms = len(capacity_df[capacity_df['정원 초과 발생 건수'] > 0])
        print(f"분석 3 (정원 초과 방지 검증): {affected_rooms}개 방에서 {capacity_exceeded}건 초과 발생")
    
    print(f"\n--- 세마포어 종합 평가 ---")
    if len(capacity_df) > 0 and capacity_df['정원 초과 발생 건수'].sum() == 0:
        print("🎉 세마포어 핵심 기능 (정원 초과 방지): 모든 방에서 완벽 성공")
    elif len(capacity_df) > 0:
        print("⚠️ 세마포어 핵심 기능 (정원 초과 방지): 일부 방에서 검증 실패")
    
    if len(sequential_df) > 0 and sequential_df['순차적 일관성 차이 건수'].sum() > 0:
        print("📝 순차적 일관성: 부분적 보장 특성 관찰 (세마포어 고유 특성)")
    
    if len(concurrent_df) > 0 and concurrent_df['발생 건수'].sum() > 0:
        print("🚀 동시성 활용: CAS 기반 효율적 동시 실행 관찰 (의도된 정상 동작)")

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description="세마포어 방별 통계 분석기")
    parser.add_argument('preprocessor_csv', help='세마포어 전처리 결과 CSV 파일 (preprocessor_semaphore.csv)')
    parser.add_argument('analysis_csv', help='세마포어 분석 결과 CSV 파일 (semaphore_analysis_result.csv)')
    parser.add_argument('output_xlsx', help='세마포어 방별 통계 분석 Excel 출력 파일')
    parser.add_argument('--rooms', help='분석할 방 번호 (쉼표로 구분)')
    
    args = parser.parse_args()
    
    try:
        print("🚀 세마포어 방별 통계 분석기 시작...")
        print("🎯 세마포어 3가지 방별 분석: 순차적 일관성 관찰 + 동시 실행 패턴 관찰 + 정원 초과 방지 검증")
        print(f"입력 파일 1: {args.preprocessor_csv}")
        print(f"입력 파일 2: {args.analysis_csv}")
        print(f"출력 파일: {args.output_xlsx}")
        
        # 1. 세마포어 데이터 로드 및 검증
        preprocessor_df, analysis_df = load_and_validate_semaphore_data(args.preprocessor_csv, args.analysis_csv)
        
        # 2. 방 번호 필터링 (선택사항)
        if args.rooms:
            room_numbers = [int(room.strip()) for room in args.rooms.split(',')]
            preprocessor_df = preprocessor_df[preprocessor_df['roomNumber'].isin(room_numbers)]
            analysis_df = analysis_df[analysis_df['roomNumber'].isin(room_numbers)]
            print(f"🔍 방 번호 {room_numbers}로 필터링 적용")
        
        # 3. 방별 전체 요청수 집계
        total_requests_df = calculate_total_requests_per_room(preprocessor_df)
        
        # 4. 세마포어 3가지 방별 분석
        print("\n=== 🎯 세마포어 3가지 방별 분석 실행 ===")
        sequential_df = analyze_sequential_consistency_per_room(preprocessor_df, analysis_df, total_requests_df)
        concurrent_df = analyze_concurrent_execution_per_room(preprocessor_df, analysis_df, total_requests_df)
        capacity_df = analyze_capacity_prevention_per_room(preprocessor_df, analysis_df, total_requests_df)
        
        # 5. Excel 파일 생성
        create_excel_output(sequential_df, concurrent_df, capacity_df, args.output_xlsx)
        
        # 6. 요약 통계 출력
        print_summary_statistics(sequential_df, concurrent_df, capacity_df, total_requests_df)
        
        print("\n🎉 세마포어 방별 통계 분석 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()