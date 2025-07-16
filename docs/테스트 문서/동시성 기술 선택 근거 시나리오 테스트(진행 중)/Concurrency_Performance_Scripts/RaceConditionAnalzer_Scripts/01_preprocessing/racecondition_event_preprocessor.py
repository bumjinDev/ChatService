import pandas as pd  # 데이터 처리를 위한 라이브러리
import re             # 정규 표현식을 위한 라이브러리
import os             # 운영체제 기능을 위한 라이브러리
import shutil         # 파일 복사/이동을 위한 라이브러리
import argparse       # 명령줄 인자 처리를 위한 라이브러리
from openpyxl import load_workbook  # Excel 파일 처리를 위한 라이브러리

# 상수 정의
LOG_FILE = 'ChatService.log'  # 기본 로그 파일명
NEW_LOG_PATH = r'E:\devSpace\ChatServiceTest\log\ChatService.log'  # 새 로그 파일 경로

def replace_log_file():
    """
    로그 파일을 새로운 버전으로 교체하는 함수
    - 기존 로그 파일이 있으면 삭제
    - 새 로그 파일을 복사해서 가져옴
    """
    # 기존 로그 파일이 존재하면 삭제
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    
    # 새 로그 파일을 현재 디렉토리로 복사
    shutil.copy(NEW_LOG_PATH, LOG_FILE)

def parse_logs(filepath, room_number=None):
    """
    로그 파일을 파싱해서 핵심 이벤트들만 추출하는 함수
    
    🔧 진짜 임계구역 3개 이벤트만 파싱:
    - PRE_JOIN_CURRENT_STATE: 진짜 임계구역 시작
    - JOIN_SUCCESS_EXISTING: 진짜 임계구역 끝 (성공)
    - JOIN_FAIL_OVER_CAPACITY_EXISTING: 진짜 임계구역 끝 (실패)
    
    입력:
    - filepath: 로그 파일 경로
    - room_number: 특정 방 번호만 필터링 (None이면 모든 방)
    
    출력: DataFrame - 파싱된 이벤트 데이터
    """
    
    # 정규 표현식 패턴 정의 (3개 핵심 이벤트만)
    pattern = re.compile(
        r'timestampIso=(?P<timestamp>\S+).*?'  # 시간 정보 추출
        r'event=(?P<event>PRE_JOIN_CURRENT_STATE|JOIN_SUCCESS_EXISTING|JOIN_FAIL_OVER_CAPACITY_EXISTING).*?'  # 핵심 이벤트만
        r'roomNumber=(?P<roomNumber>\d+).*?'    # 방 번호
        r'userId=(?P<userId>\S+).*?'            # 사용자 ID
        r'currentPeople=(?P<currentPeople>\d+).*?'  # 현재 인원수
        r'maxPeople=(?P<maxPeople>\d+)'         # 최대 정원
    )
    
    records = []  # 파싱된 레코드들을 저장할 리스트
    
    print("🔍 디버깅: 로그 파싱 시작")
    line_count = 0
    
    # 로그 파일을 한 줄씩 읽으면서 파싱
    with open(filepath, encoding='utf-8') as f:
        for line in f:
            line_count += 1
            
            # 정규식으로 매칭 시도
            match = pattern.search(line)
            if match:
                # 매칭된 그룹들을 딕셔너리로 변환
                data = match.groupdict()
                
                # 문자열로 추출된 숫자들을 정수형으로 변환
                data['roomNumber'] = int(data['roomNumber'])
                data['currentPeople'] = int(data['currentPeople'])
                data['maxPeople'] = int(data['maxPeople'])
                
                # 🔧 나노초 정밀도 정보 추출 (정렬 및 분석용)
                nano_match = re.search(r'nanoTime=(\d+)', line)
                
                if nano_match:
                    extracted_nano = int(nano_match.group(1))
                    data['nanoTime'] = extracted_nano
                    
                    # 🔍 디버깅: 추출된 nanoTime 값 출력
                    print(f"   라인 {line_count}: {data['event']} - {data['userId']} - nanoTime: {extracted_nano}")
                
                # 방 번호 필터링 적용
                if room_number is None or data['roomNumber'] == room_number:
                    records.append(data)
    
    print(f"🔍 디버깅: 총 {len(records)}개 이벤트 파싱 완료")
    
    # 리스트를 DataFrame으로 변환해서 반환
    return pd.DataFrame(records)

def normalize_timestamp_format(df_result):
    """
    시간 형식을 기존 형식으로 통일하는 함수
    2025-07-09T13:36:41.721432200Z → 2025-07-09 13:36:41.721432200+00:00
    """
    print("🔍 디버깅: 시간 형식 정규화 시작")
    
    if 'prev_entry_time' in df_result.columns:
        df_result['prev_entry_time'] = pd.to_datetime(df_result['prev_entry_time'])
        df_result['prev_entry_time'] = df_result['prev_entry_time'].dt.strftime('%Y-%m-%d %H:%M:%S.%f+00:00')
    
    if 'curr_entry_time' in df_result.columns:
        df_result['curr_entry_time'] = pd.to_datetime(df_result['curr_entry_time'])
        df_result['curr_entry_time'] = df_result['curr_entry_time'].dt.strftime('%Y-%m-%d %H:%M:%S.%f+00:00')
    
    print("🔍 디버깅: 시간 형식 정규화 완료")
    return df_result

def build_paired_data_true_critical_section(df):
    """
    🔧 시간순 단순 매칭 기반 페어링 로직 함수
    - pair_idx 제거하고 시간순 단순 매칭 적용
    - 나노초 정밀도 정렬 및 bin 할당 적용
    
    입력: df (DataFrame) - 파싱된 로그 데이터
    출력: DataFrame - 페어링된 입장 요청 데이터
    """
    
    print("🔍 디버깅: 페어링 시작")
    
    # 빈 데이터면 빈 DataFrame 반환
    if df.empty:
        return pd.DataFrame()
    
    # 입력 데이터 디버깅
    print(f"🔍 디버깅: 입력 데이터 {len(df)}건")
    if 'nanoTime' in df.columns:
        nano_values = df['nanoTime'].dropna().values
        print(f"🔍 디버깅: 입력 nanoTime 값 범위: {nano_values.min()} ~ {nano_values.max()}")
        print(f"🔍 디버깅: 입력 nanoTime 샘플: {nano_values[:5]}")
    
    # === 전체 이벤트를 nanoTime 기준으로 시간순 정렬 ===
    if 'nanoTime' in df.columns:
        print("   nanoTime 기준으로 전체 이벤트 정렬 중...")
        df_sorted = df.sort_values(['roomNumber', 'nanoTime']).reset_index(drop=True)
        
        # 🔍 디버깅: 정렬 후 nanoTime 값 확인
        sorted_nano_values = df_sorted['nanoTime'].dropna().values
        print(f"🔍 디버깅: 정렬 후 nanoTime 값 범위: {sorted_nano_values.min()} ~ {sorted_nano_values.max()}")
        print(f"🔍 디버깅: 정렬 후 nanoTime 샘플: {sorted_nano_values[:5]}")
        
    else:
        print("   timestamp 기준으로 전체 이벤트 정렬 중...")
        df_sorted = df.sort_values(['roomNumber', 'timestamp']).reset_index(drop=True)
    
    # === 방별로 시간순 단순 매칭 수행 ===
    result_list = []
    
    for room_num in df_sorted['roomNumber'].unique():
        print(f"   방 {room_num} 페어링 중...")
        room_df = df_sorted[df_sorted['roomNumber'] == room_num].copy()
        
        # 방별 이벤트 시간순 매칭
        i = 0
        while i < len(room_df):
            current_row = room_df.iloc[i]
            
            # PRE_JOIN_CURRENT_STATE 이벤트 찾기
            if current_row['event'] == 'PRE_JOIN_CURRENT_STATE':
                pre_event = current_row
                
                # 🔍 디버깅: PRE_JOIN 이벤트 nanoTime 확인
                pre_nano = pre_event.get('nanoTime', 'None')
                print(f"     PRE_JOIN - {pre_event['userId']} - nanoTime: {pre_nano}")
                
                # 다음 SUCCESS 또는 FAIL 이벤트 찾기
                for j in range(i + 1, len(room_df)):
                    next_row = room_df.iloc[j]
                    
                    # 같은 사용자의 SUCCESS/FAIL 이벤트 찾음
                    if (next_row['userId'] == pre_event['userId'] and 
                        next_row['event'] in ['JOIN_SUCCESS_EXISTING', 'JOIN_FAIL_OVER_CAPACITY_EXISTING']):
                        
                        # 🔍 디버깅: SUCCESS/FAIL 이벤트 nanoTime 확인
                        end_nano = next_row.get('nanoTime', 'None')
                        print(f"     {next_row['event']} - {next_row['userId']} - nanoTime: {end_nano}")
                        
                        # 페어링 완성
                        paired_record = create_paired_record(pre_event, next_row)
                        
                        # 🔍 디버깅: 페어링 결과 nanoTime 확인
                        paired_nano_pre = paired_record.get('nanoTime_pre', 'None')
                        paired_nano_end = paired_record.get('nanoTime_end', 'None')
                        print(f"     페어링 결과 - nanoTime_pre: {paired_nano_pre}, nanoTime_end: {paired_nano_end}")
                        
                        result_list.append(paired_record)
                        break
            
            i += 1
    
    if not result_list:
        return pd.DataFrame()
    
    # === 결과 DataFrame 생성 ===
    result = pd.DataFrame(result_list)
    
    # 🔍 디버깅: 결과 DataFrame nanoTime 값 확인
    print(f"🔍 디버깅: 결과 DataFrame 생성 완료 - {len(result)}건")
    if 'nanoTime_pre' in result.columns:
        nano_pre_values = result['nanoTime_pre'].dropna().values
        print(f"🔍 디버깅: nanoTime_pre 값 범위: {nano_pre_values.min()} ~ {nano_pre_values.max()}")
        print(f"🔍 디버깅: nanoTime_pre 샘플: {nano_pre_values[:5]}")
    
    if 'nanoTime_end' in result.columns:
        nano_end_values = result['nanoTime_end'].dropna().values
        print(f"🔍 디버깅: nanoTime_end 값 범위: {nano_end_values.min()} ~ {nano_end_values.max()}")
        print(f"🔍 디버깅: nanoTime_end 샘플: {nano_end_values[:5]}")
    
    # === 🔧 나노초 정밀도 기반 재정렬 및 순번 부여 ===
    if 'nanoTime_pre' in result.columns:
        print("   나노초 정밀도 기준으로 최종 정렬 중...")
        result = result.sort_values(['roomNumber', 'nanoTime_pre']).reset_index(drop=True)
        
        # 🔍 디버깅: 최종 정렬 후 nanoTime 값 확인
        sorted_nano_pre = result['nanoTime_pre'].dropna().values
        print(f"🔍 디버깅: 최종 정렬 후 nanoTime_pre 값 범위: {sorted_nano_pre.min()} ~ {sorted_nano_pre.max()}")
        print(f"🔍 디버깅: 최종 정렬 후 nanoTime_pre 샘플: {sorted_nano_pre[:5]}")
        
    else:
        print("   타임스탬프 기준으로 최종 정렬 중...")
        result = result.sort_values(['roomNumber', 'timestamp_pre']).reset_index(drop=True)
    
    # 방별 입장 순번 부여 (나노초 정렬 순서 기준)
    result['room_entry_sequence'] = result.groupby('roomNumber').cumcount() + 1
    
    # === 🔧 방별 개별 bin 할당 (나노초 순서 기준) ===
    print("   방별 개별 bin 할당 중...")
    result['bin'] = result.groupby('roomNumber').cumcount() // 20 + 1
    result['bin'] = result['bin'].clip(upper=10)  # 최대 10으로 제한
    
    # 🔍 디버깅: bin 할당 후 nanoTime 값 확인
    print(f"🔍 디버깅: bin 할당 후 nanoTime 값 확인")
    if 'nanoTime_pre' in result.columns:
        bin_nano_pre = result['nanoTime_pre'].dropna().values
        print(f"🔍 디버깅: bin 할당 후 nanoTime_pre 값 범위: {bin_nano_pre.min()} ~ {bin_nano_pre.max()}")
        print(f"🔍 디버깅: bin 할당 후 nanoTime_pre 샘플: {bin_nano_pre[:5]}")
    
    # === 🔧 최종 컬럼 정리 ===
    final_columns = {
        'roomNumber': 'roomNumber',
        'bin': 'bin',
        'userId': 'user_id',
        'currentPeople_pre': 'prev_people',
        'curr_people': 'curr_people',
        'expected_people': 'expected_people',
        'maxPeople_pre': 'max_people',
        'room_entry_sequence': 'room_entry_sequence',
        'join_result': 'join_result',
        'timestamp_pre': 'prev_entry_time',
        'timestamp_end': 'curr_entry_time'
    }
    
    # === 🔧 나노초 정밀도 필드 추가 ===
    if 'nanoTime_pre' in result.columns:
        final_columns['nanoTime_pre'] = 'true_critical_section_nanoTime_start'
        print("🔍 디버깅: nanoTime_pre → true_critical_section_nanoTime_start 매핑")
    if 'nanoTime_end' in result.columns:
        final_columns['nanoTime_end'] = 'true_critical_section_nanoTime_end'
        print("🔍 디버깅: nanoTime_end → true_critical_section_nanoTime_end 매핑")
    
    # === 최종 컬럼 선택 및 이름 변경 ===
    existing_columns = {old: new for old, new in final_columns.items() if old in result.columns}
    
    # 🔍 디버깅: 컬럼 매핑 전 nanoTime 값 확인
    print(f"🔍 디버깅: 컬럼 매핑 전 최종 확인")
    if 'nanoTime_pre' in result.columns:
        pre_mapping_values = result['nanoTime_pre'].dropna().values
        print(f"🔍 디버깅: 매핑 전 nanoTime_pre 값 범위: {pre_mapping_values.min()} ~ {pre_mapping_values.max()}")
        print(f"🔍 디버깅: 매핑 전 nanoTime_pre 샘플: {pre_mapping_values[:5]}")
    
    # 원하는 순서대로 컬럼 정렬
    desired_order = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 'expected_people', 'max_people', 
                     'room_entry_sequence', 'join_result', 'prev_entry_time', 'curr_entry_time',
                     'true_critical_section_nanoTime_start', 'true_critical_section_nanoTime_end']
    
    # DataFrame 재구성
    result = result[list(existing_columns.keys())].rename(columns=existing_columns)
    
    # 🔍 디버깅: 컬럼 매핑 후 nanoTime 값 확인
    print(f"🔍 디버깅: 컬럼 매핑 후 최종 확인")
    if 'true_critical_section_nanoTime_start' in result.columns:
        post_mapping_values = result['true_critical_section_nanoTime_start'].dropna().values
        print(f"🔍 디버깅: 매핑 후 true_critical_section_nanoTime_start 값 범위: {post_mapping_values.min()} ~ {post_mapping_values.max()}")
        print(f"🔍 디버깅: 매핑 후 true_critical_section_nanoTime_start 샘플: {post_mapping_values[:5]}")
    
    # 존재하는 컬럼들만 원하는 순서로 재배열
    final_order = [col for col in desired_order if col in result.columns]
    result = result[final_order]
    
    # === 🔧 시간 형식 기존 형식으로 통일 ===
    result = normalize_timestamp_format(result)
    
    # 🔍 디버깅: 최종 결과 확인
    print(f"🔍 디버깅: 최종 결과 - {len(result)}건")
    if 'true_critical_section_nanoTime_start' in result.columns:
        final_values = result['true_critical_section_nanoTime_start'].dropna().values
        print(f"🔍 디버깅: 최종 true_critical_section_nanoTime_start 값 범위: {final_values.min()} ~ {final_values.max()}")
        print(f"🔍 디버깅: 최종 true_critical_section_nanoTime_start 샘플: {final_values[:5]}")
    
    return result

def create_paired_record(pre_event, end_event):
    """
    PRE_JOIN_CURRENT_STATE와 SUCCESS/FAIL 이벤트를 매칭해서 하나의 레코드 생성
    """
    record = {}
    
    # 기본 정보
    record['roomNumber'] = pre_event['roomNumber']
    record['userId'] = pre_event['userId']
    record['maxPeople_pre'] = pre_event['maxPeople']
    
    # 인원수 정보
    record['currentPeople_pre'] = pre_event['currentPeople']
    record['curr_people'] = end_event['currentPeople']
    
    # 성공/실패에 따른 처리
    if end_event['event'] == 'JOIN_SUCCESS_EXISTING':
        record['join_result'] = 'SUCCESS'
        record['expected_people'] = pre_event['currentPeople'] + 1
    else:  # JOIN_FAIL_OVER_CAPACITY_EXISTING
        record['join_result'] = 'FAIL_OVER_CAPACITY'
        record['expected_people'] = None
    
    # 시간 정보
    record['timestamp_pre'] = pre_event['timestamp']
    record['timestamp_end'] = end_event['timestamp']
    
    # 나노초 정보
    if 'nanoTime' in pre_event and pd.notna(pre_event['nanoTime']):
        record['nanoTime_pre'] = pre_event['nanoTime']
        print(f"     create_paired_record - nanoTime_pre 설정: {pre_event['nanoTime']}")
    if 'nanoTime' in end_event and pd.notna(end_event['nanoTime']):
        record['nanoTime_end'] = end_event['nanoTime']
        print(f"     create_paired_record - nanoTime_end 설정: {end_event['nanoTime']}")
    
    return record

def get_true_critical_section_desc_table():
    """
    🔧 핵심 이벤트 나노초 포함 설명 테이블 생성 함수
    """
    return [
        ["속성명", "분석 목적", "도출 방법"],
        ["roomNumber", "방 번호 식별", "로그 필드: roomNumber"],
        ["bin", "분석 구간 구분", "각 방별로 나노초 순서 기준 20개씩 10구간"],
        ["user_id", "사용자 식별", "로그 필드: userId"],
        ["prev_people", "입장 전 인원수", "PRE_JOIN_CURRENT_STATE의 currentPeople"],
        ["curr_people", "입장 후 인원수", "SUCCESS/FAIL 이벤트의 currentPeople"],
        ["expected_people", "기대 인원수", "prev_people + 1 (성공시)"],
        ["max_people", "최대 정원", "로그 필드: maxPeople"],
        ["room_entry_sequence", "방별 입장 순번", "나노초 정밀도 기준 방별 순번"],
        ["join_result", "입장 결과", "SUCCESS 또는 FAIL_OVER_CAPACITY"],
        ["prev_entry_time", "임계구역 시작 시간", "PRE_JOIN_CURRENT_STATE 타임스탬프"],
        ["curr_entry_time", "임계구역 종료 시간", "SUCCESS/FAIL 타임스탬프"],
        ["true_critical_section_nanoTime_start", "임계구역 시작 나노초", "PRE_JOIN_CURRENT_STATE의 nanoTime"],
        ["true_critical_section_nanoTime_end", "임계구역 끝 나노초", "SUCCESS/FAIL의 nanoTime"]
    ]

def save_with_side_table(df_result, out_xlsx, desc_table):
    """
    Excel 파일에 데이터와 설명 테이블을 함께 저장하는 함수
    """
    df_result.to_excel(out_xlsx, index=False)
    
    wb = load_workbook(out_xlsx)
    ws = wb.active
    
    start_col = len(df_result.columns) + 2
    
    for i, row in enumerate(desc_table):
        for j, val in enumerate(row):
            ws.cell(row=i + 1, column=start_col + j, value=val)
    
    wb.save(out_xlsx)

def analyze_results(df):
    """
    처리된 데이터에 대한 간단한 통계 분석을 수행하는 함수
    """
    if df.empty:
        print("분석할 데이터가 없습니다.")
        return
    
    # 기본 통계 계산
    total_requests = len(df)
    success_count = len(df[df['join_result'] == 'SUCCESS'])
    fail_count = len(df[df['join_result'] == 'FAIL_OVER_CAPACITY'])
    
    print(f"\n=== 🔧 나노초 정밀도 기반 전처리 결과 ===")
    print(f"전체 입장 요청 수: {total_requests}")
    print(f"성공: {success_count}건 ({success_count/total_requests*100:.1f}%)")
    print(f"실패: {fail_count}건 ({fail_count/total_requests*100:.1f}%)")
    
    # 나노초 정밀도 통계
    if 'true_critical_section_nanoTime_start' in df.columns:
        nano_count = df['true_critical_section_nanoTime_start'].notna().sum()
        print(f"나노초 정밀도 데이터: {nano_count}건")
    
    # 레이스 컨디션 예비 분석
    if success_count > 0:
        valid_success = df[(df['join_result'] == 'SUCCESS') & df['expected_people'].notna()]
        if not valid_success.empty:
            race_conditions = valid_success[valid_success['curr_people'] != valid_success['expected_people']]
            race_count = len(race_conditions)
            print(f"잠재적 레이스 컨디션: {race_count}건 ({race_count/success_count*100:.1f}%)")
    
    # 방별 bin 분포 확인
    print(f"\n=== 방별 bin 분포 ===")
    for room_num in sorted(df['roomNumber'].unique()):
        room_data = df[df['roomNumber'] == room_num]
        bin_counts = room_data['bin'].value_counts().sort_index()
        print(f"방 {room_num}: bin {list(bin_counts.index)} → {list(bin_counts.values)}건")
    
    # 방별 통계
    room_stats = df.groupby('roomNumber').agg({
        'user_id': 'count',
        'join_result': lambda x: (x == 'SUCCESS').sum()
    }).rename(columns={'user_id': 'total_requests', 'join_result': 'success_count'})
    
    print(f"\n=== 방별 통계 ===")
    for room_num, stats in room_stats.iterrows():
        success_rate = stats['success_count'] / stats['total_requests'] * 100
        print(f"방 {room_num}: 총 {stats['total_requests']}건, 성공 {stats['success_count']}건 ({success_rate:.1f}%)")

def main():
    """
    🔧 디버깅 출력 추가된 메인 함수
    """
    parser = argparse.ArgumentParser(description="Race Condition 이벤트 전처리기 (디버깅 버전)")
    parser.add_argument('--room', type=int, help='특정 방 번호만 처리 (옵션)')
    parser.add_argument('--csv', type=str, help='CSV 파일명 (필수)')
    parser.add_argument('--xlsx', type=str, help='Excel 파일명 (옵션)')
    parser.add_argument('--output-dir', type=str, help='출력 파일 저장 디렉토리 (옵션)')
    
    args = parser.parse_args()
    
    # CSV 또는 XLSX 옵션이 없으면 종료
    if not args.csv and not args.xlsx:
        print("❌ 파일 저장 옵션을 지정해주세요:")
        print("   --csv 파일명.csv")
        print("   --xlsx 파일명.xlsx")
        print("   또는 둘 다 지정 가능")
        return
    
    try:
        print("🔧 디버깅 버전 Race Condition 이벤트 전처리기 시작...")
        print("📋 시간순 단순 매칭 및 방별 개별 bin 할당 적용")
        print("🕐 시간 형식 기존 형식으로 통일")
        print("🎯 3개 핵심 이벤트만 처리")
        print("🔍 디버깅 출력 활성화")
        
        # 출력 디렉토리 생성
        if args.output_dir:
            print(f"📁 출력 디렉토리 설정: {args.output_dir}")
            try:
                os.makedirs(args.output_dir, exist_ok=True)
                print("   출력 디렉토리 생성 완료")
            except OSError as e:
                print(f"❌ 출력 디렉토리 생성 실패: {e}")
                return
        
        # 1단계: 로그 파일 교체
        print("1. 로그 파일 교체 중...")
        replace_log_file()
        
        # 2단계: 로그 파싱 (3개 핵심 이벤트만)
        print("2. 핵심 3개 이벤트 파싱 중...")
        df = parse_logs(LOG_FILE, room_number=args.room)
        print(f"   파싱된 이벤트 수: {len(df)}")
        
        # 3단계: 시간순 단순 매칭 기반 페어링
        print("3. 시간순 단순 매칭 페어링 처리 중...")
        result = build_paired_data_true_critical_section(df)
        print(f"   페어링된 요청 수: {len(result)}")
        
        # 4단계: 결과 저장
        print("4. 결과 저장 중...")
        
        if args.csv:
            csv_path = os.path.join(args.output_dir, args.csv) if args.output_dir else args.csv
            result.to_csv(csv_path, index=False, encoding='utf-8-sig')
            print(f"   CSV 저장 완료: {csv_path}")
        
        if args.xlsx:
            xlsx_path = os.path.join(args.output_dir, args.xlsx) if args.output_dir else args.xlsx
            desc_table = get_true_critical_section_desc_table()
            save_with_side_table(result, xlsx_path, desc_table)
            print(f"   Excel 저장 완료: {xlsx_path}")
        
        # 5단계: 결과 분석
        print("5. 결과 분석 중...")
        analyze_results(result)
        
        print("\n✅ 디버깅 버전 전처리 완료!")
        print("🎯 3개 핵심 이벤트 나노초 데이터 포함!")
        print("🕐 시간 형식 기존 형식으로 통일 완료!")
        print("🔧 시간순 단순 매칭 적용 완료!")
        print("📁 출력 디렉토리 지정 기능 추가 완료!")
        print("🔍 디버깅 출력으로 nanoTime 값 추적 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()