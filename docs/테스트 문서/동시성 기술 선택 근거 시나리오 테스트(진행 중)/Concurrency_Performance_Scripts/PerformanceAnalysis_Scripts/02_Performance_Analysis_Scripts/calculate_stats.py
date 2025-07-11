#!/usr/bin/env python3
"""
성능 분석 지표 계산 스크립트 (v6.0) - 나노초 단위 분석
- 모든 시간 계산을 나노초 단위로 유지
- 통계 출력도 나노초 단위로 표시

[스크립트 목적]
이 스크립트는 성능 테스트 결과 CSV 파일을 읽어서 다양한 통계를 계산하고,
그 결과를 Excel 파일로 저장합니다.

[주요 기능]
1. CSV 파일에서 성능 데이터 읽기
2. 성공/실패 케이스별 분류
3. 대기 시간, 처리 시간 등의 통계 계산
4. 방(room)별, 구간(bin)별 상세 통계 생성
5. 결과를 Excel 파일로 저장
"""

# 필요한 라이브러리들을 가져옵니다 (import)
import pandas as pd      # 데이터 분석을 위한 라이브러리
import numpy as np       # 수치 계산을 위한 라이브러리
import os               # 파일/폴더 작업을 위한 라이브러리
import argparse         # 명령줄 인자 처리를 위한 라이브러리
from datetime import datetime  # 날짜/시간 처리를 위한 라이브러리
import sys              # 시스템 관련 기능을 위한 라이브러리
from openpyxl import load_workbook  # Excel 파일 편집을 위한 라이브러리
from decimal import Decimal  # 정확한 숫자 계산을 위한 라이브러리


def calculate_rate(count, total):
    """
    전체 대비 비율을 계산하는 함수
    
    매개변수:
        count: 특정 항목의 개수
        total: 전체 개수
    
    반환값:
        백분율 (0~100 사이의 숫자)
    
    예시:
        calculate_rate(30, 100) → 30.0 (30%)
    """
    # pandas Series인 경우와 일반 숫자인 경우를 구분해서 처리
    if isinstance(count, pd.Series) or isinstance(total, pd.Series):
        # numpy의 where 함수: 조건에 따라 다른 값을 반환
        # total이 0이 아니면 비율 계산, 0이면 0 반환 (0으로 나누기 방지)
        return np.where(total != 0, (count / total) * 100, 0)
    
    # 일반 숫자인 경우
    return (count / total) * 100 if total > 0 else 0


def get_stats_with_precision(series, metric_name="", unit="ns"):
    """
    데이터 시리즈의 통계값을 계산하는 함수
    
    매개변수:
        series: 통계를 계산할 데이터 시리즈 (pandas Series)
        metric_name: 출력할 메트릭 이름 (예: "Wait Time")
        unit: 시간 단위 ("ns"=나노초, "us"=마이크로초, "ms"=밀리초)
    
    반환값:
        통계값 딕셔너리 {'Mean': 평균, 'Median': 중앙값, 'Max': 최댓값}
    """
    # 데이터가 없으면 0으로 채운 결과 반환
    if len(series) == 0:
        return {'Mean': 0.0, 'Median': 0.0, 'Max': 0.0}
    
    # NaN(Not a Number, 값이 없음)이 아닌 유효한 값들만 추출
    valid_series = series.dropna()
    if len(valid_series) == 0:
        return {'Mean': 0.0, 'Median': 0.0, 'Max': 0.0}
    
    # 단위 변환 계수 설정 (나노초를 기준으로)
    # 1나노초 = 0.001마이크로초 = 0.000001밀리초
    conversion = {
        'ns': 1.0,          # 나노초 (기본 단위)
        'us': 0.001,        # 마이크로초로 변환
        'ms': 0.000001      # 밀리초로 변환
    }
    factor = conversion.get(unit, 1.0)
    
    # 기본 통계값 계산
    mean_val = float(valid_series.mean())    # 평균
    median_val = float(valid_series.median()) # 중앙값
    max_val = float(valid_series.max())      # 최댓값
    
    # 계산된 통계를 지정된 단위로 변환하여 반환
    return {
        'Mean': mean_val * factor,
        'Median': median_val * factor,
        'Max': max_val * factor
    }


def parse_nano_time_precise(nano_str):
    """
    나노초 문자열을 정수로 변환하는 함수
    
    매개변수:
        nano_str: 나노초를 나타내는 문자열 (예: "1234567890")
    
    반환값:
        정수형 나노초 값 또는 NaN (변환 실패시)
    
    설명:
        CSV 파일에서 읽은 나노초 값은 문자열이므로,
        이를 정수로 변환해야 계산이 가능합니다.
    """
    # 값이 없거나 'nan' 문자열인 경우 NaN 반환
    if pd.isna(nano_str) or nano_str == 'nan':
        return np.nan
    
    try:
        if isinstance(nano_str, str):
            # 과학적 표기법 (예: 1.23E+9) 처리
            if 'E' in nano_str or 'e' in nano_str:
                decimal_val = Decimal(nano_str)
                return int(decimal_val)
            else:
                # 일반 숫자 문자열을 정수로 변환
                return int(float(nano_str))
        else:
            # 이미 숫자인 경우 정수로 변환
            return int(nano_str)
    except (ValueError, TypeError):
        # 변환 실패시 NaN 반환
        return np.nan


def calculate_time_diff_nano(start_nano, end_nano):
    """
    두 나노초 시간값의 차이를 계산하는 함수
    
    매개변수:
        start_nano: 시작 시간 (나노초)
        end_nano: 종료 시간 (나노초)
    
    반환값:
        시간 차이 (나노초) 또는 NaN
    
    예시:
        calculate_time_diff_nano(1000, 3000) → 2000
    """
    # 시작 또는 종료 시간이 없으면 NaN 반환
    if pd.isna(start_nano) or pd.isna(end_nano):
        return np.nan
    
    try:
        # 종료 시간에서 시작 시간을 빼서 경과 시간 계산
        diff_nano = int(end_nano) - int(start_nano)
        return diff_nano
    except (ValueError, TypeError):
        return np.nan


def set_join_result_from_events(df):
    """
    이벤트 타입을 기반으로 join_result 컬럼을 설정하는 함수
    
    매개변수:
        df: 성능 데이터가 담긴 DataFrame
    
    반환값:
        join_result 컬럼이 설정된 DataFrame
    
    설명:
        각 요청의 결과를 분류합니다:
        - SUCCESS: 성공적으로 처리됨
        - FAIL_OVER_CAPACITY: 정원 초과로 실패
        - FAIL_ENTRY: 진입 자체를 실패 (락 획득 실패 등)
    """
    def determine_join_result(row):
        """개별 행(row)의 join_result를 결정하는 내부 함수"""
        # 이미 join_result가 설정되어 있고 유효한 값이면 그대로 사용
        if pd.notna(row.get('join_result')) and row.get('join_result') != '':
            existing_result = str(row['join_result']).strip()
            if existing_result in ['SUCCESS', 'FAIL_OVER_CAPACITY', 'FAIL_ENTRY']:
                return existing_result
        
        # critical_leave_event_type을 기반으로 결과 판단
        if pd.isna(row.get('critical_leave_event_type')) or row.get('critical_leave_event_type') == '':
            return 'FAIL_ENTRY'
        elif str(row['critical_leave_event_type']).strip() == 'SUCCESS':
            return 'SUCCESS'
        elif str(row['critical_leave_event_type']).strip() == 'FAIL_OVER_CAPACITY':
            return 'FAIL_OVER_CAPACITY'
        else:
            return 'UNKNOWN'
    
    # DataFrame 복사본 생성 (원본 데이터 보호)
    df_result = df.copy()
    
    # 각 행에 대해 join_result 결정 함수 적용
    df_result['join_result'] = df_result.apply(determine_join_result, axis=1)
    
    return df_result


def create_summary_stats(df_total, df_success, df_lock_failed, df_capacity_failed):
    """
    전체 요약 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: 성공한 요청들
        df_lock_failed: 진입 실패한 요청들
        df_capacity_failed: 정원 초과로 실패한 요청들
    
    반환값:
        요약 통계 DataFrame
    """
    total_requests = len(df_total)
    success_count = len(df_success)
    lock_failed_count = len(df_lock_failed)
    capacity_failed_count = len(df_capacity_failed)
    
    summary_data = {
        'Category': [
            'Total Requests',
            'Success',
            'Entry Failed (Lock, etc)',
            'Capacity Failed'
        ],
        'Count': [
            total_requests,
            success_count,
            lock_failed_count,
            capacity_failed_count
        ],
        'Percentage (%)': [
            100.0,
            calculate_rate(success_count, total_requests),
            calculate_rate(lock_failed_count, total_requests),
            calculate_rate(capacity_failed_count, total_requests)
        ]
    }
    
    return pd.DataFrame(summary_data)


def create_success_stats(df_success):
    """
    성공한 요청들의 시간 통계를 생성하는 함수
    
    매개변수:
        df_success: 성공한 요청들의 DataFrame
    
    반환값:
        성공 통계 DataFrame
    """
    if len(df_success) == 0:
        return pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # 대기 시간과 처리 시간 통계 계산
    wait_time_stats = get_stats_with_precision(df_success['wait_time_ns'], "Wait Time", "ns")
    dwell_time_stats = get_stats_with_precision(df_success['dwell_time_ns'], "Dwell Time", "ns")
    
    # DataFrame 형태로 구성
    success_stats_data = {
        'Metric': ['Wait Time'] * 3 + ['Dwell Time (Critical Section)'] * 3,
        'Statistic': ['Mean', 'Median', 'Max'] * 2,
        'Value': [
            wait_time_stats['Mean'], 
            wait_time_stats['Median'], 
            wait_time_stats['Max'],
            dwell_time_stats['Mean'], 
            dwell_time_stats['Median'], 
            dwell_time_stats['Max']
        ],
        'Unit': ['ns'] * 6
    }
    
    return pd.DataFrame(success_stats_data)


def create_capacity_failed_stats(df_capacity_failed):
    """
    정원 초과로 실패한 요청들의 시간 통계를 생성하는 함수
    
    매개변수:
        df_capacity_failed: 정원 초과 실패 요청들의 DataFrame
    
    반환값:
        정원 초과 실패 통계 DataFrame
    """
    if len(df_capacity_failed) == 0:
        return pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # 대기 시간과 실패 처리 시간 통계 계산
    capacity_wait_stats = get_stats_with_precision(
        df_capacity_failed['wait_time_ns'], 
        "Wait Time (Capacity Failed)", 
        "ns"
    )
    fail_proc_stats = get_stats_with_precision(
        df_capacity_failed['fail_processing_time_ns'], 
        "Fail Processing Time", 
        "ns"
    )
    
    # DataFrame 형태로 구성
    capacity_failed_stats_data = {
        'Metric': ['Wait Time'] * 3 + ['Fail Processing Time'] * 3,
        'Statistic': ['Mean', 'Median', 'Max'] * 2,
        'Value': [
            capacity_wait_stats['Mean'], 
            capacity_wait_stats['Median'], 
            capacity_wait_stats['Max'],
            fail_proc_stats['Mean'], 
            fail_proc_stats['Median'], 
            fail_proc_stats['Max']
        ],
        'Unit': ['ns'] * 6
    }
    
    return pd.DataFrame(capacity_failed_stats_data)


def create_per_room_stats(df_total, df_success, df_capacity_failed, df_lock_failed):
    """
    방(room)별 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: 성공한 요청들
        df_capacity_failed: 정원 초과 실패 요청들
        df_lock_failed: 진입 실패 요청들
    
    반환값:
        방별 통계 DataFrame
    """
    # 모든 방 번호를 정렬된 순서로 가져오기
    all_rooms = sorted(df_total['roomNumber'].unique())
    
    room_stats_list = []
    
    # 각 방별로 통계 계산
    for room in all_rooms:
        # 해당 방의 데이터만 필터링
        room_total = df_total[df_total['roomNumber'] == room]
        room_success = df_success[df_success['roomNumber'] == room] if len(df_success) > 0 else pd.DataFrame()
        room_capacity_failed = df_capacity_failed[df_capacity_failed['roomNumber'] == room] if len(df_capacity_failed) > 0 else pd.DataFrame()
        room_lock_failed = df_lock_failed[df_lock_failed['roomNumber'] == room] if len(df_lock_failed) > 0 else pd.DataFrame()
        
        # 통계 계산
        stats = {
            'roomNumber': room,
            'total_requests': len(room_total),
            'success_count': len(room_success),
            'capacity_failed_count': len(room_capacity_failed),
            'entry_failed_count': len(room_lock_failed),
            'success_rate(%)': calculate_rate(len(room_success), len(room_total)),
            'capacity_failed_rate(%)': calculate_rate(len(room_capacity_failed), len(room_total)),
            'entry_failed_rate(%)': calculate_rate(len(room_lock_failed), len(room_total)),
            'success_avg_wait_time(ns)': room_success['wait_time_ns'].mean() if len(room_success) > 0 else 0,
            'success_median_wait_time(ns)': room_success['wait_time_ns'].median() if len(room_success) > 0 else 0,
            'success_max_wait_time(ns)': room_success['wait_time_ns'].max() if len(room_success) > 0 else 0,
            'success_avg_dwell_time(ns)': room_success['dwell_time_ns'].mean() if len(room_success) > 0 else 0,
            'success_median_dwell_time(ns)': room_success['dwell_time_ns'].median() if len(room_success) > 0 else 0,
            'success_max_dwell_time(ns)': room_success['dwell_time_ns'].max() if len(room_success) > 0 else 0,
            'capacity_failed_avg_wait_time(ns)': room_capacity_failed['wait_time_ns'].mean() if len(room_capacity_failed) > 0 else 0,
            'capacity_failed_max_wait_time(ns)': room_capacity_failed['wait_time_ns'].max() if len(room_capacity_failed) > 0 else 0,
            'capacity_failed_avg_fail_processing_time(ns)': room_capacity_failed['fail_processing_time_ns'].mean() if len(room_capacity_failed) > 0 else 0,
            'capacity_failed_max_fail_processing_time(ns)': room_capacity_failed['fail_processing_time_ns'].max() if len(room_capacity_failed) > 0 else 0
        }
        
        room_stats_list.append(stats)
    
    return pd.DataFrame(room_stats_list)


def create_per_bin_stats(df_total, df_success, df_capacity_failed, df_lock_failed):
    """
    방(room)과 구간(bin)별 상세 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: 성공한 요청들
        df_capacity_failed: 정원 초과 실패 요청들
        df_lock_failed: 진입 실패 요청들
    
    반환값:
        방-구간별 통계 DataFrame
    """
    # 'bin' 컬럼이 없으면 빈 DataFrame 반환
    if 'bin' not in df_total.columns:
        return pd.DataFrame()
    
    # 모든 방-구간 조합 가져오기
    all_combinations = df_total.groupby(['roomNumber', 'bin']).size().index.tolist()
    
    bin_stats_list = []
    
    # 각 방-구간 조합별로 통계 계산
    for room, bin_num in all_combinations:
        # 해당 방-구간 조합의 데이터만 필터링
        combo_total = df_total[(df_total['roomNumber'] == room) & (df_total['bin'] == bin_num)]
        combo_success = df_success[(df_success['roomNumber'] == room) & (df_success['bin'] == bin_num)] if len(df_success) > 0 else pd.DataFrame()
        combo_capacity_failed = df_capacity_failed[(df_capacity_failed['roomNumber'] == room) & (df_capacity_failed['bin'] == bin_num)] if len(df_capacity_failed) > 0 else pd.DataFrame()
        combo_lock_failed = df_lock_failed[(df_lock_failed['roomNumber'] == room) & (df_lock_failed['bin'] == bin_num)] if len(df_lock_failed) > 0 else pd.DataFrame()
        
        # 통계 계산
        stats = {
            'roomNumber': room,
            'bin': bin_num,
            'total_requests': len(combo_total),
            'success_count': len(combo_success),
            'capacity_failed_count': len(combo_capacity_failed),
            'entry_failed_count': len(combo_lock_failed),
            'success_rate(%)': calculate_rate(len(combo_success), len(combo_total)),
            'capacity_failed_rate(%)': calculate_rate(len(combo_capacity_failed), len(combo_total)),
            'entry_failed_rate(%)': calculate_rate(len(combo_lock_failed), len(combo_total)),
            'success_avg_wait_time(ns)': combo_success['wait_time_ns'].mean() if len(combo_success) > 0 else 0,
            'success_median_wait_time(ns)': combo_success['wait_time_ns'].median() if len(combo_success) > 0 else 0,
            'success_max_wait_time(ns)': combo_success['wait_time_ns'].max() if len(combo_success) > 0 else 0,
            'success_avg_dwell_time(ns)': combo_success['dwell_time_ns'].mean() if len(combo_success) > 0 else 0,
            'success_median_dwell_time(ns)': combo_success['dwell_time_ns'].median() if len(combo_success) > 0 else 0,
            'success_max_dwell_time(ns)': combo_success['dwell_time_ns'].max() if len(combo_success) > 0 else 0,
            'capacity_failed_avg_wait_time(ns)': combo_capacity_failed['wait_time_ns'].mean() if len(combo_capacity_failed) > 0 else 0,
            'capacity_failed_max_wait_time(ns)': combo_capacity_failed['wait_time_ns'].max() if len(combo_capacity_failed) > 0 else 0,
            'capacity_failed_avg_fail_processing_time(ns)': combo_capacity_failed['fail_processing_time_ns'].mean() if len(combo_capacity_failed) > 0 else 0,
            'capacity_failed_max_fail_processing_time(ns)': combo_capacity_failed['fail_processing_time_ns'].max() if len(combo_capacity_failed) > 0 else 0
        }
        
        bin_stats_list.append(stats)
    
    return pd.DataFrame(bin_stats_list)


def create_per_thread_critical_details(df_total, df_success, df_capacity_failed, df_lock_failed):
    """
    각 스레드별 임계구역 접근 상세 내역을 생성하는 함수 - 새로 추가
    
    매개변수:
        df_total: 전체 데이터
        df_success: 성공한 요청들
        df_capacity_failed: 정원 초과 실패 요청들
        df_lock_failed: 진입 실패 요청들
    
    반환값:
        각 스레드별 임계구역 접근 상세 DataFrame
    """
    # 'bin' 컬럼이 없으면 빈 DataFrame 반환
    if 'bin' not in df_total.columns:
        return pd.DataFrame()
    
    thread_details_list = []
    
    # 전체 데이터를 순회하면서 각 스레드별 상세 정보 생성
    for index, row in df_total.iterrows():
        # 기본 정보
        room_number = row['roomNumber']
        bin_number = row['bin']
        
        # 사용자/스레드 ID 찾기
        thread_id = None
        possible_id_columns = ['userId', 'user_id', 'threadId', 'thread_id', 'clientId', 'client_id', 'requestId', 'request_id']
        for col in possible_id_columns:
            if col in row and pd.notna(row[col]):
                thread_id = row[col]
                break
        
        if thread_id is None:
            thread_id = f"thread_{index}"  # 기본값
        
        # 임계구역 접근 결과 판단
        join_result = row.get('join_result', 'UNKNOWN')
        critical_leave_event = row.get('critical_leave_event_type', 'UNKNOWN')
        
        # 임계구역 진입 여부
        entered_critical_section = pd.notna(row.get('critical_enter_nanoTime'))
        exited_critical_section = pd.notna(row.get('critical_leave_nanoTime'))
        
        # 시간 계산
        wait_time_ns = np.nan
        dwell_time_ns = np.nan
        fail_processing_time_ns = np.nan
        
        # 대기 시간 계산 (진입한 경우)
        if entered_critical_section and pd.notna(row.get('waiting_start_nanoTime')):
            wait_time_ns = calculate_time_diff_nano(
                row['waiting_start_nanoTime'], 
                row['critical_enter_nanoTime']
            )
        
        # 작업 시간 또는 실패 처리 시간 계산
        if entered_critical_section and exited_critical_section:
            processing_time = calculate_time_diff_nano(
                row['critical_enter_nanoTime'], 
                row['critical_leave_nanoTime']
            )
            
            if join_result == 'SUCCESS':
                dwell_time_ns = processing_time
            elif join_result == 'FAIL_OVER_CAPACITY':
                fail_processing_time_ns = processing_time
        
        # 임계구역 작업 성공 여부
        critical_section_success = (join_result == 'SUCCESS')
        critical_section_failure_reason = None
        
        if not critical_section_success:
            if join_result == 'FAIL_OVER_CAPACITY':
                critical_section_failure_reason = 'CAPACITY_EXCEEDED'
            elif join_result == 'FAIL_ENTRY':
                critical_section_failure_reason = 'ENTRY_FAILED'
            else:
                critical_section_failure_reason = 'UNKNOWN_FAILURE'
        
        # 현재 인원수와 최대 인원수 정보 (가능한 경우)
        current_people = row.get('currentPeople', np.nan)
        max_people = row.get('maxPeople', np.nan)
        
        # 각 스레드별 상세 정보
        thread_detail = {
            'roomNumber': room_number,
            'bin': bin_number,
            'thread_id': thread_id,
            'critical_section_entered': entered_critical_section,
            'critical_section_exited': exited_critical_section,
            'critical_section_success': critical_section_success,
            'failure_reason': critical_section_failure_reason,
            'join_result': join_result,
            'critical_leave_event_type': critical_leave_event,
            'wait_time_ns': wait_time_ns,
            'dwell_time_ns': dwell_time_ns,
            'fail_processing_time_ns': fail_processing_time_ns,
            'current_people': current_people,
            'max_people': max_people,
            'waiting_start_nanoTime': row.get('waiting_start_nanoTime', np.nan),
            'critical_enter_nanoTime': row.get('critical_enter_nanoTime', np.nan),
            'critical_leave_nanoTime': row.get('critical_leave_nanoTime', np.nan),
            'increment_before_nanoTime': row.get('increment_before_nanoTime', np.nan),
            'increment_after_nanoTime': row.get('increment_after_nanoTime', np.nan)
        }
        
        thread_details_list.append(thread_detail)
    
    # DataFrame 생성 및 정렬
    result_df = pd.DataFrame(thread_details_list)
    if not result_df.empty:
        # 방 번호 → 구간 → 대기 시작 시간 순으로 정렬
        sort_columns = ['roomNumber', 'bin']
        if 'waiting_start_nanoTime' in result_df.columns:
            sort_columns.append('waiting_start_nanoTime')
        result_df = result_df.sort_values(sort_columns)
    
    return result_df


def create_time_unit_comparison(df_success):
    """
    다양한 시간 단위로 비교 통계를 생성하는 함수
    
    매개변수:
        df_success: 성공한 요청들의 DataFrame
    
    반환값:
        시간 단위 비교 DataFrame
    
    설명:
        동일한 시간 데이터를 나노초, 마이크로초, 밀리초 단위로
        각각 표시하여 가독성을 높입니다.
    """
    comparison_stats_list = []
    
    if len(df_success) > 0:
        # 각 시간 단위별로 통계 계산
        for metric_name, column_name in [('Success - Wait Time', 'wait_time_ns'), 
                                         ('Success - Dwell Time', 'dwell_time_ns')]:
            for unit, unit_name in [('ns', 'nanoseconds'), 
                                    ('us', 'microseconds'), 
                                    ('ms', 'milliseconds')]:
                stats = get_stats_with_precision(df_success[column_name], metric_name, unit)
                comparison_stats_list.append({
                    'Metric': metric_name,
                    'Unit': unit_name,
                    'Mean': stats['Mean'],
                    'Median': stats['Median'],
                    'Max': stats['Max']
                })
    
    return pd.DataFrame(comparison_stats_list)


def format_excel_file(output_path):
    """
    Excel 파일의 포맷을 설정하는 함수
    
    매개변수:
        output_path: Excel 파일 경로
    
    설명:
        숫자 포맷을 설정하여 가독성을 높입니다.
        - 백분율: 0.00%
        - 나노초: #,##0 (천 단위 구분자)
        - 마이크로초/밀리초: 소수점 포함
    """
    wb = load_workbook(output_path)
    
    # Overall_Summary 시트의 백분율 포맷
    if 'Overall_Summary' in wb.sheetnames:
        ws = wb['Overall_Summary']
        for row in range(2, ws.max_row + 1):
            cell = ws[f'C{row}']  # Percentage 컬럼
            if cell.value is not None:
                cell.number_format = '0.00"%"'
    
    # 통계 시트들의 값 포맷
    for sheet_name in ['Overall_Success_Stats', 'Overall_Capacity_Failed_Stats']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                cell = ws[f'C{row}']  # Value 컬럼
                if cell.value is not None:
                    cell.number_format = '#,##0'  # 천 단위 구분자
    
    # Per_Room_Stats 시트의 포맷
    if 'Per_Room_Stats' in wb.sheetnames:
        ws = wb['Per_Room_Stats']
        # 비율 컬럼들 (F, G, H)
        for col in ['F', 'G', 'H']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # 시간 컬럼들 (I~S) - 정원초과 실패 최대값 컬럼 추가로 범위 확장
        for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '#,##0'
    
    # Per_Bin_Stats 시트의 포맷
    if 'Per_Bin_Stats' in wb.sheetnames:
        ws = wb['Per_Bin_Stats']
        # 비율 컬럼들 (G, H, I)
        for col in ['G', 'H', 'I']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # 시간 컬럼들 (J~T) - 정원초과 실패 최대값 컬럼 추가로 범위 확장
        for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '#,##0'
    
    # Per_Thread_Critical_Details 시트의 포맷 (새로 추가)
    if 'Per_Thread_Critical_Details' in wb.sheetnames:
        ws = wb['Per_Thread_Critical_Details']
        
        if ws.max_row > 0:
            header_row = [cell.value for cell in ws[1]]
            
            # 불리언 컬럼들 (TRUE/FALSE 표시)
            boolean_columns = ['critical_section_entered', 'critical_section_exited', 'critical_section_success']
            for col_name in boolean_columns:
                try:
                    col_index = header_row.index(col_name) + 1
                    col_letter = chr(64 + col_index) if col_index <= 26 else f"A{chr(64 + col_index - 26)}"
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None:
                            # 불리언 값을 문자열로 표시
                            if isinstance(cell.value, bool):
                                cell.value = "TRUE" if cell.value else "FALSE"
                except ValueError:
                    continue
            
            # 나노초 시간 컬럼들에 천 단위 구분자 적용
            nano_time_columns = ['wait_time_ns', 'dwell_time_ns', 'fail_processing_time_ns',
                               'waiting_start_nanoTime', 'critical_enter_nanoTime', 'critical_leave_nanoTime',
                               'increment_before_nanoTime', 'increment_after_nanoTime']
            
            for col_name in nano_time_columns:
                try:
                    col_index = header_row.index(col_name) + 1
                    col_letter = chr(64 + col_index) if col_index <= 26 else f"A{chr(64 + col_index - 26)}"
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = '#,##0'
                except ValueError:
                    continue
    
    # Time_Unit_Comparison 시트의 포맷
    if 'Time_Unit_Comparison' in wb.sheetnames:
        ws = wb['Time_Unit_Comparison']
        for row in range(2, ws.max_row + 1):
            unit_cell = ws[f'B{row}']  # Unit 컬럼
            if unit_cell.value:
                # 단위에 따라 다른 포맷 적용
                for col in ['C', 'D', 'E']:  # Mean, Median, Max
                    cell = ws[f'{col}{row}']
                    if cell.value is not None:
                        if 'nanoseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0'
                        elif 'microseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0.000'
                        elif 'milliseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0.000000'
    
    wb.save(output_path)


def process_performance_data(csv_path, label):
    """
    단일 CSV 파일을 처리하여 성능 통계를 계산하고 Excel로 저장하는 메인 처리 함수
    
    매개변수:
        csv_path: 분석할 CSV 파일 경로
        label: 출력 파일에 사용할 레이블
    
    반환값:
        성공 여부 (True/False)
    """
    print(f"\n처리 중: {csv_path} (레이블: {label})")
    
    # 1. CSV 파일 로드
    try:
        # 나노초 컬럼들을 문자열로 읽어서 정밀도 유지
        dtype_spec = {
            'waiting_start_nanoTime': str,
            'critical_enter_nanoTime': str,
            'critical_leave_nanoTime': str,
            'increment_before_nanoTime': str,
            'increment_after_nanoTime': str
        }
        
        df_total = pd.read_csv(csv_path, dtype=dtype_spec)
        print(f"  - 총 {len(df_total)}개의 레코드 로드됨")
        
        # 나노초 컬럼들을 정수로 변환
        nano_columns = [col for col in dtype_spec.keys() if col in df_total.columns]
        for col in nano_columns:
            df_total[col] = df_total[col].apply(parse_nano_time_precise)
        
    except FileNotFoundError:
        print(f"오류: CSV 파일을 찾을 수 없습니다 - {csv_path}")
        return False
    except Exception as e:
        print(f"오류: CSV 파일 로드 실패 - {e}")
        return False
    
    # 2. join_result 컬럼 설정
    df_total = set_join_result_from_events(df_total)
    
    # 3. 데이터를 결과별로 분류
    df_success = df_total[df_total['join_result'] == 'SUCCESS'].copy()
    df_lock_failed = df_total[df_total['join_result'] == 'FAIL_ENTRY'].copy()
    df_capacity_failed = df_total[df_total['join_result'] == 'FAIL_OVER_CAPACITY'].copy()
    
    print(f"  - 성공: {len(df_success)}, 진입실패: {len(df_lock_failed)}, 정원초과실패: {len(df_capacity_failed)}")
    
    # 4. 성공 그룹의 시간 계산
    if len(df_success) > 0:
        # 대기 시간 = 임계 영역 진입 시간 - 대기 시작 시간
        df_success['wait_time_ns'] = df_success.apply(
            lambda row: calculate_time_diff_nano(
                row['waiting_start_nanoTime'], 
                row['critical_enter_nanoTime']
            ),
            axis=1
        )
        
        # 체류 시간 = 임계 영역 나간 시간 - 임계 영역 진입 시간
        df_success['dwell_time_ns'] = df_success.apply(
            lambda row: calculate_time_diff_nano(
                row['critical_enter_nanoTime'], 
                row['critical_leave_nanoTime']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링 (음수가 아닌 값)
        valid_success = df_success[
            (df_success['wait_time_ns'].notna()) & 
            (df_success['dwell_time_ns'].notna()) &
            (df_success['wait_time_ns'] >= 0) & 
            (df_success['dwell_time_ns'] >= 0)
        ]
        
        print(f"  - 성공 그룹 시간 계산 완료 (유효 데이터: {len(valid_success)}개/{len(df_success)}개)")
        df_success = valid_success
    
    # 5. 정원초과 실패 그룹의 시간 계산
    if len(df_capacity_failed) > 0:
        # 대기 시간 계산
        df_capacity_failed['wait_time_ns'] = df_capacity_failed.apply(
            lambda row: calculate_time_diff_nano(
                row['waiting_start_nanoTime'], 
                row['critical_enter_nanoTime']
            ),
            axis=1
        )
        
        # 실패 처리 시간 = 임계 영역 나간 시간 - 임계 영역 진입 시간
        df_capacity_failed['fail_processing_time_ns'] = df_capacity_failed.apply(
            lambda row: calculate_time_diff_nano(
                row['critical_enter_nanoTime'], 
                row['critical_leave_nanoTime']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링
        valid_capacity_failed = df_capacity_failed[
            (df_capacity_failed['wait_time_ns'].notna()) & 
            (df_capacity_failed['fail_processing_time_ns'].notna()) &
            (df_capacity_failed['wait_time_ns'] >= 0) & 
            (df_capacity_failed['fail_processing_time_ns'] >= 0)
        ]
        
        print(f"  - 정원초과 실패 그룹 시간 계산 완료 (유효 데이터: {len(valid_capacity_failed)}개/{len(df_capacity_failed)}개)")
        df_capacity_failed = valid_capacity_failed
    
    # 6. 각종 통계 DataFrame 생성
    df_summary = create_summary_stats(df_total, df_success, df_lock_failed, df_capacity_failed)
    df_success_stats = create_success_stats(df_success)
    df_capacity_failed_stats = create_capacity_failed_stats(df_capacity_failed)
    df_per_room_stats = create_per_room_stats(df_total, df_success, df_capacity_failed, df_lock_failed)
    df_per_bin_stats = create_per_bin_stats(df_total, df_success, df_capacity_failed, df_lock_failed)
    df_per_thread_critical_details = create_per_thread_critical_details(df_total, df_success, df_capacity_failed, df_lock_failed)  # 새로 추가
    df_comparison_stats = create_time_unit_comparison(df_success)
    
    # 7. 데이터 검증 출력
    total_requests = len(df_total)
    success_count = len(df_success)
    lock_failed_count = len(df_lock_failed)
    capacity_failed_count = len(df_capacity_failed)
    
    print("\n  - 데이터 검증:")
    print(f"    전체 요청 수: {total_requests}")
    print(f"    성공: {success_count} ({calculate_rate(success_count, total_requests):.2f}%)")
    print(f"    정원초과 실패: {capacity_failed_count} ({calculate_rate(capacity_failed_count, total_requests):.2f}%)")
    print(f"    진입 실패: {lock_failed_count} ({calculate_rate(lock_failed_count, total_requests):.2f}%)")
    
    # 8. Excel 파일로 저장
    output_dir = 'performance_reports'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_path = os.path.join(output_dir, f"{label}_stats_nano.xlsx")
    
    try:
        # pandas의 ExcelWriter를 사용하여 여러 시트를 한 파일에 저장
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Overall_Summary', index=False)
            df_success_stats.to_excel(writer, sheet_name='Overall_Success_Stats', index=False)
            df_capacity_failed_stats.to_excel(writer, sheet_name='Overall_Capacity_Failed_Stats', index=False)
            df_per_room_stats.to_excel(writer, sheet_name='Per_Room_Stats', index=False)
            
            # 빈 DataFrame이 아닌 경우에만 저장
            if not df_per_bin_stats.empty:
                df_per_bin_stats.to_excel(writer, sheet_name='Per_Bin_Stats', index=False)
            if not df_per_thread_critical_details.empty:  # 새로 추가된 시트
                df_per_thread_critical_details.to_excel(writer, sheet_name='Per_Thread_Critical_Details', index=False)
            if not df_comparison_stats.empty:
                df_comparison_stats.to_excel(writer, sheet_name='Time_Unit_Comparison', index=False)
        
        # Excel 파일 포맷 설정
        format_excel_file(output_path)
        
        print(f"  - Excel 파일 저장 완료: {output_path}")
        return True
        
    except Exception as e:
        print(f"오류: Excel 파일 저장 실패 - {e}")
        return False


def main():
    """
    프로그램의 메인 함수
    
    명령줄에서 인자를 받아서 처리를 시작합니다.
    
    사용 예시:
        python calculate_performance_stats.py \
            --inputs "test1.csv,test2.csv" \
            --labels "Test1,Test2"
    """
    # 명령줄 인자 파서 설정
    parser = argparse.ArgumentParser(
        description='성능 테스트 결과 CSV 파일을 분석하여 나노초 단위 통계 지표를 계산하고 Excel로 저장합니다.'
    )
    
    # --inputs 인자: 분석할 CSV 파일들 (콤마로 구분)
    parser.add_argument(
        '--inputs',
        type=str,
        required=True,
        help='분석할 CSV 파일 경로들 (콤마로 구분)'
    )
    
    # --labels 인자: 각 파일의 출력 레이블 (콤마로 구분)
    parser.add_argument(
        '--labels',
        type=str,
        required=True,
        help='각 CSV 파일에 해당하는 출력 레이블 (콤마로 구분)'
    )
    
    # --compare 인자: 비교 참조 파일 (선택사항, 현재 미사용)
    parser.add_argument(
        '--compare',
        type=str,
        help='비교할 참조 Excel 파일 경로 (선택사항)'
    )
    
    # 인자 파싱
    args = parser.parse_args()
    
    # 입력 파일과 레이블을 리스트로 변환
    input_files = [f.strip() for f in args.inputs.split(',')]
    labels = [l.strip() for l in args.labels.split(',')]
    
    # 입력 검증: 파일 수와 레이블 수가 일치해야 함
    if len(input_files) != len(labels):
        print("오류: 입력 파일 수와 레이블 수가 일치하지 않습니다.")
        sys.exit(1)
    
    # 시작 시간 기록
    start_time = datetime.now()
    print(f"성능 분석 시작: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"버전: v6.0 - 나노초 단위 분석")
    
    # 각 파일 처리
    success_count = 0
    for csv_path, label in zip(input_files, labels):
        if process_performance_data(csv_path, label):
            success_count += 1
    
    # 종료 시간 및 소요 시간 계산
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    
    # 최종 결과 출력
    print(f"\n성능 분석 완료: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"소요 시간: {elapsed_time}")
    print(f"처리 결과: {success_count}/{len(input_files)} 파일 성공")


# 이 스크립트가 직접 실행될 때만 main() 함수 호출
# (다른 스크립트에서 import할 때는 실행되지 않음)
if __name__ == "__main__":
    main()