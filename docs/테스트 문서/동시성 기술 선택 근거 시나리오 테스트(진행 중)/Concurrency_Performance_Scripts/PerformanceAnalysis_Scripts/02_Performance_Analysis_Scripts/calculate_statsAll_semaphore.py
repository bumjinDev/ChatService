#!/usr/bin/env python3
"""
세마포어 성능 분석 지표 계산 스크립트 (v1.0)
- 세마포어 tryAcquire() 기반 permit 획득/거절 성능 분석
- 나노초 단위 정밀도 측정
- 인원수 컬럼 제거된 버전

[스크립트 목적]
세마포어 전처리기에서 생성된 CSV 파일을 읽어서 세마포어 특화 통계를 계산하고,
결과를 Excel 파일로 저장합니다.

[주요 기능]
1. 세마포어 CSV 데이터 로드 (2개 나노초 시점)
2. SUCCESS/FAIL_OVER_CAPACITY 분류 (3개 카테고리)
3. permit 처리 시간 등 세마포어 특화 메트릭 계산
4. 방별, 구간별 상세 통계 생성
5. 결과를 Excel 파일로 저장

[세마포어 데이터 구조]
- true_critical_section_nanoTime_start: ATTEMPT 시점
- true_critical_section_nanoTime_end: SUCCESS/FAIL 시점  
- join_result: SUCCESS/FAIL_OVER_CAPACITY/UNKNOWN
"""

import pandas as pd
import numpy as np
import os
import argparse
from datetime import datetime
import sys
from openpyxl import load_workbook
from decimal import Decimal


def calculate_rate(count, total):
    """
    전체 대비 비율을 계산하는 함수
    
    매개변수:
        count: 특정 항목의 개수
        total: 전체 개수
    
    반환값:
        백분율 (0~100 사이의 숫자)
    """
    if isinstance(count, pd.Series) or isinstance(total, pd.Series):
        return np.where(total != 0, (count / total) * 100, 0)
    
    return (count / total) * 100 if total > 0 else 0


def get_semaphore_stats_with_precision(series, metric_name="", unit="ns"):
    """
    세마포어 데이터 시리즈의 통계값을 계산하는 함수
    
    매개변수:
        series: 통계를 계산할 데이터 시리즈 (pandas Series)
        metric_name: 출력할 메트릭 이름
        unit: 시간 단위 ("ns"=나노초, "us"=마이크로초, "ms"=밀리초)
    
    반환값:
        통계값 딕셔너리 {'Mean': 평균, 'Median': 중앙값, 'Max': 최댓값, 'Sum': 총합}
    """
    if len(series) == 0:
        return {'Mean': 0.0, 'Median': 0.0, 'Max': 0.0, 'Sum': 0.0}
    
    valid_series = series.dropna()
    if len(valid_series) == 0:
        return {'Mean': 0.0, 'Median': 0.0, 'Max': 0.0, 'Sum': 0.0}
    
    # 단위 변환 계수
    conversion = {
        'ns': 1.0,          # 나노초 (기본)
        'us': 0.001,        # 마이크로초
        'ms': 0.000001      # 밀리초
    }
    factor = conversion.get(unit, 1.0)
    
    # 기본 통계값 계산
    mean_val = float(valid_series.mean())
    median_val = float(valid_series.median())
    max_val = float(valid_series.max())
    sum_val = float(valid_series.sum())
    
    return {
        'Mean': mean_val * factor,
        'Median': median_val * factor,
        'Max': max_val * factor,
        'Sum': sum_val * factor
    }


def parse_semaphore_nano_time(nano_str):
    """
    세마포어 나노초 문자열을 정수로 변환하는 함수
    
    매개변수:
        nano_str: 나노초를 나타내는 문자열
    
    반환값:
        정수형 나노초 값 또는 NaN
    """
    if pd.isna(nano_str) or nano_str == 'nan' or nano_str == '':
        return np.nan
    
    try:
        if isinstance(nano_str, str):
            if 'E' in nano_str or 'e' in nano_str:
                decimal_val = Decimal(nano_str)
                return int(decimal_val)
            else:
                return int(float(nano_str))
        else:
            return int(nano_str)
    except (ValueError, TypeError):
        return np.nan


def calculate_semaphore_time_diff(start_nano, end_nano):
    """
    세마포어 두 나노초 시간값의 차이를 계산하는 함수
    
    매개변수:
        start_nano: 시작 시간 (ATTEMPT 나노초)
        end_nano: 종료 시간 (SUCCESS/FAIL 나노초)
    
    반환값:
        처리 시간 (나노초) 또는 NaN
    """
    if pd.isna(start_nano) or pd.isna(end_nano):
        return np.nan
    
    try:
        diff_nano = int(end_nano) - int(start_nano)
        return diff_nano if diff_nano >= 0 else np.nan
    except (ValueError, TypeError):
        return np.nan


def classify_semaphore_results(df):
    """
    세마포어 결과를 분류하는 함수
    
    매개변수:
        df: 세마포어 성능 데이터가 담긴 DataFrame
    
    반환값:
        (df_success, df_failed, df_unknown) 튜플
    
    설명:
        세마포어는 3가지 결과만 존재:
        - SUCCESS: permit 획득 성공
        - FAIL_OVER_CAPACITY: permit 획득 실패 (정원 초과)
        - UNKNOWN: 불완전한 데이터
    """
    df_success = df[df['join_result'] == 'SUCCESS'].copy()
    df_failed = df[df['join_result'] == 'FAIL_OVER_CAPACITY'].copy()
    df_unknown = df[df['join_result'] == 'UNKNOWN'].copy()
    
    return df_success, df_failed, df_unknown


def create_semaphore_summary_stats(df_total, df_success, df_failed, df_unknown):
    """
    세마포어 전체 요약 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: permit 획득 성공 요청들
        df_failed: permit 획득 실패 요청들
        df_unknown: 불완전한 요청들
    
    반환값:
        세마포어 요약 통계 DataFrame
    """
    total_requests = len(df_total)
    success_count = len(df_success)
    failed_count = len(df_failed)
    unknown_count = len(df_unknown)
    
    summary_data = {
        'Category': [
            'Total Requests',
            'Permit Acquired (SUCCESS)',
            'Permit Rejected (FAIL_OVER_CAPACITY)',
            'Incomplete Data (UNKNOWN)'
        ],
        'Count': [
            total_requests,
            success_count,
            failed_count,
            unknown_count
        ],
        'Percentage (%)': [
            100.0,
            calculate_rate(success_count, total_requests),
            calculate_rate(failed_count, total_requests),
            calculate_rate(unknown_count, total_requests)
        ]
    }
    
    return pd.DataFrame(summary_data)


def create_semaphore_success_stats(df_success):
    """
    permit 획득 성공한 요청들의 통계를 생성하는 함수
    
    매개변수:
        df_success: permit 획득 성공 요청들의 DataFrame
    
    반환값:
        세마포어 성공 통계 DataFrame
    """
    if len(df_success) == 0:
        return pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # permit 처리 시간 통계
    processing_time_stats = get_semaphore_stats_with_precision(
        df_success['permit_processing_time_ns'], 
        "Permit Processing Time", 
        "ns"
    )
    
    success_stats_data = {
        'Metric': ['Permit Processing Time'] * 4,
        'Statistic': ['Mean', 'Median', 'Max', 'Sum'],
        'Value': [
            processing_time_stats['Mean'],
            processing_time_stats['Median'],
            processing_time_stats['Max'],
            processing_time_stats['Sum']
        ],
        'Unit': ['ns'] * 4
    }
    
    return pd.DataFrame(success_stats_data)


def create_semaphore_failure_stats(df_failed):
    """
    permit 획득 실패한 요청들의 통계를 생성하는 함수
    
    매개변수:
        df_failed: permit 획득 실패 요청들의 DataFrame
    
    반환값:
        세마포어 실패 통계 DataFrame
    """
    if len(df_failed) == 0:
        return pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # permit 거절 처리 시간 통계
    rejection_time_stats = get_semaphore_stats_with_precision(
        df_failed['permit_rejection_time_ns'], 
        "Permit Rejection Time", 
        "ns"
    )
    
    failure_stats_data = {
        'Metric': ['Permit Rejection Time'] * 4,
        'Statistic': ['Mean', 'Median', 'Max', 'Sum'],
        'Value': [
            rejection_time_stats['Mean'],
            rejection_time_stats['Median'],
            rejection_time_stats['Max'],
            rejection_time_stats['Sum']
        ],
        'Unit': ['ns'] * 4
    }
    
    return pd.DataFrame(failure_stats_data)


def create_semaphore_per_room_stats(df_total, df_success, df_failed):
    """
    세마포어 방별 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: permit 획득 성공 요청들
        df_failed: permit 획득 실패 요청들
    
    반환값:
        세마포어 방별 통계 DataFrame
    """
    all_rooms = sorted(df_total['roomNumber'].unique())
    room_stats_list = []
    
    for room in all_rooms:
        # 해당 방의 데이터만 필터링
        room_total = df_total[df_total['roomNumber'] == room]
        room_success = df_success[df_success['roomNumber'] == room] if len(df_success) > 0 else pd.DataFrame()
        room_failed = df_failed[df_failed['roomNumber'] == room] if len(df_failed) > 0 else pd.DataFrame()
        
        # 기본 통계
        total_count = len(room_total)
        success_count = len(room_success)
        failed_count = len(room_failed)
        
        stats = {
            'roomNumber': room,
            'total_requests': total_count,
            'success_count': success_count,
            'failed_count': failed_count,
            'permit_success_rate(%)': calculate_rate(success_count, total_count),
            'permit_failure_rate(%)': calculate_rate(failed_count, total_count),
            
            # 성공 시 permit 처리 시간 통계
            'success_avg_permit_processing_time(ns)': room_success['permit_processing_time_ns'].mean() if success_count > 0 else 0,
            'success_median_permit_processing_time(ns)': room_success['permit_processing_time_ns'].median() if success_count > 0 else 0,
            'success_max_permit_processing_time(ns)': room_success['permit_processing_time_ns'].max() if success_count > 0 else 0,
            'success_total_permit_processing_time(ns)': room_success['permit_processing_time_ns'].sum() if success_count > 0 else 0,
            
            # 실패 시 permit 거절 시간 통계  
            'failed_avg_permit_rejection_time(ns)': room_failed['permit_rejection_time_ns'].mean() if failed_count > 0 else 0,
            'failed_median_permit_rejection_time(ns)': room_failed['permit_rejection_time_ns'].median() if failed_count > 0 else 0,
            'failed_max_permit_rejection_time(ns)': room_failed['permit_rejection_time_ns'].max() if failed_count > 0 else 0,
            'failed_total_permit_rejection_time(ns)': room_failed['permit_rejection_time_ns'].sum() if failed_count > 0 else 0
        }
        
        room_stats_list.append(stats)
    
    return pd.DataFrame(room_stats_list)


def create_semaphore_per_bin_stats(df_total, df_success, df_failed):
    """
    세마포어 방-구간별 상세 통계를 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
        df_success: permit 획득 성공 요청들
        df_failed: permit 획득 실패 요청들
    
    반환값:
        세마포어 방-구간별 통계 DataFrame
    """
    if 'bin' not in df_total.columns:
        return pd.DataFrame()
    
    all_combinations = df_total.groupby(['roomNumber', 'bin']).size().index.tolist()
    bin_stats_list = []
    
    for room, bin_num in all_combinations:
        # 해당 방-구간 조합의 데이터만 필터링
        combo_total = df_total[(df_total['roomNumber'] == room) & (df_total['bin'] == bin_num)]
        combo_success = df_success[(df_success['roomNumber'] == room) & (df_success['bin'] == bin_num)] if len(df_success) > 0 else pd.DataFrame()
        combo_failed = df_failed[(df_failed['roomNumber'] == room) & (df_failed['bin'] == bin_num)] if len(df_failed) > 0 else pd.DataFrame()
        
        # 기본 통계
        total_count = len(combo_total)
        success_count = len(combo_success)
        failed_count = len(combo_failed)
        
        stats = {
            'roomNumber': room,
            'bin': bin_num,
            'total_requests': total_count,
            'success_count': success_count,
            'failed_count': failed_count,
            'permit_success_rate(%)': calculate_rate(success_count, total_count),
            'permit_failure_rate(%)': calculate_rate(failed_count, total_count),
            
            # 성공 시 permit 처리 시간 통계
            'success_avg_permit_processing_time(ns)': combo_success['permit_processing_time_ns'].mean() if success_count > 0 else 0,
            'success_median_permit_processing_time(ns)': combo_success['permit_processing_time_ns'].median() if success_count > 0 else 0,
            'success_max_permit_processing_time(ns)': combo_success['permit_processing_time_ns'].max() if success_count > 0 else 0,
            'success_total_permit_processing_time(ns)': combo_success['permit_processing_time_ns'].sum() if success_count > 0 else 0,
            
            # 실패 시 permit 거절 시간 통계
            'failed_avg_permit_rejection_time(ns)': combo_failed['permit_rejection_time_ns'].mean() if failed_count > 0 else 0,
            'failed_median_permit_rejection_time(ns)': combo_failed['permit_rejection_time_ns'].median() if failed_count > 0 else 0,
            'failed_max_permit_rejection_time(ns)': combo_failed['permit_rejection_time_ns'].max() if failed_count > 0 else 0,
            'failed_total_permit_rejection_time(ns)': combo_failed['permit_rejection_time_ns'].sum() if failed_count > 0 else 0
        }
        
        bin_stats_list.append(stats)
    
    return pd.DataFrame(bin_stats_list)


def create_semaphore_thread_details(df_total):
    """
    세마포어 각 스레드별 permit 획득 상세 내역을 생성하는 함수
    
    매개변수:
        df_total: 전체 데이터
    
    반환값:
        세마포어 스레드별 상세 DataFrame
    """
    if 'bin' not in df_total.columns:
        return pd.DataFrame()
    
    thread_details_list = []
    
    for index, row in df_total.iterrows():
        # 기본 정보
        room_number = row['roomNumber']
        bin_number = row['bin']
        user_id = row['user_id']
        join_result = row['join_result']
        
        # permit 처리 시간
        processing_time = np.nan
        if join_result == 'SUCCESS' and 'permit_processing_time_ns' in row:
            processing_time = row['permit_processing_time_ns']
        elif join_result == 'FAIL_OVER_CAPACITY' and 'permit_rejection_time_ns' in row:
            processing_time = row['permit_rejection_time_ns']
        
        # permit 획득 성공 여부
        permit_acquired = (join_result == 'SUCCESS')
        permit_failure_reason = None if permit_acquired else ('CAPACITY_EXCEEDED' if join_result == 'FAIL_OVER_CAPACITY' else 'UNKNOWN')
        
        thread_detail = {
            'roomNumber': room_number,
            'bin': bin_number,
            'user_id': user_id,
            'permit_acquired': permit_acquired,
            'permit_failure_reason': permit_failure_reason,
            'join_result': join_result,
            'permit_processing_time_ns': processing_time,
            'attempt_nanoTime': row.get('true_critical_section_nanoTime_start', np.nan),
            'result_nanoTime': row.get('true_critical_section_nanoTime_end', np.nan)
        }
        
        thread_details_list.append(thread_detail)
    
    # DataFrame 생성 및 정렬
    result_df = pd.DataFrame(thread_details_list)
    if not result_df.empty:
        # 방 번호 → 구간 → 시도 시간 순으로 정렬
        sort_columns = ['roomNumber', 'bin']
        if 'attempt_nanoTime' in result_df.columns:
            sort_columns.append('attempt_nanoTime')
        result_df = result_df.sort_values(sort_columns, na_position='last')
    
    return result_df


def create_semaphore_time_comparison(df_success, df_failed):
    """
    세마포어 다양한 시간 단위로 비교 통계를 생성하는 함수
    
    매개변수:
        df_success: permit 획득 성공 요청들
        df_failed: permit 획득 실패 요청들
    
    반환값:
        세마포어 시간 단위 비교 DataFrame
    """
    comparison_stats_list = []
    
    # 성공 시 permit 처리 시간
    if len(df_success) > 0:
        for unit, unit_name in [('ns', 'nanoseconds'), ('us', 'microseconds'), ('ms', 'milliseconds')]:
            stats = get_semaphore_stats_with_precision(df_success['permit_processing_time_ns'], 'Permit Processing Time (Success)', unit)
            comparison_stats_list.append({
                'Metric': 'Permit Processing Time (Success)',
                'Unit': unit_name,
                'Mean': stats['Mean'],
                'Median': stats['Median'],
                'Max': stats['Max']
            })
    
    # 실패 시 permit 거절 시간
    if len(df_failed) > 0:
        for unit, unit_name in [('ns', 'nanoseconds'), ('us', 'microseconds'), ('ms', 'milliseconds')]:
            stats = get_semaphore_stats_with_precision(df_failed['permit_rejection_time_ns'], 'Permit Rejection Time (Failed)', unit)
            comparison_stats_list.append({
                'Metric': 'Permit Rejection Time (Failed)',
                'Unit': unit_name,
                'Mean': stats['Mean'],
                'Median': stats['Median'],
                'Max': stats['Max']
            })
    
    return pd.DataFrame(comparison_stats_list)


def format_semaphore_excel_file(output_path):
    """
    세마포어 Excel 파일의 포맷을 설정하는 함수
    
    매개변수:
        output_path: Excel 파일 경로
    """
    wb = load_workbook(output_path)
    
    # Semaphore_Per_Room_Stats 시트 포맷
    if 'Semaphore_Per_Room_Stats' in wb.sheetnames:
        ws = wb['Semaphore_Per_Room_Stats']
        
        # 비율 컬럼들 (D, E) - 퍼센트 포맷
        for col in ['D', 'E']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # 나노초 시간 컬럼들 (F~M) - 천 단위 구분자
        for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '#,##0'
    
    # Semaphore_Per_Bin_Stats 시트 포맷
    if 'Semaphore_Per_Bin_Stats' in wb.sheetnames:
        ws = wb['Semaphore_Per_Bin_Stats']
        
        # 비율 컬럼들 (E, F) - 퍼센트 포맷
        for col in ['E', 'F']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # 나노초 시간 컬럼들 (G~N) - 천 단위 구분자
        for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '#,##0'
    
    # 통계 시트들의 값 포맷
    for sheet_name in ['Semaphore_Success_Stats', 'Semaphore_Failure_Stats']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                cell = ws[f'C{row}']  # Value 컬럼
                if cell.value is not None:
                    unit_cell = ws[f'D{row}']  # Unit 컬럼
                    if unit_cell.value == '%':
                        cell.number_format = '0.00'  # 퍼센트는 숫자로
                    else:
                        cell.number_format = '#,##0'  # 나노초는 천 단위 구분자
    
    # Semaphore_Thread_Details 시트 포맷
    if 'Semaphore_Thread_Details' in wb.sheetnames:
        ws = wb['Semaphore_Thread_Details']
        
        if ws.max_row > 0:
            header_row = [cell.value for cell in ws[1]]
            
            # 불리언 컬럼 포맷
            if 'permit_acquired' in header_row:
                col_index = header_row.index('permit_acquired') + 1
                col_letter = chr(64 + col_index)
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col_letter}{row}']
                    if isinstance(cell.value, bool):
                        cell.value = "TRUE" if cell.value else "FALSE"
            
            # 나노초 컬럼들 포맷
            nano_columns = ['permit_processing_time_ns', 'attempt_nanoTime', 'result_nanoTime']
            for col_name in nano_columns:
                if col_name in header_row:
                    col_index = header_row.index(col_name) + 1
                    col_letter = chr(64 + col_index) if col_index <= 26 else f"A{chr(64 + col_index - 26)}"
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        if cell.value is not None and isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                            cell.number_format = '#,##0'
    
    # Semaphore_Time_Comparison 시트 포맷
    if 'Semaphore_Time_Comparison' in wb.sheetnames:
        ws = wb['Semaphore_Time_Comparison']
        for row in range(2, ws.max_row + 1):
            unit_cell = ws[f'B{row}']  # Unit 컬럼
            if unit_cell.value:
                # 단위에 따라 다른 포맷 적용
                for col in ['C', 'D', 'E']:  # Mean, Median, Max
                    cell = ws[f'{col}{row}']
                    if cell.value is not None:
                        if 'nanoseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0'
                        elif 'microseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0.000'
                        elif 'milliseconds' in str(unit_cell.value):
                            cell.number_format = '#,##0.000000'
    
    wb.save(output_path)


def process_semaphore_performance_data(csv_path, label):
    """
    세마포어 CSV 파일을 처리하여 성능 통계를 계산하고 Excel로 저장하는 메인 처리 함수
    
    매개변수:
        csv_path: 분석할 세마포어 CSV 파일 경로
        label: 출력 파일에 사용할 레이블
    
    반환값:
        성공 여부 (True/False)
    """
    print(f"\n세마포어 데이터 처리 중: {csv_path} (레이블: {label})")
    
    # 1. 세마포어 CSV 파일 로드
    try:
        # 세마포어 나노초 컬럼들을 문자열로 읽어서 정밀도 유지
        dtype_spec = {
            'true_critical_section_nanoTime_start': str,
            'true_critical_section_nanoTime_end': str
        }
        
        df_total = pd.read_csv(csv_path, dtype=dtype_spec)
        print(f"  - 총 {len(df_total)}개의 세마포어 레코드 로드됨")
        
        # 필수 컬럼 검증
        required_columns = ['roomNumber', 'user_id', 'join_result']
        missing_columns = [col for col in required_columns if col not in df_total.columns]
        if missing_columns:
            print(f"오류: 필수 컬럼 누락 - {missing_columns}")
            return False
        
        # 나노초 컬럼들을 정수로 변환
        for col in ['true_critical_section_nanoTime_start', 'true_critical_section_nanoTime_end']:
            if col in df_total.columns:
                df_total[col] = df_total[col].apply(parse_semaphore_nano_time)
        
    except FileNotFoundError:
        print(f"오류: CSV 파일을 찾을 수 없습니다 - {csv_path}")
        return False
    except Exception as e:
        print(f"오류: CSV 파일 로드 실패 - {e}")
        return False
    
    # 2. 세마포어 결과별로 분류
    df_success, df_failed, df_unknown = classify_semaphore_results(df_total)
    
    print(f"  - permit 획득 성공: {len(df_success)}")
    print(f"  - permit 획득 실패: {len(df_failed)}")
    print(f"  - 불완전한 데이터: {len(df_unknown)}")
    
    # 3. 성공 그룹의 세마포어 메트릭 계산
    if len(df_success) > 0:
        # permit 처리 시간 계산
        df_success['permit_processing_time_ns'] = df_success.apply(
            lambda row: calculate_semaphore_time_diff(
                row['true_critical_section_nanoTime_start'], 
                row['true_critical_section_nanoTime_end']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링
        valid_success = df_success[
            (df_success['permit_processing_time_ns'].notna()) & 
            (df_success['permit_processing_time_ns'] >= 0)
        ]
        
        print(f"  - 성공 그룹 메트릭 계산 완료 (유효 데이터: {len(valid_success)}개/{len(df_success)}개)")
        df_success = valid_success
    
    # 4. 실패 그룹의 세마포어 메트릭 계산
    if len(df_failed) > 0:
        # permit 거절 처리 시간 계산
        df_failed['permit_rejection_time_ns'] = df_failed.apply(
            lambda row: calculate_semaphore_time_diff(
                row['true_critical_section_nanoTime_start'], 
                row['true_critical_section_nanoTime_end']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링
        valid_failed = df_failed[
            (df_failed['permit_rejection_time_ns'].notna()) & 
            (df_failed['permit_rejection_time_ns'] >= 0)
        ]
        
        print(f"  - 실패 그룹 메트릭 계산 완료 (유효 데이터: {len(valid_failed)}개/{len(df_failed)}개)")
        df_failed = valid_failed
    
    # 5. 세마포어 통계 DataFrame 생성
    df_summary = create_semaphore_summary_stats(df_total, df_success, df_failed, df_unknown)
    df_success_stats = create_semaphore_success_stats(df_success)
    df_failure_stats = create_semaphore_failure_stats(df_failed)
    df_per_room_stats = create_semaphore_per_room_stats(df_total, df_success, df_failed)
    df_per_bin_stats = create_semaphore_per_bin_stats(df_total, df_success, df_failed)
    df_thread_details = create_semaphore_thread_details(df_total)
    df_time_comparison = create_semaphore_time_comparison(df_success, df_failed)
    
    # 6. 데이터 검증 출력
    total_requests = len(df_total)
    success_count = len(df_success)
    failed_count = len(df_failed)
    unknown_count = len(df_unknown)
    
    print("\n  - 세마포어 데이터 검증:")
    print(f"    전체 요청 수: {total_requests}")
    print(f"    permit 획득 성공: {success_count} ({calculate_rate(success_count, total_requests):.2f}%)")
    print(f"    permit 획득 실패: {failed_count} ({calculate_rate(failed_count, total_requests):.2f}%)")
    print(f"    불완전한 데이터: {unknown_count} ({calculate_rate(unknown_count, total_requests):.2f}%)")
    
    # 7. Excel 파일로 저장
    output_dir = 'semaphore_performance_reports'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_path = os.path.join(output_dir, f"{label}_semaphore_stats.xlsx")
    
    try:
        # pandas의 ExcelWriter를 사용하여 세마포어 전용 시트들 저장
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Semaphore_Summary', index=False)
            df_success_stats.to_excel(writer, sheet_name='Semaphore_Success_Stats', index=False)
            df_failure_stats.to_excel(writer, sheet_name='Semaphore_Failure_Stats', index=False)
            df_per_room_stats.to_excel(writer, sheet_name='Semaphore_Per_Room_Stats', index=False)
            
            # 빈 DataFrame이 아닌 경우에만 저장
            if not df_per_bin_stats.empty:
                df_per_bin_stats.to_excel(writer, sheet_name='Semaphore_Per_Bin_Stats', index=False)
            if not df_thread_details.empty:
                df_thread_details.to_excel(writer, sheet_name='Semaphore_Thread_Details', index=False)
            if not df_time_comparison.empty:
                df_time_comparison.to_excel(writer, sheet_name='Semaphore_Time_Comparison', index=False)
        
        # Excel 파일 포맷 설정
        format_semaphore_excel_file(output_path)
        
        print(f"  - 세마포어 Excel 파일 저장 완료: {output_path}")
        return True
        
    except Exception as e:
        print(f"오류: Excel 파일 저장 실패 - {e}")
        return False


def main():
    """
    세마포어 성능 분석 프로그램의 메인 함수
    
    명령줄에서 인자를 받아서 세마포어 전용 처리를 시작합니다.
    
    사용 예시:
        python semaphore_calculate_stats.py \
            --inputs "preprocessor_semaphore.csv" \
            --labels "Semaphore_Test1"
    """
    # 명령줄 인자 파서 설정
    parser = argparse.ArgumentParser(
        description='세마포어 성능 테스트 결과 CSV 파일을 분석하여 나노초 단위 통계 지표를 계산하고 Excel로 저장합니다.'
    )
    
    # --inputs 인자: 분석할 세마포어 CSV 파일들 (콤마로 구분)
    parser.add_argument(
        '--inputs',
        type=str,
        required=True,
        help='분석할 세마포어 CSV 파일 경로들 (콤마로 구분)'
    )
    
    # --labels 인자: 각 파일의 출력 레이블 (콤마로 구분)
    parser.add_argument(
        '--labels',
        type=str,
        required=True,
        help='각 세마포어 CSV 파일에 해당하는 출력 레이블 (콤마로 구분)'
    )
    
    # 인자 파싱
    args = parser.parse_args()
    
    # 입력 파일과 레이블을 리스트로 변환
    input_files = [f.strip() for f in args.inputs.split(',')]
    labels = [l.strip() for l in args.labels.split(',')]
    
    # 입력 검증: 파일 수와 레이블 수가 일치해야 함
    if len(input_files) != len(labels):
        print("오류: 입력 파일 수와 레이블 수가 일치하지 않습니다.")
        sys.exit(1)
    
    # 시작 시간 기록
    start_time = datetime.now()
    print(f"세마포어 성능 분석 시작: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"버전: v1.0 - 세마포어 전용 + permit 기반 분석 + 나노초 정밀도")
    
    # 각 파일 처리
    success_count = 0
    for csv_path, label in zip(input_files, labels):
        if process_semaphore_performance_data(csv_path, label):
            success_count += 1
    
    # 종료 시간 및 소요 시간 계산
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    
    # 최종 결과 출력
    print(f"\n세마포어 성능 분석 완료: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"소요 시간: {elapsed_time}")
    print(f"처리 결과: {success_count}/{len(input_files)} 파일 성공")
    
    if success_count > 0:
        print(f"\n📊 생성된 Excel 시트:")
        print(f"  - Semaphore_Summary: 전체 요약 (permit 성공/실패 비율)")
        print(f"  - Semaphore_Success_Stats: permit 획득 성공 통계")
        print(f"  - Semaphore_Failure_Stats: permit 획득 실패 통계")
        print(f"  - Semaphore_Per_Room_Stats: 방별 세마포어 성능")
        print(f"  - Semaphore_Per_Bin_Stats: 구간별 세마포어 성능")
        print(f"  - Semaphore_Thread_Details: 스레드별 permit 획득 상세")
        print(f"  - Semaphore_Time_Comparison: 시간 단위별 비교")


# 이 스크립트가 직접 실행될 때만 main() 함수 호출
if __name__ == "__main__":
    main()