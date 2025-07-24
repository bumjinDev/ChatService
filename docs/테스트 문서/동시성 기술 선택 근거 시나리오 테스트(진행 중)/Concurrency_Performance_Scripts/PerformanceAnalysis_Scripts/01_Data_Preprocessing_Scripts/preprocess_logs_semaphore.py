#!/usr/bin/env python3
"""
세마포어 성능 측정 로그 전처리 스크립트

[목적]
세마포어 기반 멀티스레드 환경에서 permit 획득/해제 로그를 파싱하여
성능 분석용 데이터로 변환합니다.

[주요 기능]
1. 세마포어 성능 로그에서 3가지 이벤트 추출 (ATTEMPT, SUCCESS, FAIL)
2. 사용자별, 방별로 이벤트 그룹화 및 페어링
3. 나노초 정밀도 시간 정렬
4. CSV 및 Excel 형식으로 결과 저장
5. 사용자 지정 출력 디렉토리 지원

[세마포어 이벤트 의미]
- SEMAPHORE_EXISTING_ATTEMPT: tryAcquire() 호출 직전 (permit 획득 시도)
- SEMAPHORE_EXISTING_SUCCESS: permit 획득 성공
- SEMAPHORE_EXISTING_FAIL: permit 획득 실패 (즉시 거절)

[출력 컬럼 구성]
- roomNumber: 방 번호
- bin: 방별 10개 구간 분할
- room_entry_sequence: 방별 처리 순번
- user_id: 사용자 식별자
- prev_people: ATTEMPT 시점 현재 인원
- curr_people: SUCCESS/FAIL 후 인원
- max_people: 방 최대 정원
- join_result: SUCCESS/FAIL_OVER_CAPACITY
- true_critical_section_nanoTime_start: ATTEMPT 시점 나노초 타임스탬프
- true_critical_section_nanoTime_end: SUCCESS/FAIL 시점 나노초 타임스탬프
"""

import pandas as pd
import re
import os
import shutil
import argparse
import sys
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook

# ===== 상수 정의 =====
# 파일 경로 상수
LOG_FILE = 'ChatService.log'
NEW_LOG_PATH = r'E:\devSpace\ChatServiceTest\log\ChatService.log'

# 세마포어 이벤트 타입 상수
EVENT_SEMAPHORE_ATTEMPT = 'SEMAPHORE_EXISTING_ATTEMPT'
EVENT_SEMAPHORE_SUCCESS = 'SEMAPHORE_EXISTING_SUCCESS'
EVENT_SEMAPHORE_FAIL = 'SEMAPHORE_EXISTING_FAIL'

# 결과 타입 상수
RESULT_SUCCESS = 'SUCCESS'
RESULT_FAIL_CAPACITY = 'FAIL_OVER_CAPACITY'
RESULT_UNKNOWN = 'UNKNOWN'

# 분석 구간 수
BINS_COUNT = 10

# 세마포어 성능 로그 정규식 패턴
SEMAPHORE_PERFORMANCE_PATTERN = re.compile(
    r'SEMAPHORE_PERFORMANCE_MARK tag=(?P<tag>SEMAPHORE_EXISTING_ATTEMPT|SEMAPHORE_EXISTING_SUCCESS|SEMAPHORE_EXISTING_FAIL)'
    r' timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<event>[\w_]+)'
    r'.* roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r' currentPeople=(?P<currentPeople>\d+)'
    r' maxPeople=(?P<maxPeople>\d+)'
    r' nanoTime=(?P<nanoTime>\d+)'
    r' epochNano=(?P<epochNano>\d+)'
    r' threadId=(?P<threadId>\d+)'
)


def test_semaphore_pattern():
    """SEMAPHORE_PERFORMANCE_PATTERN 정규식 테스트 함수"""
    test_logs = [
        "SEMAPHORE_PERFORMANCE_MARK tag=SEMAPHORE_EXISTING_ATTEMPT timestampIso=2025-07-23T02:24:37.452142100Z event=PRE_ACQUIRE_EXISTING_ROOM className=RoomJoinServiceSemaphore methodName=confirmJoinRoom roomNumber=1203 userId=user123 currentPeople=3 maxPeople=5 nanoTime=38434598060100 epochNano=1752891877453106700 threadId=247",
        "SEMAPHORE_PERFORMANCE_MARK tag=SEMAPHORE_EXISTING_SUCCESS timestampIso=2025-07-23T02:24:37.452142200Z event=POST_ACQUIRE_EXISTING_ROOM_SUCCESS className=RoomJoinServiceSemaphore methodName=confirmJoinRoom roomNumber=1203 userId=user123 currentPeople=4 maxPeople=5 nanoTime=38434598060200 epochNano=1752891877453106800 threadId=247",
        "SEMAPHORE_PERFORMANCE_MARK tag=SEMAPHORE_EXISTING_FAIL timestampIso=2025-07-23T02:24:37.452142300Z event=POST_ACQUIRE_EXISTING_ROOM_FAIL className=RoomJoinServiceSemaphore methodName=confirmJoinRoom roomNumber=1203 userId=user456 currentPeople=5 maxPeople=5 nanoTime=38434598060300 epochNano=1752891877453106900 threadId=248"
    ]
    
    success_count = 0
    for i, test_log in enumerate(test_logs):
        match = SEMAPHORE_PERFORMANCE_PATTERN.search(test_log)
        if match:
            result = match.groupdict()
            print(f"✅ 테스트 {i+1} 매칭 성공: tag={result['tag']}, event={result['event']}, currentPeople={result['currentPeople']}")
            success_count += 1
        else:
            print(f"❌ 테스트 {i+1} 매칭 실패")
    
    return success_count == len(test_logs)


def replace_log_file() -> None:
    """
    기존 로그 파일을 새 로그 파일로 교체
    """
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    shutil.copy(NEW_LOG_PATH, LOG_FILE)
    print(f"로그 파일 교체 완료: {NEW_LOG_PATH} → {LOG_FILE}")


def parse_log_line(line: str) -> Optional[Dict[str, Any]]:
    """
    로그 라인을 파싱하여 세마포어 성능 이벤트 정보 추출
    """
    # SEMAPHORE_PERFORMANCE_MARK 패턴 매칭
    match = SEMAPHORE_PERFORMANCE_PATTERN.search(line)
    if match:
        data = match.groupdict()
        return data
    
    return None


def parse_semaphore_events(filepath: str, room_number: Optional[int] = None) -> pd.DataFrame:
    """
    로그 파일에서 세마포어 성능 이벤트를 파싱하여 DataFrame으로 변환
    """
    records = []
    
    try:
        with open(filepath, encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                data = parse_log_line(line)
                
                if data:
                    # 방 번호를 정수로 변환
                    data['roomNumber'] = int(data['roomNumber'])
                    
                    # 나노초 및 기타 수치 값들을 적절한 타입으로 변환
                    data['currentPeople'] = int(data['currentPeople'])
                    data['maxPeople'] = int(data['maxPeople'])
                    data['nanoTime'] = data['nanoTime']  # 문자열로 유지 (정밀도 보존)
                    data['epochNano'] = data['epochNano']  # 문자열로 유지
                    data['threadId'] = int(data['threadId'])
                    
                    # 방 번호 필터링
                    if room_number is None or data['roomNumber'] == room_number:
                        records.append(data)
    
    except FileNotFoundError:
        print(f"오류: 로그 파일을 찾을 수 없습니다 - {filepath}")
        return pd.DataFrame()
    except Exception as e:
        print(f"오류: 로그 파일 파싱 실패 - {e}")
        return pd.DataFrame()
    
    df = pd.DataFrame(records)
    print(f"📊 세마포어 성능 이벤트 파싱 완료: {len(df)}개 이벤트")
    
    if not df.empty:
        # 이벤트 타입별 통계
        event_counts = df['tag'].value_counts()
        for event_type, count in event_counts.items():
            print(f"  - {event_type}: {count}건")
    
    return df


def process_semaphore_user_events(user_id: str, user_data: pd.DataFrame) -> Optional[Dict[str, Any]]:
    """
    단일 사용자의 세마포어 이벤트들을 처리하여 성능 프로필 생성
    """
    # nanoTime 기준으로 정렬
    user_data = user_data.sort_values(['nanoTime_num']).reset_index(drop=True)
    
    # 이벤트별로 그룹화
    events = {}
    for _, row in user_data.iterrows():
        event_name = row['tag']
        events[event_name] = {
            'timestamp': row['timestamp'],
            'nanoTime': str(int(float(row['nanoTime']))),
            'epochNano': str(int(float(row['epochNano']))),
            'currentPeople': row['currentPeople'],
            'maxPeople': row['maxPeople'],
            'threadId': row['threadId']
        }
    
    # ATTEMPT 이벤트가 없으면 유효하지 않은 세션으로 간주
    if EVENT_SEMAPHORE_ATTEMPT not in events:
        return None
    
    # 기본 프로필 정보
    profile = {
        'user_id': user_id,
        'roomNumber': user_data.iloc[0]['roomNumber']
    }
    
    # ATTEMPT 시점 정보 (시작점)
    attempt_event = events[EVENT_SEMAPHORE_ATTEMPT]
    profile.update({
        'true_critical_section_nanoTime_start': attempt_event['nanoTime']
    })
    
    # SUCCESS 또는 FAIL 결과 확인
    if EVENT_SEMAPHORE_SUCCESS in events:
        success_event = events[EVENT_SEMAPHORE_SUCCESS]
        profile.update({
            'join_result': RESULT_SUCCESS,
            'true_critical_section_nanoTime_end': success_event['nanoTime']
        })
    elif EVENT_SEMAPHORE_FAIL in events:
        fail_event = events[EVENT_SEMAPHORE_FAIL]
        profile.update({
            'join_result': RESULT_FAIL_CAPACITY,
            'true_critical_section_nanoTime_end': fail_event['nanoTime']
        })
    else:
        # ATTEMPT만 있고 결과가 없는 경우 (불완전한 세션)
        profile.update({
            'join_result': RESULT_UNKNOWN,
            'true_critical_section_nanoTime_end': attempt_event['nanoTime']  # 같은 시점으로 설정
        })
    
    return profile


def build_semaphore_performance_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    파싱된 세마포어 로그 데이터를 성능 분석용 데이터로 변환
    
    처리 과정:
        1. 타임스탬프 변환 및 nanoTime 기준 정렬
        2. 사용자별 이벤트 페어링 (ATTEMPT + SUCCESS/FAIL)
        3. 세마포어 특화 컬럼 생성 (prev_people, curr_people 등)
        4. 방별 나노초 시간 기준 정렬
        5. bin 구간 및 room_entry_sequence 부여
    """
    if df.empty:
        return pd.DataFrame()
    
    print("🔧 세마포어 성능 데이터 구축 시작...")
    
    # 타임스탬프를 datetime 객체로 변환
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    
    # 나노초 값들을 숫자로 변환 (정렬용)
    df['nanoTime_num'] = df['nanoTime'].astype(float)
    df['epochNano_num'] = df['epochNano'].astype(float)
    
    # 방별 + nanoTime 기준 정렬 (기본 정렬)
    df = df.sort_values(['roomNumber', 'nanoTime_num']).reset_index(drop=True)
    
    performance_results = []
    
    # 각 방별로 처리
    for room_num in df['roomNumber'].unique():
        print(f"🏠 방 {room_num} 처리 중...")
        
        room_events = df[df['roomNumber'] == room_num].copy()
        
        # 사용자별로 그룹화하여 세마포어 이벤트 페어링
        for user_id, user_data in room_events.groupby('userId'):
            profile = process_semaphore_user_events(user_id, user_data)
            if profile:  # 유효한 프로필만 추가
                performance_results.append(profile)
                print(f"   ✅ 세마포어 프로필 생성: user_id={profile['user_id']}, result={profile['join_result']}")
    
    if not performance_results:
        print("⚠️ 유효한 세마포어 세션이 없습니다.")
        return pd.DataFrame()
    
    # DataFrame 생성 및 정리
    result_df = pd.DataFrame(performance_results)
    
    # 나노초 컬럼들을 문자열로 확실히 변환
    nano_columns = [col for col in result_df.columns if 'nanoTime' in col]
    for col in nano_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col].astype(str)
    
    # 세마포어 특화 정렬 및 구간 할당
    result_df = sort_and_assign_semaphore_bins(result_df)
    
    # 컬럼 순서 조정
    result_df = reorder_columns(result_df)
    
    print(f"✅ 세마포어 성능 데이터 구축 완료: {len(result_df)}개 세션")
    return result_df


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    컬럼 순서를 지정된 순서로 재정렬
    """
    desired_order = [
        'roomNumber',
        'bin', 
        'room_entry_sequence',
        'user_id',
        'join_result',
        'true_critical_section_nanoTime_start',
        'true_critical_section_nanoTime_end'
    ]
    
    # 존재하는 컬럼만 선택하여 재정렬
    existing_columns = [col for col in desired_order if col in df.columns]
    return df[existing_columns]


def sort_and_assign_semaphore_bins(df: pd.DataFrame) -> pd.DataFrame:
    """
    세마포어 데이터를 나노초 시간 기준으로 정렬하고 bin 구간 할당
    """
    if df.empty:
        return df
    
    print("📊 세마포어 데이터 정렬 및 bin 할당 중...")
    
    # ATTEMPT 시점 나노초 기준으로 정렬 (전체 경쟁 시작 순서)
    if 'true_critical_section_nanoTime_start' in df.columns:
        # 나노초 값을 숫자로 변환하여 정렬
        df['start_nanoTime_num'] = pd.to_numeric(df['true_critical_section_nanoTime_start'], errors='coerce')
        df['start_nanoTime_num'] = df['start_nanoTime_num'].fillna(float('inf'))
        
        # 방 번호 → 시작 나노초 순으로 정렬
        df = df.sort_values(['roomNumber', 'start_nanoTime_num']).reset_index(drop=True)
        
        # 정렬용 임시 컬럼 제거
        df = df.drop(columns=['start_nanoTime_num'])
        
        print("✅ ATTEMPT 시점 나노초 기준 정렬 완료")
    else:
        # fallback: 방 번호만으로 정렬
        df = df.sort_values(['roomNumber']).reset_index(drop=True)
        print("⚠️ 나노초 컬럼이 없어서 방 번호만으로 정렬")
    
    # 방별로 bin과 room_entry_sequence 할당
    for room_num in df['roomNumber'].unique():
        room_mask = df['roomNumber'] == room_num
        room_data = df[room_mask]
        room_indices = room_data.index
        total_requests = len(room_data)
        
        # room_entry_sequence 할당 (1부터 시작)
        df.loc[room_indices, 'room_entry_sequence'] = range(1, total_requests + 1)
        
        # bin 할당 (10개 구간으로 균등 분할)
        if total_requests <= BINS_COUNT:
            # 요청이 10개 이하면 각각을 하나의 bin으로
            bins = range(1, total_requests + 1)
        else:
            # 10개 구간으로 균등 분할
            bins = pd.cut(range(total_requests), bins=BINS_COUNT, labels=range(1, BINS_COUNT + 1)).astype(int)
        
        df.loc[room_indices, 'bin'] = bins
        
        print(f"  방 {room_num}: {total_requests}개 세션에 순서 번호와 {len(set(bins))}개 bin 할당")
    
    print("✅ bin과 room_entry_sequence 할당 완료")
    return df


def get_semaphore_desc_table() -> List[List[str]]:
    """
    Excel 파일에 추가할 세마포어 컬럼 설명 테이블 생성
    """
    return [
        ["속성명", "측정 목적", "도출 방법"],
        ["roomNumber", "방 번호 식별", "로그 필드: roomNumber"],
        ["bin", "방별 분석 구간", "각 방의 요청을 10개 구간으로 균등 분할"],
        ["room_entry_sequence", "방별 처리 순번", "ATTEMPT 나노초 시간 순서"],
        ["user_id", "사용자 식별", "로그 필드: userId"],
        ["join_result", "입장 결과 구분", "SUCCESS/FAIL_OVER_CAPACITY/UNKNOWN"],
        ["true_critical_section_nanoTime_start", "permit 획득 시도 시점", "ATTEMPT 이벤트의 nanoTime"],
        ["true_critical_section_nanoTime_end", "permit 획득 결과 시점", "SUCCESS/FAIL 이벤트의 nanoTime"],
        ["세마포어 특징", "CAS 기반 동시성 제어", "tryAcquire()의 즉시 성공/실패 특성"]
    ]


def convert_nano_value(x: Any) -> str:
    """
    나노초 값을 문자열로 안전하게 변환
    """
    if pd.isna(x) or x == '' or str(x).lower() == 'nan':
        return ''
    try:
        return str(int(float(x)))
    except (ValueError, TypeError):
        return str(x)


def save_to_csv(df: pd.DataFrame, filepath: str) -> None:
    """
    DataFrame을 CSV 파일로 저장 (나노초 정밀도 유지)
    """
    # 나노초 컬럼들을 문자열로 변환하여 지수 표기법 방지
    df_copy = df.copy()
    nano_columns = [col for col in df.columns if 'nanoTime' in col]
    
    for col in nano_columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(convert_nano_value)
    
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # UTF-8 with BOM으로 저장 (Excel에서 한글 깨짐 방지)
    df_copy.to_csv(filepath, index=False, encoding='utf-8-sig')
    print(f"CSV 파일 저장 완료: {filepath}")


def save_with_side_table(df: pd.DataFrame, filepath: str, desc_table: List[List[str]]) -> str:
    """
    DataFrame을 Excel로 저장하고 설명 테이블 추가
    """
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # timezone 제거한 복사본 생성
    df_excel = df.copy()
    for col in df_excel.columns:
        if pd.api.types.is_datetime64_tz_dtype(df_excel[col]):
            df_excel[col] = df_excel[col].dt.tz_localize(None)
    
    # Excel 파일 생성
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False)
        
        # 나노초 컬럼들을 텍스트 형식으로 설정
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        nano_columns = [col for col in df.columns if 'nanoTime' in col]
        for col_name in nano_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=col_idx)
                    cell.number_format = '@'  # 텍스트 형식
    
    # 설명 테이블 추가
    wb = load_workbook(filepath)
    ws = wb.active
    
    start_col = len(df.columns) + 2
    
    for i, row in enumerate(desc_table):
        for j, val in enumerate(row):
            ws.cell(row=i + 1, column=start_col + j, value=val)
    
    wb.save(filepath)
    print(f"Excel 파일 저장 완료: {filepath}")
    
    return filepath


def analyze_semaphore_results(df: pd.DataFrame) -> None:
    """
    세마포어 처리 결과 분석 및 출력
    """
    if df.empty:
        print("분석할 데이터가 없습니다.")
        return
    
    total_operations = len(df)
    
    print(f"\n{'='*60}")
    print(f"세마포어 성능 분석 결과")
    print(f"{'='*60}")
    print(f"전체 처리 작업 수: {total_operations}")
    
    # 입장 결과별 분석
    if 'join_result' in df.columns:
        print(f"\n[세마포어 permit 획득 결과 분석]")
        join_result_counts = df['join_result'].value_counts()
        for result, count in join_result_counts.items():
            print(f"  {result}: {count}건 ({count/total_operations*100:.1f}%)")
    
    # 세마포어 성능 특성 분석
    print(f"\n[세마포어 성능 특성 분석]")
    if 'true_critical_section_nanoTime_start' in df.columns and 'true_critical_section_nanoTime_end' in df.columns:
        # 처리 시간 계산 (나노초)
        df['processing_time_nano'] = pd.to_numeric(df['true_critical_section_nanoTime_end']) - pd.to_numeric(df['true_critical_section_nanoTime_start'])
        processing_times = df['processing_time_nano'].dropna()
        
        if len(processing_times) > 0:
            avg_processing_time = processing_times.mean()
            min_processing_time = processing_times.min()
            max_processing_time = processing_times.max()
            
            print(f"  평균 처리 시간: {avg_processing_time:.0f} 나노초 ({avg_processing_time/1_000_000:.3f} ms)")
            print(f"  최소 처리 시간: {min_processing_time:.0f} 나노초")
            print(f"  최대 처리 시간: {max_processing_time:.0f} 나노초")
    
    # 방별 성능 통계
    print(f"\n[방별 성능 통계]")
    room_stats = df.groupby('roomNumber').agg({
        'user_id': 'count',
        'bin': 'nunique'
    })
    room_stats.columns = ['total_operations', 'bin_count']
    
    for room_num, stats in room_stats.iterrows():
        print(f"  방 {room_num}: 총 {stats['total_operations']}회, 구간 수: {stats['bin_count']}")
        
        if 'join_result' in df.columns:
            room_data = df[df['roomNumber'] == room_num]
            room_results = room_data['join_result'].value_counts()
            for result, count in room_results.items():
                print(f"    - {result}: {count}건")
    
    # 동시성 수준 분석
    if False:  # prev_people, max_people 컬럼 제거로 인한 비활성화
        pass
    
    # 나노초 정밀도 확인
    print(f"\n[나노초 정밀도 확인 (첫 3개 샘플)]")
    nano_columns = [col for col in df.columns if 'nanoTime' in col]
    if nano_columns:
        for i in range(min(3, len(df))):
            print(f"  샘플 {i+1}:")
            for col in nano_columns:
                if col in df.columns:
                    print(f"    {col}: {df.iloc[i][col]}")


def main():
    """
    메인 실행 함수 (세마포어 전용)
    """
    parser = argparse.ArgumentParser(
        description="세마포어 성능 측정 로그 전처리기: ATTEMPT → SUCCESS/FAIL 페어링",
        epilog="예시: python semaphore_preprocess_logs.py --output_dir results --room 1 --xlsx semaphore_output.xlsx"
    )
    parser.add_argument('--output_dir', type=str, default='semaphore_results', 
                        help='출력 디렉토리 경로 (기본값: semaphore_results)')
    parser.add_argument('--room', type=int, help='특정 방 번호만 처리 (옵션)')
    parser.add_argument('--csv', type=str, help='추가 CSV 파일명 (옵션)')
    parser.add_argument('--xlsx', type=str, help='Excel 파일명 (옵션)')
    parser.add_argument('--test', action='store_true', help='정규식 패턴 테스트 실행')
    
    args = parser.parse_args()
    
    # 정규식 테스트 옵션
    if args.test:
        print("🧪 SEMAPHORE_PERFORMANCE_PATTERN 정규식 테스트 실행...")
        test_result = test_semaphore_pattern()
        if test_result:
            print("✅ 정규식 테스트 통과")
        else:
            print("❌ 정규식 테스트 실패 - 패턴을 재검토하세요")
        return
    
    try:
        # 1. 로그 파일 교체
        replace_log_file()
        
        # 2. 세마포어 로그 파싱
        print(f"\n세마포어 성능 로그 파싱 중...")
        df = parse_semaphore_events(LOG_FILE, room_number=args.room)
        print(f"파싱 완료: {len(df)}개 세마포어 이벤트")
        
        if df.empty:
            print("⚠️ 세마포어 성능 로그를 찾을 수 없습니다. 로그 설정을 확인하세요.")
            return
        
        # 3. 세마포어 성능 데이터 구축
        print(f"\n세마포어 성능 데이터 구축 중...")
        result = build_semaphore_performance_data(df)
        
        if result.empty:
            print("⚠️ 유효한 세마포어 세션이 생성되지 않았습니다.")
            return
            
        print(f"구축 완료: {len(result)}개 세마포어 세션")
        
        # 4. 출력 디렉토리 설정 및 생성
        output_dir = args.output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"출력 디렉토리 생성: {output_dir}")
        
        # 5. 기본 CSV 파일 저장 (세마포어 전용 명명)
        if args.room:
            base_filename = f'preprocessor_performance_semaphore_romm_{args.room}.csv'
        else:
            base_filename = 'preprocessor_performance_semaphore.csv'
        
        csv_path = os.path.join(output_dir, base_filename)
        save_to_csv(result, csv_path)
        
        # 6. 추가 CSV 파일 저장 (옵션)
        if args.csv:
            additional_csv_path = os.path.join(output_dir, args.csv)
            save_to_csv(result, additional_csv_path)
        
        # 7. Excel 파일 저장 (옵션)
        if args.xlsx:
            xlsx_path = os.path.join(output_dir, args.xlsx)
            desc_table = get_semaphore_desc_table()
            save_with_side_table(result, xlsx_path, desc_table)
        
        # 8. 세마포어 결과 분석 출력
        analyze_semaphore_results(result)
        
        print(f"\n{'='*60}")
        print(f"세마포어 전용 전처리 완료!")
        print(f"정렬 순서: ATTEMPT 나노초 시간 기준")
        print(f"출력 디렉토리: {os.path.abspath(output_dir)}")
        print(f"{'='*60}")
        
    except FileNotFoundError as e:
        print(f"\n오류: 파일을 찾을 수 없습니다 - {e}")
        sys.exit(1)
    except PermissionError as e:
        print(f"\n오류: 파일 접근 권한이 없습니다 - {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n예상치 못한 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# 스크립트가 직접 실행될 때만 main() 호출
if __name__ == '__main__':
    main()