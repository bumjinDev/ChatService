#!/usr/bin/env python3
"""
임계 영역 성능 측정 로그 전처리 스크립트

[목적]
멀티스레드 환경에서 임계 영역(Critical Section) 접근 로그를 파싱하여
성능 분석용 데이터로 변환합니다.

[주요 기능]
1. 로그 파일에서 5가지 이벤트 추출 (WAITING_START, CRITICAL_ENTER, CRITICAL_LEAVE, INCREMENT_BEFORE, INCREMENT_AFTER)
2. 사용자별, 방별로 이벤트 그룹화
3. 시간순 정렬 및 구간(bin) 분할
4. CSV 및 Excel 형식으로 결과 저장
5. 사용자 지정 출력 디렉토리 지원
6. PRE_CHECK_FAIL_OVER_CAPACITY 독립 이벤트 처리 추가

[이벤트 의미]
- WAITING_START: 임계 영역 진입 대기 시작
- CRITICAL_ENTER: 임계 영역 진입 성공
- CRITICAL_LEAVE: 임계 영역 퇴장
- INCREMENT_BEFORE: 카운터 증가 작업 시작
- INCREMENT_AFTER: 카운터 증가 작업 완료
- PRE_CHECK_FAIL_OVER_CAPACITY: 락 외부 사전 차단 (독립 이벤트)
"""

import pandas as pd
import re
import os
import shutil
import argparse
import sys
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook

# ===== 상수 정의 =====
# 파일 경로 상수
LOG_FILE = 'ChatService.log'
NEW_LOG_PATH = r'E:\devSpace\ChatServiceTest\log\ChatService.log'

# 이벤트 타입 상수
EVENT_WAITING_START = 'WAITING_START'
EVENT_CRITICAL_ENTER = 'CRITICAL_ENTER'
EVENT_CRITICAL_LEAVE = 'CRITICAL_LEAVE'
EVENT_INCREMENT_BEFORE = 'INCREMENT_BEFORE'
EVENT_INCREMENT_AFTER = 'INCREMENT_AFTER'

# 결과 타입 상수
RESULT_SUCCESS = 'SUCCESS'
RESULT_FAIL_CAPACITY = 'FAIL_OVER_CAPACITY'
RESULT_PRE_CHECK_FAIL = 'PRE_CHECK_FAIL'  # 새로 추가
RESULT_UNKNOWN = 'UNKNOWN'

# 분석 구간 수
BINS_COUNT = 10

# 정규 표현식 패턴 (컴파일하여 성능 향상)
CRITICAL_PATTERN = re.compile(
    r'CRITICAL_SECTION_MARK tag=(?P<tag>WAITING_START|CRITICAL_ENTER|CRITICAL_LEAVE)'
    r' timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<event>\w+)'
    r'.* roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r'.* nanoTime=(?P<nanoTime>\d+)'
    r'.* epochNano=(?P<epochNano>\d+)'
)

INCREMENT_PATTERN = re.compile(
    r'timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<tag>INCREMENT_BEFORE|INCREMENT_AFTER)'
    r' roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r'.* epochNano=(?P<epochNano>\d+)'
    r'.* nanoTime=(?P<nanoTime>\d+)'
)

# PRE_CHECK_FAIL_OVER_CAPACITY 이벤트 패턴 추가
PRE_CHECK_FAIL_PATTERN = re.compile(
    r'timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<tag>PRE_CHECK_FAIL_OVER_CAPACITY)'
    r' roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r'.* epochNano=(?P<epochNano>\d+)'
    r'.* nanoTime=(?P<nanoTime>\d+)'
)


def replace_log_file() -> None:
    """
    기존 로그 파일을 새 로그 파일로 교체
    
    설명:
        분석 전에 최신 로그 파일로 교체하여
        항상 최신 데이터를 분석할 수 있도록 합니다.
    """
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    shutil.copy(NEW_LOG_PATH, LOG_FILE)
    print(f"로그 파일 교체 완료: {NEW_LOG_PATH} → {LOG_FILE}")


def parse_log_line(line: str) -> Optional[Dict[str, Any]]:
    """
    로그 라인을 파싱하여 이벤트 정보 추출
    
    Args:
        line: 로그 파일의 한 줄
    
    Returns:
        파싱된 이벤트 정보 딕셔너리 또는 None
    """
    # CRITICAL_SECTION_MARK 패턴 매칭
    match = CRITICAL_PATTERN.search(line)
    if match:
        data = match.groupdict()
        data['event_type'] = data['event']  # SUCCESS, FAIL_OVER_CAPACITY 등
        return data
    
    # INCREMENT 이벤트 패턴 매칭
    match = INCREMENT_PATTERN.search(line)
    if match:
        data = match.groupdict()
        data['event_type'] = None  # INCREMENT 이벤트는 event_type이 없음
        return data
    
    # PRE_CHECK_FAIL_OVER_CAPACITY 이벤트 패턴 매칭 (새로 추가)
    match = PRE_CHECK_FAIL_PATTERN.search(line)
    if match:
        data = match.groupdict()
        data['event_type'] = 'PRE_CHECK_FAIL_OVER_CAPACITY'  # 독립 이벤트 타입
        return data
    
    return None


def parse_five_events_clean(filepath: str, room_number: Optional[int] = None) -> pd.DataFrame:
    """
    로그 파일에서 5가지 이벤트 + PRE_CHECK_FAIL 이벤트를 파싱하여 DataFrame으로 변환
    
    Args:
        filepath: 로그 파일 경로
        room_number: 특정 방 번호만 필터링 (None이면 전체)
    
    Returns:
        파싱된 이벤트 데이터가 담긴 DataFrame
    
    특징:
        - 나노초 값을 문자열로 유지하여 정밀도 보존
        - 방 번호로 필터링 가능
        - PRE_CHECK_FAIL_OVER_CAPACITY 독립 이벤트 포함
    """
    records = []
    
    try:
        with open(filepath, encoding='utf-8') as f:
            for line in f:
                data = parse_log_line(line)
                
                if data:
                    # 방 번호를 정수로 변환
                    data['roomNumber'] = int(data['roomNumber'])
                    
                    # 나노초 값들을 문자열로 유지 (정밀도 보존)
                    # Python의 int는 임의 정밀도를 지원하지만,
                    # pandas/numpy는 64비트 제한이 있어 문자열로 보존
                    data['nanoTime'] = data['nanoTime']
                    data['epochNano'] = data['epochNano']
                    
                    # 방 번호 필터링
                    if room_number is None or data['roomNumber'] == room_number:
                        records.append(data)
    
    except FileNotFoundError:
        print(f"오류: 로그 파일을 찾을 수 없습니다 - {filepath}")
        return pd.DataFrame()
    except Exception as e:
        print(f"오류: 로그 파일 파싱 실패 - {e}")
        return pd.DataFrame()
    
    return pd.DataFrame(records)


def create_pre_check_fail_record(event_data: pd.Series, room_sequence: int) -> Dict[str, Any]:
    """
    PRE_CHECK_FAIL_OVER_CAPACITY 이벤트로부터 독립 레코드 생성
    
    Args:
        event_data: PRE_CHECK_FAIL_OVER_CAPACITY 이벤트 데이터
        room_sequence: 방별 진입 순번
    
    Returns:
        PRE_CHECK_FAIL 독립 레코드
    """
    record = {
        # 기본 정보
        'user_id': event_data['userId'],
        'roomNumber': event_data['roomNumber'],
        'room_entry_sequence': room_sequence,
        'join_result': RESULT_PRE_CHECK_FAIL,
        
        # 정렬용 시간 정보 (실제 PRE_CHECK_FAIL_OVER_CAPACITY 로그 시간)
        'waiting_start_time': event_data['timestamp'],
        'waiting_start_nanoTime': str(int(float(event_data['nanoTime']))),
        'waiting_start_epochNano': str(int(float(event_data['epochNano']))),
        
        # 나머지 모든 이벤트 컬럼들 - 빈 값 처리
        'waiting_start_event_type': '',
        'critical_enter_time': '',
        'critical_enter_nanoTime': '0',
        'critical_enter_epochNano': '0',
        'critical_enter_event_type': '',
        'critical_leave_time': '',
        'critical_leave_nanoTime': '0',
        'critical_leave_epochNano': '0',
        'critical_leave_event_type': '',
        'increment_before_time': '',
        'increment_before_nanoTime': '0',
        'increment_before_epochNano': '0',
        'increment_before_event_type': '',
        'increment_after_time': '',
        'increment_after_nanoTime': '0',
        'increment_after_epochNano': '0',
        'increment_after_event_type': ''
    }
    
    return record


def process_user_events(user_id: str, user_data: pd.DataFrame) -> Dict[str, Any]:
    """
    단일 사용자의 이벤트들을 처리하여 성능 프로필 생성
    
    Args:
        user_id: 사용자 ID
        user_data: 해당 사용자의 이벤트 데이터
    
    Returns:
        사용자의 성능 측정 프로필
    """
    # 이벤트를 시간순으로 정렬
    user_data = user_data.sort_values(['timestamp', 'nanoTime_num', 'epochNano_num']).reset_index(drop=True)
    
    # 이벤트별로 그룹화
    events = {}
    for _, row in user_data.iterrows():
        event_name = row['tag']
        events[event_name] = {
            'timestamp': row['timestamp'],
            'nanoTime': str(int(float(row['nanoTime']))),
            'epochNano': str(int(float(row['epochNano']))),
            'event_type': row.get('event_type')
        }
    
    # 기본 프로필 정보
    profile = {
        'user_id': user_id,
        'roomNumber': user_data.iloc[0]['roomNumber']
    }
    
    # 각 이벤트별 정보 추가
    event_mappings = {
        EVENT_WAITING_START: 'waiting_start',
        EVENT_CRITICAL_ENTER: 'critical_enter',
        EVENT_CRITICAL_LEAVE: 'critical_leave',
        EVENT_INCREMENT_BEFORE: 'increment_before',
        EVENT_INCREMENT_AFTER: 'increment_after'
    }
    
    for event_tag, prefix in event_mappings.items():
        if event_tag in events:
            event_data = events[event_tag]
            profile.update({
                f'{prefix}_time': event_data['timestamp'],
                f'{prefix}_nanoTime': event_data['nanoTime'],
                f'{prefix}_epochNano': event_data['epochNano']
            })
            
            # event_type이 있는 경우 추가
            if event_data['event_type'] is not None:
                profile[f'{prefix}_event_type'] = event_data['event_type']
    
    # join_result 설정 (CRITICAL_LEAVE의 event_type 기반)
    if EVENT_CRITICAL_LEAVE in events:
        leave_type = events[EVENT_CRITICAL_LEAVE]['event_type']
        if leave_type == RESULT_SUCCESS:
            profile['join_result'] = RESULT_SUCCESS
        elif leave_type == RESULT_FAIL_CAPACITY:
            profile['join_result'] = RESULT_FAIL_CAPACITY
        else:
            profile['join_result'] = RESULT_UNKNOWN
    
    return profile


def assign_bins(df: pd.DataFrame) -> pd.DataFrame:
    """
    각 방의 요청을 10개 구간으로 균등 분할
    
    Args:
        df: 전체 데이터 DataFrame
    
    Returns:
        bin 컬럼이 추가된 DataFrame
    
    설명:
        시간대별 성능 변화를 분석하기 위해
        각 방의 요청을 10개 구간으로 나눕니다.
    """
    df = df.copy()
    
    for room_num in df['roomNumber'].unique():
        room_mask = df['roomNumber'] == room_num
        room_data = df[room_mask]
        
        total_requests = len(room_data)
        
        # 요청이 10개 이하면 각각을 하나의 bin으로
        if total_requests <= BINS_COUNT:
            bins = range(1, total_requests + 1)
        else:
            # 10개 구간으로 균등 분할
            bins = pd.cut(range(total_requests), bins=BINS_COUNT, labels=range(1, BINS_COUNT + 1)).astype(int)
        
        df.loc[room_mask, 'bin'] = bins
    
    return df


def build_clean_performance_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    파싱된 로그 데이터를 성능 분석용 데이터로 변환
    
    Args:
        df: 파싱된 이벤트 DataFrame
    
    Returns:
        성능 분석용 DataFrame (기존 페어링 + PRE_CHECK_FAIL 독립 이벤트 포함)
    
    처리 과정:
        1. 타임스탬프 변환 및 정렬
        2. PRE_CHECK_FAIL_OVER_CAPACITY 독립 이벤트 분리 처리
        3. 나머지 5개 이벤트 사용자별 그룹화 및 페어링
        4. 방별 진입 순서 계산
        5. 독립 이벤트와 페어링 결과 병합
        6. 10개 구간으로 분할
    """
    if df.empty:
        return pd.DataFrame()
    
    # 타임스탬프를 datetime 객체로 변환
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    
    # 나노초 값들을 숫자로 변환 (정렬용)
    df['nanoTime_num'] = df['nanoTime'].astype(float)
    df['epochNano_num'] = df['epochNano'].astype(float)
    
    # 전체 데이터를 시간순 정렬
    df = df.sort_values(['timestamp', 'nanoTime_num', 'epochNano_num']).reset_index(drop=True)
    
    # PRE_CHECK_FAIL_OVER_CAPACITY 독립 이벤트와 나머지 이벤트 분리
    pre_check_fail_events = df[df['event_type'] == 'PRE_CHECK_FAIL_OVER_CAPACITY'].copy()
    other_events = df[df['event_type'] != 'PRE_CHECK_FAIL_OVER_CAPACITY'].copy()
    
    performance_results = []
    
    # 각 방별로 처리
    for room_num in df['roomNumber'].unique():
        room_sequence = 0
        
        # === 1. PRE_CHECK_FAIL 독립 이벤트 처리 ===
        room_pre_check_fail = pre_check_fail_events[pre_check_fail_events['roomNumber'] == room_num].copy()
        for _, pre_check_event in room_pre_check_fail.iterrows():
            room_sequence += 1
            pre_check_record = create_pre_check_fail_record(pre_check_event, room_sequence)
            performance_results.append(pre_check_record)
        
        # === 2. 기존 5개 이벤트 페어링 처리 ===
        room_other_events = other_events[other_events['roomNumber'] == room_num].copy()
        
        if not room_other_events.empty:
            # 사용자별로 그룹화하고 첫 이벤트 시간으로 정렬
            user_groups = []
            for user_id, user_data in room_other_events.groupby('userId'):
                first_event = user_data.iloc[0]
                user_groups.append((
                    first_event['timestamp'],
                    first_event['nanoTime_num'],
                    first_event['epochNano_num'],
                    user_id,
                    user_data
                ))
            
            # 첫 이벤트 시간순으로 정렬
            user_groups.sort(key=lambda x: (x[0], x[1], x[2]))
            
            # 각 사용자의 이벤트 처리
            for _, _, _, user_id, user_data in user_groups:
                room_sequence += 1
                
                profile = process_user_events(user_id, user_data)
                profile['room_entry_sequence'] = room_sequence
                
                performance_results.append(profile)
    
    if not performance_results:
        return pd.DataFrame()
    
    # DataFrame 생성 및 정리
    result_df = pd.DataFrame(performance_results)
    
    # 나노초 컬럼들을 문자열로 확실히 변환
    nano_columns = [col for col in result_df.columns if 'nanoTime' in col or 'epochNano' in col]
    for col in nano_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col].astype(str)
    
    # 구간(bin) 할당
    result_df = assign_bins(result_df)
    
    # 컬럼 순서 정리
    base_columns = ['roomNumber', 'bin', 'user_id', 'room_entry_sequence', 'join_result']
    event_columns = []
    
    # 이벤트별 컬럼들을 순서대로 추가
    for prefix in ['waiting_start', 'critical_enter', 'critical_leave', 'increment_before', 'increment_after']:
        for suffix in ['_time', '_nanoTime', '_epochNano', '_event_type']:
            col_name = prefix + suffix
            if col_name in result_df.columns:
                event_columns.append(col_name)
    
    # 존재하는 컬럼만 선택
    all_columns = base_columns + event_columns
    existing_columns = [col for col in all_columns if col in result_df.columns]
    result_df = result_df[existing_columns]
    
    # 최종 정렬
    result_df = sort_final_dataframe(result_df)
    
    return result_df


def sort_final_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    최종 DataFrame을 방 번호와 시간순으로 정렬
    
    Args:
        df: 정렬할 DataFrame
    
    Returns:
        정렬된 DataFrame
    """
    # 정렬을 위한 임시 숫자 컬럼 생성
    if 'waiting_start_time' in df.columns:
        df['waiting_start_nanoTime_num'] = pd.to_numeric(df['waiting_start_nanoTime'])
        df['waiting_start_epochNano_num'] = pd.to_numeric(df['waiting_start_epochNano'])
        
        df = df.sort_values([
            'roomNumber', 
            'waiting_start_time', 
            'waiting_start_nanoTime_num', 
            'waiting_start_epochNano_num'
        ])
        
        # 임시 컬럼 제거
        df = df.drop(columns=['waiting_start_nanoTime_num', 'waiting_start_epochNano_num'])
    elif 'critical_enter_time' in df.columns:
        df['critical_enter_nanoTime_num'] = pd.to_numeric(df['critical_enter_nanoTime'])
        df['critical_enter_epochNano_num'] = pd.to_numeric(df['critical_enter_epochNano'])
        
        df = df.sort_values([
            'roomNumber', 
            'critical_enter_time', 
            'critical_enter_nanoTime_num', 
            'critical_enter_epochNano_num'
        ])
        
        df = df.drop(columns=['critical_enter_nanoTime_num', 'critical_enter_epochNano_num'])
    else:
        df = df.sort_values(['roomNumber', 'room_entry_sequence'])
    
    return df.reset_index(drop=True)


def get_clean_event_desc_table() -> List[List[str]]:
    """
    Excel 파일에 추가할 컬럼 설명 테이블 생성
    
    Returns:
        컬럼 설명 테이블 (2차원 리스트)
    """
    return [
        ["속성명", "측정 목적", "도출 방법"],
        ["roomNumber", "방 번호 식별", "로그 필드: roomNumber"],
        ["bin", "방별 분석 구간", "각 방의 요청을 10개 구간으로 균등 분할"],
        ["user_id", "사용자 식별", "로그 필드: userId"],
        ["room_entry_sequence", "방별 처리 순번", "방별 타임스탬프 기준 순번"],
        ["join_result", "입장 결과 구분", "SUCCESS/FAIL_OVER_CAPACITY/PRE_CHECK_FAIL/UNKNOWN"],
        ["waiting_start_*", "대기 시작 시점", "WAITING_START 이벤트 속성들 (PRE_CHECK_FAIL은 정렬용만)"],
        ["critical_enter_*", "임계구역 진입 시점", "CRITICAL_ENTER 이벤트 속성들"],
        ["critical_leave_*", "임계구역 진출 시점", "CRITICAL_LEAVE 이벤트 속성들"],
        ["increment_before_*", "증가 작업 시작 시점", "INCREMENT_BEFORE 이벤트 속성들"],
        ["increment_after_*", "증가 작업 완료 시점", "INCREMENT_AFTER 이벤트 속성들"],
        ["*_time", "이벤트 발생 시간", "timestampIso (스레드 요청 순서)"],
        ["*_nanoTime", "나노초 정밀도 시간", "System.nanoTime()"],
        ["*_epochNano", "Epoch 나노초 시간", "Epoch 기준 나노초"],
        ["*_event_type", "이벤트 상세 타입", "SUCCESS/FAIL_OVER_CAPACITY 등"],
        ["PRE_CHECK_FAIL", "락 외부 사전 차단", "이중 확인 구조로 락 진입 전 차단된 요청"]
    ]


def convert_nano_value(x: Any) -> str:
    """
    나노초 값을 문자열로 안전하게 변환
    
    Args:
        x: 변환할 값
    
    Returns:
        문자열로 변환된 나노초 값
    """
    if pd.isna(x) or x == '' or str(x).lower() == 'nan':
        return ''
    try:
        return str(int(float(x)))
    except (ValueError, TypeError):
        return str(x)


def save_to_csv(df: pd.DataFrame, filepath: str) -> None:
    """
    DataFrame을 CSV 파일로 저장 (나노초 정밀도 유지)
    
    Args:
        df: 저장할 DataFrame
        filepath: 저장 경로
    """
    # 나노초 컬럼들을 문자열로 변환하여 지수 표기법 방지
    df_copy = df.copy()
    nano_columns = [col for col in df.columns if 'nanoTime' in col or 'epochNano' in col]
    
    for col in nano_columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(convert_nano_value)
    
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # UTF-8 with BOM으로 저장 (Excel에서 한글 깨짐 방지)
    df_copy.to_csv(filepath, index=False, encoding='utf-8-sig')
    print(f"CSV 파일 저장 완료: {filepath}")


def save_with_side_table(df: pd.DataFrame, filepath: str, desc_table: List[List[str]]) -> str:
    """
    DataFrame을 Excel로 저장하고 설명 테이블 추가
    
    Args:
        df: 저장할 DataFrame
        filepath: 파일 전체 경로
        desc_table: 설명 테이블
    
    Returns:
        저장된 파일 경로
    """
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # timezone 제거한 복사본 생성
    df_excel = df.copy()
    for col in df_excel.columns:
        if pd.api.types.is_datetime64_tz_dtype(df_excel[col]):
            df_excel[col] = df_excel[col].dt.tz_localize(None)
    
    # Excel 파일 생성
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False)
        
        # 나노초 컬럼들을 텍스트 형식으로 설정
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        nano_columns = [col for col in df.columns if 'nanoTime' in col or 'epochNano' in col]
        for col_name in nano_columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '@'  # 텍스트 형식
    
    # 설명 테이블 추가
    wb = load_workbook(filepath)
    ws = wb.active
    
    start_col = len(df.columns) + 2
    
    for i, row in enumerate(desc_table):
        for j, val in enumerate(row):
            ws.cell(row=i + 1, column=start_col + j, value=val)
    
    wb.save(filepath)
    print(f"Excel 파일 저장 완료: {filepath}")
    
    return filepath


def analyze_clean_results(df: pd.DataFrame) -> None:
    """
    처리 결과 분석 및 출력
    
    Args:
        df: 분석할 DataFrame
    """
    if df.empty:
        print("분석할 데이터가 없습니다.")
        return
    
    total_operations = len(df)
    
    print(f"\n{'='*60}")
    print(f"레이스 컨디션 지표 제거된 5개 이벤트 성능 측정 분석 결과")
    print(f"{'='*60}")
    print(f"전체 처리 작업 수: {total_operations}")
    
    # 입장 결과별 분석
    if 'join_result' in df.columns:
        print(f"\n[입장 결과별 분석]")
        join_result_counts = df['join_result'].value_counts()
        for result, count in join_result_counts.items():
            print(f"  {result}: {count}건 ({count/total_operations*100:.1f}%)")
    
    # 이벤트 완성도 분석
    events = ['waiting_start_time', 'critical_enter_time', 'critical_leave_time', 
              'increment_before_time', 'increment_after_time']
    
    print(f"\n[이벤트 완성도 분석]")
    for event in events:
        if event in df.columns:
            complete_count = df[event].notna().sum()
            print(f"  {event}: {complete_count}건 ({complete_count/total_operations*100:.1f}%)")
    
    # CRITICAL_LEAVE 이벤트 타입별 분석
    if 'critical_leave_event_type' in df.columns:
        print(f"\n[CRITICAL_LEAVE 이벤트 타입별 분석]")
        critical_leave_types = df['critical_leave_event_type'].value_counts()
        for event_type, count in critical_leave_types.items():
            if pd.notna(event_type):
                print(f"  {event_type}: {count}건")
    
    # PRE_CHECK_FAIL 분석 (새로 추가)
    pre_check_fail_count = len(df[df['join_result'] == RESULT_PRE_CHECK_FAIL])
    if pre_check_fail_count > 0:
        print(f"\n[PRE_CHECK_FAIL 분석]")
        print(f"  락 외부 사전 차단: {pre_check_fail_count}건 ({pre_check_fail_count/total_operations*100:.1f}%)")
        print(f"  이중 확인 구조 효과: 락 진입 전 {pre_check_fail_count}건의 불필요한 대기 방지")
    
    # 완전한 5개 이벤트 세션 분석
    complete_sessions = 0
    success_complete = 0
    fail_complete = 0
    
    for _, row in df.iterrows():
        # PRE_CHECK_FAIL은 독립 이벤트이므로 완전한 세션 분석에서 제외
        if row['join_result'] != RESULT_PRE_CHECK_FAIL:
            if all(pd.notna(row.get(event)) and row.get(event) != '' for event in events):
                complete_sessions += 1
                if row.get('join_result') == RESULT_SUCCESS:
                    success_complete += 1
                elif row.get('join_result') == RESULT_FAIL_CAPACITY:
                    fail_complete += 1
    
    paired_operations = total_operations - pre_check_fail_count
    if paired_operations > 0:
        print(f"\n[완전한 5개 이벤트 세션 분석] (PRE_CHECK_FAIL 제외)")
        print(f"  페어링 대상 세션: {paired_operations}건")
        print(f"  완전한 세션: {complete_sessions}건 ({complete_sessions/paired_operations*100:.1f}%)")
        print(f"    - 성공: {success_complete}건")
        print(f"    - 실패: {fail_complete}건")
    
    # 방별 성능 통계
    print(f"\n[방별 성능 통계]")
    room_stats = df.groupby('roomNumber').agg({
        'user_id': 'count',
        'bin': 'nunique'
    })
    room_stats.columns = ['total_operations', 'bin_count']
    
    for room_num, stats in room_stats.iterrows():
        print(f"  방 {room_num}: 총 {stats['total_operations']}회, 구간 수: {stats['bin_count']}")
        
        if 'join_result' in df.columns:
            room_data = df[df['roomNumber'] == room_num]
            room_results = room_data['join_result'].value_counts()
            for result, count in room_results.items():
                print(f"    - {result}: {count}건")
    
    # 나노초 정밀도 확인
    print(f"\n[나노초 정밀도 확인 (첫 3개 샘플)]")
    nano_columns = [col for col in df.columns if 'nanoTime' in col]
    if nano_columns:
        for i in range(min(3, len(df))):
            print(f"  샘플 {i+1}:")
            for col in nano_columns[:2]:
                if col in df.columns:
                    print(f"    {col}: {df.iloc[i][col]}")


def main():
    """
    메인 실행 함수
    
    명령줄 인자:
        --output_dir: 출력 디렉토리 경로
        --room: 특정 방 번호만 처리
        --csv: 추가 CSV 파일명
        --xlsx: Excel 파일명
    """
    # 명령줄 인자 파서 설정
    parser = argparse.ArgumentParser(
        description="레이스 컨디션 지표 제거된 5개 이벤트 성능 측정 전처리기 + PRE_CHECK_FAIL 독립 이벤트",
        epilog="예시: py -3 preprocess_logs.py --output_dir C:\\my_analysis\\ --room 1 --xlsx output.xlsx"
    )
    parser.add_argument('--output_dir', type=str, default='results', 
                        help='출력 디렉토리 경로 (기본값: results)')
    parser.add_argument('--room', type=int, help='특정 방 번호만 처리 (옵션)')
    parser.add_argument('--csv', type=str, help='추가 CSV 파일명 (옵션)')
    parser.add_argument('--xlsx', type=str, help='Excel 파일명 (옵션)')
    
    args = parser.parse_args()
    
    try:
        # 1. 로그 파일 교체
        replace_log_file()
        
        # 2. 로그 파싱
        print(f"\n로그 파일 파싱 중...")
        df = parse_five_events_clean(LOG_FILE, room_number=args.room)
        print(f"파싱 완료: {len(df)}개 이벤트")
        
        # 3. 성능 데이터 구축
        print(f"\n성능 데이터 구축 중...")
        result = build_clean_performance_data(df)
        print(f"구축 완료: {len(result)}개 세션")
        
        # 4. 출력 디렉토리 설정 및 생성
        output_dir = args.output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"출력 디렉토리 생성: {output_dir}")
        
        # 5. 기본 CSV 파일 저장
        if args.room:
            base_filename = f'clean_five_events_performance_room{args.room}.csv'
        else:
            base_filename = 'clean_five_events_performance_all_rooms.csv'
        
        csv_path = os.path.join(output_dir, base_filename)
        save_to_csv(result, csv_path)
        
        # 6. 추가 CSV 파일 저장 (옵션)
        if args.csv:
            additional_csv_path = os.path.join(output_dir, args.csv)
            save_to_csv(result, additional_csv_path)
        
        # 7. Excel 파일 저장 (옵션)
        if args.xlsx:
            xlsx_path = os.path.join(output_dir, args.xlsx)
            desc_table = get_clean_event_desc_table()
            save_with_side_table(result, xlsx_path, desc_table)
        
        # 8. 결과 분석 출력
        analyze_clean_results(result)
        
        print(f"\n{'='*60}")
        print(f"처리 완료!")
        print(f"출력 디렉토리: {os.path.abspath(output_dir)}")
        print(f"{'='*60}")
        
    except FileNotFoundError as e:
        print(f"\n오류: 파일을 찾을 수 없습니다 - {e}")
        sys.exit(1)
    except PermissionError as e:
        print(f"\n오류: 파일 접근 권한이 없습니다 - {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n예상치 못한 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# 스크립트가 직접 실행될 때만 main() 호출
if __name__ == '__main__':
    main(