#!/usr/bin/env python3
"""
Race Condition 분석기 (완전 수정 버전)
- 디버깅된 CSV 구조에 맞춰 완전히 새로 작성
- 4가지 규칙으로 동시성 문제를 탐지
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import load_workbook

def detect_race_condition_anomalies(df):
    """
    4가지 규칙으로 이상 현상을 탐지
    
    입력: df (DataFrame) - 분석할 데이터
    출력: (anomalies, detailed_analysis) - 이상 현상 목록과 상세 분석
    """
    
    print("🔍 이상 현상 탐지 시작...")
    
    # 결과 저장 리스트
    anomalies = []
    detailed_analysis = []
    
    # 시간 컬럼 변환
    df['prev_entry_time'] = pd.to_datetime(df['prev_entry_time'])
    df['curr_entry_time'] = pd.to_datetime(df['curr_entry_time'])
    
    print(f"분석 대상 방: {df['roomNumber'].unique()}")
    
    # 방별로 분석
    for room_num in df['roomNumber'].unique():
        print(f"  방 {room_num} 분석 중...")
        room_df = df[df['roomNumber'] == room_num].copy()
        
        # === 규칙 4를 위한 시간 순서 매핑 ===
        room_df_sorted = room_df.sort_values('curr_entry_time').reset_index(drop=True)
        sorted_sequence_map = {}
        for sorted_idx, row in room_df_sorted.iterrows():
            user_id = row['user_id']
            sorted_position = sorted_idx + 1
            sorted_sequence_map[user_id] = sorted_position
        
        # === 규칙 2를 위한 경합 그룹 찾기 ===
        contention_groups = find_contention_groups(room_df)
        
        # === 각 레코드 검사 ===
        for idx in room_df.index:
            row = room_df.loc[idx]
            anomaly_types = []
            anomaly_details = {}
            
            # 규칙 1: 값 불일치
            if pd.notna(row['expected_people']):
                expected_curr = min(row['expected_people'], row['max_people'])
                if row['curr_people'] != expected_curr:
                    anomaly_types.append('값 불일치')
                    anomaly_details['lost_update_expected'] = expected_curr
                    anomaly_details['lost_update_actual'] = row['curr_people']
                    anomaly_details['lost_update_diff'] = row['curr_people'] - expected_curr
            
            # 규칙 2: 경합 발생 자체
            user_id = row['user_id']
            if user_id in contention_groups:
                anomaly_types.append('경합 발생 자체')
                contention_info = contention_groups[user_id]
                anomaly_details['contention_group_size'] = contention_info['group_size']
                anomaly_details['contention_user_ids'] = ', '.join(contention_info['user_ids'])
            
            # 규칙 3: 정원 초과 오류
            if row['curr_people'] > row['max_people']:
                anomaly_types.append('정원 초과 오류')
                anomaly_details['over_capacity_amount'] = row['curr_people'] - row['max_people']
                anomaly_details['over_capacity_curr'] = row['curr_people']
                anomaly_details['over_capacity_max'] = row['max_people']
            
            # 규칙 4: 상태 전이 오류
            if user_id in sorted_sequence_map:
                sorted_position = sorted_sequence_map[user_id]
                expected_curr_people = 1 + sorted_position  # 초기 1명 + 순번
                
                if expected_curr_people <= row['max_people']:
                    if row['curr_people'] != expected_curr_people:
                        anomaly_types.append('상태 전이 오류')
                        anomaly_details['expected_curr_by_sequence'] = expected_curr_people
                        anomaly_details['actual_curr_people'] = row['curr_people']
                        anomaly_details['curr_sequence_diff'] = row['curr_people'] - expected_curr_people
                        anomaly_details['sorted_sequence_position'] = sorted_position
            
            # 임계구역 분석
            critical_analysis = analyze_critical_section(row, room_df, idx)
            anomaly_details.update(critical_analysis)
            
            # 이상 현상 발견 시 저장
            if anomaly_types:
                result_row = row.to_dict()
                result_row['anomaly_type'] = ', '.join(anomaly_types)
                
                for key, value in anomaly_details.items():
                    result_row[key] = value
                
                anomalies.append(result_row)
                
                # 상세 분석 텍스트 생성
                detailed_text = generate_analysis_text(row, anomaly_types, anomaly_details, room_num)
                detailed_analysis.append(detailed_text)
    
    print(f"✅ 이상 현상 탐지 완료: {len(anomalies)}건 발견")
    return anomalies, detailed_analysis

def find_contention_groups(room_df):
    """진짜 임계구역 기준 경합 그룹 찾기"""
    contention_groups = {}
    
    for i, row1 in room_df.iterrows():
        start1 = row1['prev_entry_time']
        end1 = row1['curr_entry_time']
        user1 = row1['user_id']
        
        if pd.isna(start1) or pd.isna(end1):
            continue
            
        overlapping_users = [user1]
        
        for j, row2 in room_df.iterrows():
            if i == j:
                continue
                
            start2 = row2['prev_entry_time']
            end2 = row2['curr_entry_time']
            user2 = row2['user_id']
            
            if pd.isna(start2) or pd.isna(end2):
                continue
            
            # 시간 겹침 확인
            if not (end1 < start2 or end2 < start1):
                overlapping_users.append(user2)
        
        # 2명 이상 겹치면 경합
        if len(overlapping_users) >= 2:
            for user_id in overlapping_users:
                contention_groups[user_id] = {
                    'group_size': len(overlapping_users),
                    'user_ids': overlapping_users
                }
    
    return contention_groups

def analyze_critical_section(base_row, room_df, base_idx):
    """임계구역 상세 분석"""
    critical_start = base_row['prev_entry_time']
    critical_end = base_row['curr_entry_time']
    
    if pd.isna(critical_start) or pd.isna(critical_end):
        return {}
    
    # 개입 사용자 찾기
    intervening_users = []
    for idx, other_row in room_df.iterrows():
        if idx == base_idx:
            continue
            
        other_start = other_row['prev_entry_time']
        other_end = other_row['curr_entry_time']
        
        if pd.isna(other_start) or pd.isna(other_end):
            continue
        
        if not (critical_end < other_start or other_end < critical_start):
            intervening_users.append(other_row['user_id'])
    
    result = {
        'true_critical_section_start': critical_start,
        'true_critical_section_end': critical_end,
        'true_critical_section_duration': (critical_end - critical_start).total_seconds(),
        'intervening_users_in_critical_section': ', '.join(intervening_users) if intervening_users else '',
        'intervening_user_count_critical': len(intervening_users)
    }
    
    # 나노초 정밀도 분석
    start_nano = base_row.get('true_critical_section_nanoTime_start')
    end_nano = base_row.get('true_critical_section_nanoTime_end')
    
    if not pd.isna(start_nano) and not pd.isna(end_nano):
        result['true_critical_section_duration_nanos'] = end_nano - start_nano
        
        start_epoch = base_row.get('true_critical_section_epochNano_start')
        end_epoch = base_row.get('true_critical_section_epochNano_end')
        
        if not pd.isna(start_epoch) and not pd.isna(end_epoch):
            result['true_critical_section_epoch_duration_nanos'] = end_epoch - start_epoch
    
    return result

def generate_analysis_text(row, anomaly_types, anomaly_details, room_num):
    """상세 분석 텍스트 생성"""
    text = f"""
================================================================================
경쟁 상태 이상 현상 분석 (수정된 요구사항 기반)
================================================================================
방 번호: {room_num}
사용자 ID: {row['user_id']}
Bin: {row['bin']}
발생 시각: {row['curr_entry_time']}
이상 현상 유형: {', '.join(anomaly_types)}
원본 방 입장 순번: {row['room_entry_sequence']}

기본 정보:
 이전 인원 (prev_people): {row['prev_people']}명
 현재 인원 (curr_people): {row['curr_people']}명
 최대 정원 (max_people): {row['max_people']}명

진짜 임계구역 정보:
 시작: {row['prev_entry_time']}
 끝: {row['curr_entry_time']}
 나노초 시작: {row.get('true_critical_section_nanoTime_start', 'N/A')}
 나노초 끝: {row.get('true_critical_section_nanoTime_end', 'N/A')}

상세 분석:"""

    for anomaly_type in anomaly_types:
        if anomaly_type == '값 불일치':
            text += f"""
 [규칙 1: 값 불일치 (Lost Update)]
  - 예상 결과: {anomaly_details.get('lost_update_expected', 'N/A')}명
  - 실제 결과: {anomaly_details.get('lost_update_actual', 'N/A')}명
  - 차이: {anomaly_details.get('lost_update_diff', 'N/A')}명
  → 다른 사용자의 작업으로 인해 의도한 갱신이 누락되거나 덮어쓰여짐"""
        
        elif anomaly_type == '경합 발생 자체':
            text += f"""
 [규칙 2: 경합 발생 자체 (Contention Detected)]
  - 경합 그룹 크기: {anomaly_details.get('contention_group_size', 'N/A')}개 사용자
  - 경합 사용자 ID: {anomaly_details.get('contention_user_ids', 'N/A')}
  → 진짜 임계구역이 1나노초라도 겹쳐 동시성 제어 부재 상황"""
        
        elif anomaly_type == '정원 초과 오류':
            text += f"""
 [규칙 3: 정원 초과 오류 (Capacity Exceeded Error)]
  - 최대 정원: {anomaly_details.get('over_capacity_max', 'N/A')}명
  - 실제 인원: {anomaly_details.get('over_capacity_curr', 'N/A')}명
  - 초과 인원: {anomaly_details.get('over_capacity_amount', 'N/A')}명
  → 비즈니스 규칙을 명백히 위반한 심각한 오류"""
        
        elif anomaly_type == '상태 전이 오류':
            pos = anomaly_details.get('sorted_sequence_position', 'N/A')
            text += f"""
 [규칙 4: 상태 전이 오류 (Stale Read / Inconsistent State)]
  - curr_entry_time 기준 정렬 순번: {pos}번째
  - 예상 curr_people: 1 + {pos} = {anomaly_details.get('expected_curr_by_sequence', 'N/A')}명
  - 실제 curr_people: {anomaly_details.get('actual_curr_people', 'N/A')}명
  - 차이: {anomaly_details.get('curr_sequence_diff', 'N/A')}명
  → 올바른 순서의 상태를 읽지 못하고 오염된 상태 값을 기준으로 작업"""

    duration = anomaly_details.get('true_critical_section_duration', 'N/A')
    duration_formatted = f"{duration:.6f}" if duration != 'N/A' else 'N/A'
    
    text += f"""

타이밍 상세 정보:
 진짜 임계구역 지속시간: {duration_formatted}초
 나노초 지속시간: {anomaly_details.get('true_critical_section_duration_nanos', 'N/A')}ns
 Epoch 나노초 지속시간: {anomaly_details.get('true_critical_section_epoch_duration_nanos', 'N/A')}ns
 개입 사용자: {anomaly_details.get('intervening_user_count_critical', 'N/A')}개
 개입 사용자 목록: {anomaly_details.get('intervening_users_in_critical_section', 'N/A') or '없음'}
"""
    
    return text

def print_statistics(df, anomaly_df):
    """분석 결과 통계 출력"""
    print("\n=== 🔧 경쟁 상태 탐지 결과 ===")
    print(f"전체 레코드 수: {len(df)}")
    print(f"이상 현상 발견 수: {len(anomaly_df)}")
    
    if len(df) > 0:
        print(f"이상 현상 비율: {len(anomaly_df)/len(df)*100:.2f}%")
    
    # 나노초 정밀도 통계
    if 'true_critical_section_nanoTime_start' in df.columns:
        nano_count = df['true_critical_section_nanoTime_start'].notna().sum()
        print(f"나노초 정밀도 데이터: {nano_count}건 ({nano_count/len(df)*100:.1f}%)")
    
    if len(anomaly_df) > 0:
        print("\n=== 4가지 규칙별 이상 현상 분포 ===")
        
        error_counts = {
            '값 불일치': 0,
            '경합 발생 자체': 0,
            '정원 초과 오류': 0,
            '상태 전이 오류': 0
        }
        
        for anomaly_str in anomaly_df['anomaly_type']:
            for error_type in error_counts.keys():
                if error_type in anomaly_str:
                    error_counts[error_type] += 1
        
        for error_type, count in error_counts.items():
            percentage = count/len(anomaly_df)*100 if len(anomaly_df) > 0 else 0
            print(f"  - {error_type}: {count}건 ({percentage:.1f}%)")

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description="Race Condition 분석기 (완전 수정 버전)")
    parser.add_argument('input_csv', help='입력 CSV 파일')
    parser.add_argument('output_csv', help='출력 CSV 파일 (이상 현상만)')
    parser.add_argument('--detailed_output', default='detailed_analysis.txt', help='상세 분석 텍스트 파일')
    parser.add_argument('--rooms', help='분석할 방 번호 (쉼표로 구분)')
    parser.add_argument('--xlsx_output', help='Excel 출력 파일 (선택사항)')
    
    args = parser.parse_args()
    
    try:
        print("🚀 Race Condition 분석기 시작...")
        
        # CSV 파일 읽기
        df = pd.read_csv(args.input_csv)
        print(f"✅ CSV 파일 읽기 완료: {len(df)}행, {len(df.columns)}컬럼")
        
        # 필수 컬럼 확인
        required_columns = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 'expected_people', 
                           'max_people', 'prev_entry_time', 'curr_entry_time', 'room_entry_sequence']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"❌ 필수 컬럼 누락: {missing_columns}")
            return
        
        print("✅ 필수 컬럼 확인 완료")
        
        # 방 번호 필터링
        if args.rooms:
            room_numbers = [int(room.strip()) for room in args.rooms.split(',')]
            df = df[df['roomNumber'].isin(room_numbers)]
            print(f"🔍 방 번호 {room_numbers}로 필터링: {len(df)} 레코드")
        
        # 메인 분석 실행
        anomalies, detailed_analysis = detect_race_condition_anomalies(df)
        
        # 결과 저장
        if anomalies:
            anomaly_df = pd.DataFrame(anomalies)
            
            # 컬럼 정리
            basic_cols = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 'expected_people', 
                         'max_people', 'prev_entry_time', 'curr_entry_time', 
                         'true_critical_section_nanoTime_start', 'true_critical_section_epochNano_start',
                         'true_critical_section_nanoTime_end', 'true_critical_section_epochNano_end',
                         'anomaly_type', 'room_entry_sequence']
            
            detail_cols = ['lost_update_expected', 'lost_update_actual', 'lost_update_diff',
                          'contention_group_size', 'contention_user_ids',
                          'over_capacity_amount', 'over_capacity_curr', 'over_capacity_max',
                          'expected_curr_by_sequence', 'actual_curr_people', 'curr_sequence_diff',
                          'sorted_sequence_position',
                          'true_critical_section_start', 'true_critical_section_end', 'true_critical_section_duration',
                          'intervening_users_in_critical_section', 'intervening_user_count_critical',
                          'true_critical_section_duration_nanos', 'true_critical_section_epoch_duration_nanos']
            
            existing_cols = [col for col in basic_cols + detail_cols if col in anomaly_df.columns]
            anomaly_df = anomaly_df[existing_cols]
            anomaly_df = anomaly_df.fillna('')
            
            # CSV 저장
            anomaly_df.to_csv(args.output_csv, index=False, encoding='utf-8-sig')
            print(f"💾 이상 현상 {len(anomaly_df)}개가 {args.output_csv}에 저장되었습니다.")
            
            # Excel 저장 (선택사항)
            if args.xlsx_output:
                desc_table = [
                    ["규칙", "조건", "의미"],
                    ["규칙 1: 값 불일치", "curr_people ≠ expected_people", "다른 사용자 작업으로 의도한 갱신이 누락/덮어쓰여짐"],
                    ["규칙 2: 경합 발생 자체", "진짜 임계구역이 1나노초라도 겹침", "동시성 제어 부재로 잠재적 위험 노출"],
                    ["규칙 3: 정원 초과 오류", "curr_people > max_people", "비즈니스 규칙을 명백히 위반한 심각한 오류"],
                    ["규칙 4: 상태 전이 오류", "curr_people ≠ 1+curr_entry_time기준순번", "올바른 상태를 읽지 못하고 오염된 상태로 작업"]
                ]
                
                anomaly_df.to_excel(args.xlsx_output, index=False)
                wb = load_workbook(args.xlsx_output)
                ws = wb.active
                
                for i, row in enumerate(desc_table):
                    for j, val in enumerate(row):
                        ws.cell(row=4+i, column=10+j, value=val)
                
                wb.save(args.xlsx_output)
                print(f"📊 Excel 파일도 저장됨: {args.xlsx_output}")
        
        else:
            anomaly_df = pd.DataFrame()
            print("✅ 이상 현상이 발견되지 않았습니다.")
        
        # 상세 분석 텍스트 저장
        with open(args.detailed_output, 'w', encoding='utf-8') as f:
            f.write("Race Condition 상세 분석 결과\n")
            f.write(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"입력 파일: {args.input_csv}\n")
            f.write(f"총 이상 현상 수: {len(anomalies)}\n\n")
            
            for detailed_text in detailed_analysis:
                f.write(detailed_text)
                f.write("\n")
        
        print(f"📄 상세 분석 결과 저장: {args.detailed_output}")
        
        # 통계 출력
        print_statistics(df, anomaly_df)
        
        print("\n🎉 분석 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()