import pandas as pd  # 데이터 처리를 위한 라이브러리
import re             # 정규 표현식을 위한 라이브러리
import os             # 운영체제 기능을 위한 라이브러리
import shutil         # 파일 복사/이동을 위한 라이브러리
import argparse       # 명령줄 인자 처리를 위한 라이브러리
from openpyxl import load_workbook  # Excel 파일 처리를 위한 라이브러리

# 상수 정의
LOG_FILE = 'ChatService.log'  # 기본 로그 파일명
NEW_LOG_PATH = r'E:\devSpace\ChatServiceTest\log\ChatService.log'  # 새 로그 파일 경로

def replace_log_file():
    """
    로그 파일을 새로운 버전으로 교체하는 함수
    - 기존 로그 파일이 있으면 삭제
    - 새 로그 파일을 복사해서 가져옴
    """
    # 기존 로그 파일이 존재하면 삭제
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    
    # 새 로그 파일을 현재 디렉토리로 복사
    shutil.copy(NEW_LOG_PATH, LOG_FILE)

def parse_logs(filepath, room_number=None):
    """
    로그 파일을 파싱해서 핵심 이벤트들을 추출하는 함수
    
    🔧 수정된 요구사항에 따라:
    - 진짜 임계구역: PRE_JOIN_CURRENT_STATE ~ JOIN_SUCCESS_EXISTING
    - 규칙 2 경합 탐지: 진짜 임계구역 기준으로 변경
    
    입력:
    - filepath: 로그 파일 경로
    - room_number: 특정 방 번호만 필터링 (None이면 모든 방)
    
    출력: DataFrame - 파싱된 이벤트 데이터
    """
    
    # 정규 표현식 패턴 정의
    # 핵심 이벤트들만 추출 (요구사항 수정에 따라 INCREMENT 이벤트 불필요):
    # - PRE_JOIN_CURRENT_STATE: 진짜 임계구역 시작
    # - JOIN_SUCCESS_EXISTING: 진짜 임계구역 끝 (성공)
    # - JOIN_FAIL_OVER_CAPACITY_EXISTING: 진짜 임계구역 끝 (실패)
    pattern = re.compile(
        r'timestampIso=(?P<timestamp>\S+).*?'  # 시간 정보 추출
        r'event=(?P<event>PRE_JOIN_CURRENT_STATE|JOIN_SUCCESS_EXISTING|JOIN_FAIL_OVER_CAPACITY_EXISTING).*?'  # 핵심 이벤트만
        r'roomNumber=(?P<roomNumber>\d+).*?'    # 방 번호
        r'userId=(?P<userId>\S+).*?'            # 사용자 ID
        r'currentPeople=(?P<currentPeople>\d+).*?'  # 현재 인원수
        r'maxPeople=(?P<maxPeople>\d+)'         # 최대 정원
    )
    
    records = []  # 파싱된 레코드들을 저장할 리스트
    
    # 로그 파일을 한 줄씩 읽으면서 파싱
    with open(filepath, encoding='utf-8') as f:
        for line in f:
            # 정규식으로 매칭 시도
            match = pattern.search(line)
            if match:
                # 매칭된 그룹들을 딕셔너리로 변환
                data = match.groupdict()
                
                # 문자열로 추출된 숫자들을 정수형으로 변환
                data['roomNumber'] = int(data['roomNumber'])
                data['currentPeople'] = int(data['currentPeople'])
                data['maxPeople'] = int(data['maxPeople'])
                
                # 🔧 나노초 정밀도 정보 추출 (진짜 임계구역 정밀도 향상)
                # 요구사항 수정으로 INCREMENT 이벤트는 불필요하지만, 
                # 진짜 임계구역의 나노초 정밀도는 여전히 유용
                nano_match = re.search(r'nanoTime=(\d+)', line)
                epoch_match = re.search(r'epochNano=(\d+)', line)
                
                if nano_match:
                    data['nanoTime'] = int(nano_match.group(1))
                if epoch_match:
                    data['epochNano'] = int(epoch_match.group(1))
                
                # 방 번호 필터링 적용
                if room_number is None or data['roomNumber'] == room_number:
                    records.append(data)
    
    # 리스트를 DataFrame으로 변환해서 반환
    return pd.DataFrame(records)

def build_paired_data_true_critical_section(df):
    """
    🔧 수정된 요구사항에 따른 페어링 로직 함수
    - 진짜 임계구역 기준: PRE_JOIN_CURRENT_STATE ~ JOIN_SUCCESS_EXISTING
    - 규칙 2 경합 탐지: 진짜 임계구역 겹침 기준으로 변경
    
    입력: df (DataFrame) - 파싱된 로그 데이터
    출력: DataFrame - 페어링된 입장 요청 데이터
    """
    
    # 빈 데이터면 빈 DataFrame 반환
    if df.empty:
        return pd.DataFrame()
    
    # === 핵심 이벤트별로 데이터 분리 ===
    pre = df[df['event'] == 'PRE_JOIN_CURRENT_STATE'].copy()          # 진짜 임계구역 시작
    success = df[df['event'] == 'JOIN_SUCCESS_EXISTING'].copy()       # 진짜 임계구역 끝 (성공)
    fail = df[df['event'] == 'JOIN_FAIL_OVER_CAPACITY_EXISTING'].copy() # 진짜 임계구역 끝 (실패)

    # === 페어링을 위한 인덱스 생성 ===
    # 같은 방, 같은 사용자의 이벤트들을 순서대로 연결하기 위한 인덱스
    for event_df in [pre, success, fail]:
        if not event_df.empty:
            # 방 번호와 사용자 ID별로 그룹화해서 순차 번호 부여
            event_df['pair_idx'] = event_df.groupby(['roomNumber', 'userId']).cumcount()

    # === 성공 케이스 페어링 ===
    paired_success = pd.DataFrame()
    if not pre.empty and not success.empty:
        # roomNumber, userId, pair_idx가 같은 레코드들을 연결
        paired_success = pd.merge(pre, success, on=['roomNumber', 'userId', 'pair_idx'], suffixes=('_pre', '_suc'))
        
        # 성공 케이스 결과 필드 생성
        paired_success['join_result'] = 'SUCCESS'
        paired_success['curr_people'] = paired_success['currentPeople_suc']
        paired_success['expected_people'] = paired_success['currentPeople_pre'] + 1

    # === 실패 케이스 페어링 ===
    paired_fail = pd.DataFrame()
    if not pre.empty and not fail.empty:
        # roomNumber, userId, pair_idx가 같은 레코드들을 연결
        paired_fail = pd.merge(pre, fail, on=['roomNumber', 'userId', 'pair_idx'], suffixes=('_pre', '_fail'))
        
        # 실패 케이스 결과 필드 생성
        paired_fail['join_result'] = 'FAIL_OVER_CAPACITY'
        paired_fail['curr_people'] = paired_fail['currentPeople_fail']
        paired_fail['expected_people'] = None  # 실패했으므로 예상 인원수 없음

    # === 성공과 실패 케이스 통합 ===
    result = pd.concat([paired_success, paired_fail], ignore_index=True)
    
    if result.empty:
        return pd.DataFrame()
    
    # === 방별 시간 순 정렬 및 순번 부여 ===
    result['timestamp_pre'] = pd.to_datetime(result['timestamp_pre'])
    result = result.sort_values(['roomNumber', 'timestamp_pre']).reset_index(drop=True)
    result['room_entry_sequence'] = result.groupby('roomNumber').cumcount() + 1
    
    # === 분석을 위한 구간 생성 ===
    result['bin'] = pd.cut(range(len(result)), bins=10, labels=range(1, 11))
    
    # === 🔧 수정된 최종 컬럼 정리 ===
    # 원하는 순서: roomNumber, bin, user_id, prev_people, curr_people, expected_people, max_people, room_entry_sequence, join_result, prev_entry_time, curr_entry_time, true_critical_section_nanoTime_start, true_critical_section_epochNano_start, true_critical_section_nanoTime_end, true_critical_section_epochNano_end
    final_columns = {
        'roomNumber': 'roomNumber',
        'bin': 'bin',
        'userId': 'user_id',
        'currentPeople_pre': 'prev_people',
        'curr_people': 'curr_people',
        'expected_people': 'expected_people',
        'maxPeople_pre': 'max_people',
        'room_entry_sequence': 'room_entry_sequence',
        'join_result': 'join_result',
        'timestamp_pre': 'prev_entry_time'
    }
    
    # === 시간 필드 처리 ===
    if 'timestamp_suc' in result.columns:
        final_columns['timestamp_suc'] = 'curr_entry_time'
    elif 'timestamp_fail' in result.columns:
        final_columns['timestamp_fail'] = 'curr_entry_time'
    
    # === 🔧 진짜 임계구역 나노초 정밀도 필드 추가 ===
    # 진짜 임계구역의 나노초 정밀도 시작점
    if 'nanoTime_pre' in result.columns:
        final_columns['nanoTime_pre'] = 'true_critical_section_nanoTime_start'
    if 'epochNano_pre' in result.columns:
        final_columns['epochNano_pre'] = 'true_critical_section_epochNano_start'
    
    # 진짜 임계구역의 나노초 정밀도 끝점
    if 'nanoTime_suc' in result.columns:
        final_columns['nanoTime_suc'] = 'true_critical_section_nanoTime_end'
    elif 'nanoTime_fail' in result.columns:
        final_columns['nanoTime_fail'] = 'true_critical_section_nanoTime_end'
    
    if 'epochNano_suc' in result.columns:
        final_columns['epochNano_suc'] = 'true_critical_section_epochNano_end'
    elif 'epochNano_fail' in result.columns:
        final_columns['epochNano_fail'] = 'true_critical_section_epochNano_end'
    
    # === 최종 컬럼 선택 및 이름 변경 ===
    # 실제로 존재하는 컬럼들만 선택
    existing_columns = {old: new for old, new in final_columns.items() if old in result.columns}
    
    # 원하는 순서대로 컬럼 정렬
    desired_order = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 'expected_people', 'max_people', 
                     'room_entry_sequence', 'join_result', 'prev_entry_time', 'curr_entry_time',
                     'true_critical_section_nanoTime_start', 'true_critical_section_epochNano_start',
                     'true_critical_section_nanoTime_end', 'true_critical_section_epochNano_end']
    
    # DataFrame 재구성 (컬럼명 변경 후 순서 정렬)
    result = result[list(existing_columns.keys())].rename(columns=existing_columns)
    
    # 존재하는 컬럼들만 원하는 순서로 재배열
    final_order = [col for col in desired_order if col in result.columns]
    result = result[final_order]
    
    return result

def get_true_critical_section_desc_table():
    """
    🔧 수정된 요구사항에 따른 설명 테이블 생성 함수
    - 진짜 임계구역 기준 필드 설명
    - 규칙 2 경합 탐지: 진짜 임계구역 겹침 기준
    """
    return [
        ["속성명", "분석 목적", "도출 방법"],
        ["roomNumber", "방 번호 식별", "로그 필드: roomNumber"],
        ["bin", "분석 구간 구분", "전체 데이터를 10등분"],
        ["user_id", "사용자 식별", "로그 필드: userId"],
        ["prev_people", "입장 전 인원수", "PRE_JOIN_CURRENT_STATE의 currentPeople"],
        ["curr_people", "입장 후 인원수", "SUCCESS/FAIL 이벤트의 currentPeople"],
        ["expected_people", "기대 인원수", "prev_people + 1 (성공시)"],
        ["max_people", "최대 정원", "로그 필드: maxPeople"],
        ["room_entry_sequence", "방별 입장 순번", "prev_entry_time 기준 방별 순번"],
        ["join_result", "입장 결과", "SUCCESS 또는 FAIL_OVER_CAPACITY"],
        ["prev_entry_time", "진짜 임계구역 시작", "PRE_JOIN_CURRENT_STATE 타임스탬프"],
        ["curr_entry_time", "진짜 임계구역 끝", "SUCCESS/FAIL 타임스탬프"],
        ["true_critical_section_nanoTime_start", "진짜 임계구역 시작 나노초", "PRE_JOIN_CURRENT_STATE의 nanoTime"],
        ["true_critical_section_epochNano_start", "진짜 임계구역 시작 Epoch 나노초", "PRE_JOIN_CURRENT_STATE의 epochNano"],
        ["true_critical_section_nanoTime_end", "진짜 임계구역 끝 나노초", "SUCCESS/FAIL의 nanoTime"],
        ["true_critical_section_epochNano_end", "진짜 임계구역 끝 Epoch 나노초", "SUCCESS/FAIL의 epochNano"]
    ]

def save_with_side_table(df_result, out_xlsx, desc_table):
    """
    Excel 파일에 데이터와 설명 테이블을 함께 저장하는 함수
    """
    df_result.to_excel(out_xlsx, index=False)
    
    wb = load_workbook(out_xlsx)
    ws = wb.active
    
    start_col = len(df_result.columns) + 2
    
    for i, row in enumerate(desc_table):
        for j, val in enumerate(row):
            ws.cell(row=i + 1, column=start_col + j, value=val)
    
    wb.save(out_xlsx)

def analyze_results(df):
    """
    처리된 데이터에 대한 간단한 통계 분석을 수행하는 함수
    """
    if df.empty:
        print("분석할 데이터가 없습니다.")
        return
    
    # 기본 통계 계산
    total_requests = len(df)
    success_count = len(df[df['join_result'] == 'SUCCESS'])
    fail_count = len(df[df['join_result'] == 'FAIL_OVER_CAPACITY'])
    
    print(f"\n=== 🔧 수정된 요구사항 기반 전처리 결과 ===")
    print(f"전체 입장 요청 수: {total_requests}")
    print(f"성공: {success_count}건 ({success_count/total_requests*100:.1f}%)")
    print(f"실패: {fail_count}건 ({fail_count/total_requests*100:.1f}%)")
    
    # 레이스 컨디션 예비 분석
    if success_count > 0:
        valid_success = df[(df['join_result'] == 'SUCCESS') & df['expected_people'].notna()]
        if not valid_success.empty:
            race_conditions = valid_success[valid_success['curr_people'] != valid_success['expected_people']]
            race_count = len(race_conditions)
            print(f"잠재적 레이스 컨디션: {race_count}건 ({race_count/success_count*100:.1f}%)")
    
    # 진짜 임계구역 나노초 정밀도 통계
    if 'true_critical_section_nanoTime_start' in df.columns and 'true_critical_section_nanoTime_end' in df.columns:
        nano_start_count = df['true_critical_section_nanoTime_start'].notna().sum()
        nano_end_count = df['true_critical_section_nanoTime_end'].notna().sum()
        print(f"진짜 임계구역 나노초 정밀도: 시작 {nano_start_count}건, 끝 {nano_end_count}건")
    
    # 방별 통계
    room_stats = df.groupby('roomNumber').agg({
        'user_id': 'count',
        'join_result': lambda x: (x == 'SUCCESS').sum()
    }).rename(columns={'user_id': 'total_requests', 'join_result': 'success_count'})
    
    print(f"\n=== 방별 통계 ===")
    for room_num, stats in room_stats.iterrows():
        success_rate = stats['success_count'] / stats['total_requests'] * 100
        print(f"방 {room_num}: 총 {stats['total_requests']}건, 성공 {stats['success_count']}건 ({success_rate:.1f}%)")

def main():
    """
    🔧 수정된 요구사항 기반 메인 함수
    """
    parser = argparse.ArgumentParser(description="Race Condition 이벤트 전처리기 (수정된 요구사항 기반)")
    parser.add_argument('--room', type=int, help='특정 방 번호만 처리 (옵션)')
    parser.add_argument('--csv', type=str, help='추가 CSV 파일명 (옵션)')
    parser.add_argument('--xlsx', type=str, help='Excel 파일명 (옵션)')
    
    args = parser.parse_args()
    
    try:
        print("🔧 수정된 요구사항 기반 Race Condition 이벤트 전처리기 시작...")
        print("📋 규칙 2 경합 탐지: 진짜 임계구역 (PRE_JOIN_CURRENT_STATE ~ JOIN_SUCCESS_EXISTING) 기준")
        
        # 1단계: 로그 파일 교체
        print("1. 로그 파일 교체 중...")
        replace_log_file()
        
        # 2단계: 로그 파싱 (진짜 임계구역 이벤트)
        print("2. 진짜 임계구역 이벤트 파싱 중...")
        df = parse_logs(LOG_FILE, room_number=args.room)
        print(f"   파싱된 이벤트 수: {len(df)}")
        
        # 3단계: 진짜 임계구역 기준 페어링
        print("3. 진짜 임계구역 기준 페어링 처리 중...")
        result = build_paired_data_true_critical_section(df)
        print(f"   페어링된 요청 수: {len(result)}")
        
        # 4단계: 결과 저장
        print("4. 결과 저장 중...")
        
        # 기본 CSV 파일 저장
        if args.room:
            base_filename = f'racecondition_event_preprocessor_result_room{args.room}.csv'
        else:
            base_filename = 'racecondition_event_preprocessor_result.csv'
        
        result.to_csv(base_filename, index=False, encoding='utf-8-sig')
        print(f"   CSV 저장 완료: {base_filename}")
        
        # 추가 파일 저장
        if args.csv:
            result.to_csv(args.csv, index=False, encoding='utf-8-sig')
            print(f"   추가 CSV 저장 완료: {args.csv}")
        
        if args.xlsx:
            desc_table = get_true_critical_section_desc_table()
            save_with_side_table(result, args.xlsx, desc_table)
            print(f"   Excel 저장 완료: {args.xlsx}")
        
        # 5단계: 결과 분석
        print("5. 결과 분석 중...")
        analyze_results(result)
        
        print("\n✅ 수정된 요구사항 기반 전처리 완료!")
        print("🎯 진짜 임계구역 나노초 정밀도 필드 추가 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()