#!/usr/bin/env python3
"""
성능 분석 지표 계산 스크립트 (v6.0) - 나노초 단위 분석
- 모든 시간 계산을 나노초 단위로 유지
- 통계 출력도 나노초 단위로 표시
"""

import pandas as pd
import numpy as np
import os
import argparse
from datetime import datetime
import sys
from openpyxl import load_workbook
from decimal import Decimal, ROUND_HALF_UP

def calculate_rate(count, total):
    """전체 대비 비율 계산 (0으로 나누기 방지)"""
    if isinstance(count, pd.Series) or isinstance(total, pd.Series):
        return np.where(total != 0, (count / total) * 100, 0)
    return (count / total) * 100 if total > 0 else 0

def get_stats_with_precision(series, metric_name="", unit="ns"):
    """시리즈의 통계값 계산 (Mean, Median, Max) - 나노초 기준
    
    Args:
        series: 통계를 계산할 시리즈 (나노초 단위)
        metric_name: 메트릭 이름
        unit: 표시 단위 ("ns", "us", "ms")
    """
    if len(series) == 0:
        return {
            'Mean': 0.0,
            'Median': 0.0,
            'Max': 0.0
        }
    
    # NaN이 아닌 값들만 사용
    valid_series = series.dropna()
    if len(valid_series) == 0:
        return {
            'Mean': 0.0,
            'Median': 0.0,
            'Max': 0.0
        }
    
    # 단위 변환 계수 (나노초 기준)
    conversion = {
        'ns': 1.0,              # 나노초 (기본)
        'us': 0.001,            # 마이크로초
        'ms': 0.000001          # 밀리초
    }
    factor = conversion.get(unit, 1.0)
    
    # 기본 통계
    mean_val = float(valid_series.mean())
    median_val = float(valid_series.median())
    max_val = float(valid_series.max())
    
    # 로그 출력 (나노초 단위)
    print(f"    - {metric_name} 통계:")
    print(f"      Mean={mean_val:,.0f}ns, Median={median_val:,.0f}ns, Max={max_val:,.0f}ns (샘플: {len(valid_series)}개)")
    
    return {
        'Mean': mean_val * factor,
        'Median': median_val * factor,
        'Max': max_val * factor
    }

def parse_nano_time_precise(nano_str):
    """나노초 문자열을 정확하게 파싱 (정밀도 유지)"""
    if pd.isna(nano_str) or nano_str == 'nan':
        return np.nan
    
    try:
        # 문자열을 Decimal로 변환하여 정밀도 유지
        if isinstance(nano_str, str):
            # 지수 표기법을 일반 숫자로 변환
            if 'E' in nano_str or 'e' in nano_str:
                decimal_val = Decimal(nano_str)
                # 정수로 변환하여 정밀도 유지
                return int(decimal_val)
            else:
                return int(float(nano_str))
        else:
            return int(nano_str)
    except (ValueError, TypeError) as e:
        print(f"      경고: 나노초 파싱 실패 - {nano_str}: {e}")
        return np.nan

def calculate_time_diff_nano(start_nano, end_nano):
    """나노초 값 간의 시간 차이를 나노초로 반환"""
    if pd.isna(start_nano) or pd.isna(end_nano):
        return np.nan
    
    try:
        # 정수 연산으로 정밀도 유지 (나노초 단위 유지)
        diff_nano = int(end_nano) - int(start_nano)
        return diff_nano
    except (ValueError, TypeError) as e:
        return np.nan

def set_join_result_from_events(df):
    """이벤트 타입 기반으로 join_result 컬럼을 설정"""
    def determine_join_result(row):
        """개별 행의 join_result를 결정"""
        # 이미 join_result가 설정되어 있고 유효한 값이면 그대로 사용
        if pd.notna(row.get('join_result')) and row.get('join_result') != '':
            existing_result = str(row['join_result']).strip()
            if existing_result in ['SUCCESS', 'FAIL_OVER_CAPACITY', 'FAIL_ENTRY']:
                return existing_result
        
        # critical_leave_event_type 기반으로 판단
        if pd.isna(row.get('critical_leave_event_type')) or row.get('critical_leave_event_type') == '':
            return 'FAIL_ENTRY'
        elif str(row['critical_leave_event_type']).strip() == 'SUCCESS':
            return 'SUCCESS'
        elif str(row['critical_leave_event_type']).strip() == 'FAIL_OVER_CAPACITY':
            return 'FAIL_OVER_CAPACITY'
        else:
            return 'UNKNOWN'
    
    print("  - join_result 컬럼 설정 중...")
    df_result = df.copy()
    df_result['join_result'] = df_result.apply(determine_join_result, axis=1)
    
    # 분포 확인
    new_counts = df_result['join_result'].value_counts(dropna=False)
    print(f"    join_result 분포: {dict(new_counts)}")
    
    return df_result

def process_performance_data(csv_path, label):
    """단일 CSV 파일을 처리하여 성능 통계를 계산하고 Excel로 저장"""
    print(f"\n처리 중: {csv_path} (레이블: {label})")
    
    # 1. 원본 데이터 로드
    try:
        # 나노초 컬럼들을 문자열로 읽어서 정밀도 유지
        dtype_spec = {
            'waiting_start_nanoTime': str,
            'critical_enter_nanoTime': str,
            'critical_leave_nanoTime': str,
            'increment_before_nanoTime': str,
            'increment_after_nanoTime': str
        }
        
        df_total = pd.read_csv(csv_path, dtype=dtype_spec)
        print(f"  - 총 {len(df_total)}개의 레코드 로드됨")
        
        # 나노초 컬럼들을 정수로 변환 (정밀도 유지)
        nano_columns = [col for col in dtype_spec.keys() if col in df_total.columns]
        for col in nano_columns:
            df_total[col] = df_total[col].apply(parse_nano_time_precise)
        
    except FileNotFoundError:
        print(f"오류: CSV 파일을 찾을 수 없습니다 - {csv_path}")
        return False
    except Exception as e:
        print(f"오류: CSV 파일 로드 실패 - {e}")
        return False
    
    # 1.5. join_result 컬럼 설정
    df_total = set_join_result_from_events(df_total)
    
    # 2. 데이터 그룹 분류
    df_success = df_total[df_total['join_result'] == 'SUCCESS'].copy()
    df_lock_failed = df_total[df_total['join_result'] == 'FAIL_ENTRY'].copy()
    df_capacity_failed = df_total[df_total['join_result'] == 'FAIL_OVER_CAPACITY'].copy()
    
    print(f"  - 성공: {len(df_success)}, 진입실패: {len(df_lock_failed)}, 정원초과실패: {len(df_capacity_failed)}")
    
    # 3. 성공 그룹의 시간차 계산 (나노초 단위)
    if len(df_success) > 0:
        print("  - 성공 그룹 시간 계산 중...")
        
        # 나노초 기반 계산
        df_success['wait_time_ns'] = df_success.apply(
            lambda row: calculate_time_diff_nano(
                row['waiting_start_nanoTime'], 
                row['critical_enter_nanoTime']
            ),
            axis=1
        )
        
        df_success['dwell_time_ns'] = df_success.apply(
            lambda row: calculate_time_diff_nano(
                row['critical_enter_nanoTime'], 
                row['critical_leave_nanoTime']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링
        valid_success = df_success[
            (df_success['wait_time_ns'].notna()) & 
            (df_success['dwell_time_ns'].notna()) &
            (df_success['wait_time_ns'] >= 0) & 
            (df_success['dwell_time_ns'] >= 0)
        ]
        
        print(f"  - 성공 그룹 시간 계산 완료 (유효 데이터: {len(valid_success)}개/{len(df_success)}개)")
        df_success = valid_success
    
    # 4. 정원초과 실패 그룹의 시간차 계산 (나노초 단위)
    if len(df_capacity_failed) > 0:
        print("  - 정원초과 실패 그룹 시간 계산 중...")
        
        df_capacity_failed['wait_time_ns'] = df_capacity_failed.apply(
            lambda row: calculate_time_diff_nano(
                row['waiting_start_nanoTime'], 
                row['critical_enter_nanoTime']
            ),
            axis=1
        )
        
        df_capacity_failed['fail_processing_time_ns'] = df_capacity_failed.apply(
            lambda row: calculate_time_diff_nano(
                row['critical_enter_nanoTime'], 
                row['critical_leave_nanoTime']
            ),
            axis=1
        )
        
        # 유효한 데이터만 필터링
        valid_capacity_failed = df_capacity_failed[
            (df_capacity_failed['wait_time_ns'].notna()) & 
            (df_capacity_failed['fail_processing_time_ns'].notna()) &
            (df_capacity_failed['wait_time_ns'] >= 0) & 
            (df_capacity_failed['fail_processing_time_ns'] >= 0)
        ]
        
        print(f"  - 정원초과 실패 그룹 시간 계산 완료 (유효 데이터: {len(valid_capacity_failed)}개/{len(df_capacity_failed)}개)")
        df_capacity_failed = valid_capacity_failed
    
    # 5. Sheet 1: Overall_Summary 계산
    total_requests = len(df_total)
    success_count = len(df_success)
    lock_failed_count = len(df_lock_failed)
    capacity_failed_count = len(df_capacity_failed)
    
    summary_data = {
        'Category': [
            'Total Requests',
            'Success',
            'Entry Failed (Lock, etc)',
            'Capacity Failed'
        ],
        'Count': [
            total_requests,
            success_count,
            lock_failed_count,
            capacity_failed_count
        ],
        'Percentage (%)': [
            100.0,
            calculate_rate(success_count, total_requests),
            calculate_rate(lock_failed_count, total_requests),
            calculate_rate(capacity_failed_count, total_requests)
        ]
    }
    df_summary = pd.DataFrame(summary_data)
    
    # 6. Sheet 2: Overall_Success_Stats 계산 (나노초 단위)
    if len(df_success) > 0:
        print("  - 성공 그룹 통계 계산 중...")
        wait_time_stats = get_stats_with_precision(df_success['wait_time_ns'], "Wait Time", "ns")
        dwell_time_stats = get_stats_with_precision(df_success['dwell_time_ns'], "Dwell Time", "ns")
        
        # 기본 통계만
        success_stats_data = {
            'Metric': ['Wait Time'] * 3 + ['Dwell Time (Critical Section)'] * 3,
            'Statistic': ['Mean', 'Median', 'Max'] * 2,
            'Value': [
                wait_time_stats['Mean'], 
                wait_time_stats['Median'], 
                wait_time_stats['Max'],
                dwell_time_stats['Mean'], 
                dwell_time_stats['Median'], 
                dwell_time_stats['Max']
            ],
            'Unit': ['ns'] * 6  # 나노초 단위로 변경
        }
        df_success_stats = pd.DataFrame(success_stats_data)
    else:
        df_success_stats = pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # 7. Sheet 3: Overall_Capacity_Failed_Stats 계산 (나노초 단위)
    if len(df_capacity_failed) > 0:
        print("  - 정원초과 실패 그룹 통계 계산 중...")
        capacity_wait_stats = get_stats_with_precision(df_capacity_failed['wait_time_ns'], "Wait Time (Capacity Failed)", "ns")
        fail_proc_stats = get_stats_with_precision(df_capacity_failed['fail_processing_time_ns'], "Fail Processing Time", "ns")
        
        capacity_failed_stats_data = {
            'Metric': ['Wait Time'] * 3 + ['Fail Processing Time'] * 3,
            'Statistic': ['Mean', 'Median', 'Max'] * 2,
            'Value': [
                capacity_wait_stats['Mean'], 
                capacity_wait_stats['Median'], 
                capacity_wait_stats['Max'],
                fail_proc_stats['Mean'], 
                fail_proc_stats['Median'], 
                fail_proc_stats['Max']
            ],
            'Unit': ['ns'] * 6  # 나노초 단위로 변경
        }
        df_capacity_failed_stats = pd.DataFrame(capacity_failed_stats_data)
    else:
        df_capacity_failed_stats = pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # 8. Sheet 4: Per_Room_Stats 계산 (나노초 단위)
    print("  - Per_Room_Stats 계산 중...")
    all_rooms = sorted(df_total['roomNumber'].unique())
    
    room_stats_list = []
    for room in all_rooms:
        room_total = df_total[df_total['roomNumber'] == room]
        room_success = df_success[df_success['roomNumber'] == room] if len(df_success) > 0 else pd.DataFrame()
        room_capacity_failed = df_capacity_failed[df_capacity_failed['roomNumber'] == room] if len(df_capacity_failed) > 0 else pd.DataFrame()
        room_lock_failed = df_lock_failed[df_lock_failed['roomNumber'] == room] if len(df_lock_failed) > 0 else pd.DataFrame()
        
        stats = {
            'roomNumber': room,
            'total_requests': len(room_total),
            'success_count': len(room_success),
            'capacity_failed_count': len(room_capacity_failed),
            'entry_failed_count': len(room_lock_failed),
            'success_rate(%)': calculate_rate(len(room_success), len(room_total)),
            'capacity_failed_rate(%)': calculate_rate(len(room_capacity_failed), len(room_total)),
            'entry_failed_rate(%)': calculate_rate(len(room_lock_failed), len(room_total)),
            'success_avg_wait_time(ns)': room_success['wait_time_ns'].mean() if len(room_success) > 0 else 0,
            'success_median_wait_time(ns)': room_success['wait_time_ns'].median() if len(room_success) > 0 else 0,
            'success_avg_dwell_time(ns)': room_success['dwell_time_ns'].mean() if len(room_success) > 0 else 0,
            'success_median_dwell_time(ns)': room_success['dwell_time_ns'].median() if len(room_success) > 0 else 0,
            'capacity_failed_avg_wait_time(ns)': room_capacity_failed['wait_time_ns'].mean() if len(room_capacity_failed) > 0 else 0,
            'capacity_failed_avg_fail_processing_time(ns)': room_capacity_failed['fail_processing_time_ns'].mean() if len(room_capacity_failed) > 0 else 0
        }
        
        room_stats_list.append(stats)
    
    df_per_room_stats = pd.DataFrame(room_stats_list)
    
    # 9. Sheet 5: Per_Bin_Stats 계산 (나노초 단위)
    print("  - Per_Bin_Stats 계산 중...")
    if 'bin' in df_total.columns:
        all_combinations = df_total.groupby(['roomNumber', 'bin']).size().index.tolist()
        
        bin_stats_list = []
        for room, bin_num in all_combinations:
            combo_total = df_total[(df_total['roomNumber'] == room) & (df_total['bin'] == bin_num)]
            combo_success = df_success[(df_success['roomNumber'] == room) & (df_success['bin'] == bin_num)] if len(df_success) > 0 else pd.DataFrame()
            combo_capacity_failed = df_capacity_failed[(df_capacity_failed['roomNumber'] == room) & (df_capacity_failed['bin'] == bin_num)] if len(df_capacity_failed) > 0 else pd.DataFrame()
            combo_lock_failed = df_lock_failed[(df_lock_failed['roomNumber'] == room) & (df_lock_failed['bin'] == bin_num)] if len(df_lock_failed) > 0 else pd.DataFrame()
            
            stats = {
                'roomNumber': room,
                'bin': bin_num,
                'total_requests': len(combo_total),
                'success_count': len(combo_success),
                'capacity_failed_count': len(combo_capacity_failed),
                'entry_failed_count': len(combo_lock_failed),
                'success_rate(%)': calculate_rate(len(combo_success), len(combo_total)),
                'capacity_failed_rate(%)': calculate_rate(len(combo_capacity_failed), len(combo_total)),
                'entry_failed_rate(%)': calculate_rate(len(combo_lock_failed), len(combo_total)),
                'success_avg_wait_time(ns)': combo_success['wait_time_ns'].mean() if len(combo_success) > 0 else 0,
                'success_median_wait_time(ns)': combo_success['wait_time_ns'].median() if len(combo_success) > 0 else 0,
                'success_avg_dwell_time(ns)': combo_success['dwell_time_ns'].mean() if len(combo_success) > 0 else 0,
                'success_median_dwell_time(ns)': combo_success['dwell_time_ns'].median() if len(combo_success) > 0 else 0,
                'capacity_failed_avg_wait_time(ns)': combo_capacity_failed['wait_time_ns'].mean() if len(combo_capacity_failed) > 0 else 0,
                'capacity_failed_avg_fail_processing_time(ns)': combo_capacity_failed['fail_processing_time_ns'].mean() if len(combo_capacity_failed) > 0 else 0
            }
            
            bin_stats_list.append(stats)
        
        df_per_bin_stats = pd.DataFrame(bin_stats_list)
    else:
        df_per_bin_stats = pd.DataFrame()
    
    # 10. Sheet 6: Time_Unit_Comparison (다양한 단위로 표시)
    print("  - 시간 단위 비교 통계 계산 중...")
    comparison_stats_list = []
    
    if len(df_success) > 0:
        # 나노초, 마이크로초, 밀리초 단위로 각각 계산
        wait_time_ns_stats = get_stats_with_precision(df_success['wait_time_ns'], "Wait Time", "ns")
        wait_time_us_stats = get_stats_with_precision(df_success['wait_time_ns'], "Wait Time", "us")
        wait_time_ms_stats = get_stats_with_precision(df_success['wait_time_ns'], "Wait Time", "ms")
        
        dwell_time_ns_stats = get_stats_with_precision(df_success['dwell_time_ns'], "Dwell Time", "ns")
        dwell_time_us_stats = get_stats_with_precision(df_success['dwell_time_ns'], "Dwell Time", "us")
        dwell_time_ms_stats = get_stats_with_precision(df_success['dwell_time_ns'], "Dwell Time", "ms")
        
        # Wait Time 통계 (모든 단위)
        comparison_stats_list.extend([
            {
                'Metric': 'Success - Wait Time',
                'Unit': 'nanoseconds',
                'Mean': wait_time_ns_stats['Mean'],
                'Median': wait_time_ns_stats['Median'],
                'Max': wait_time_ns_stats['Max']
            },
            {
                'Metric': 'Success - Wait Time',
                'Unit': 'microseconds',
                'Mean': wait_time_us_stats['Mean'],
                'Median': wait_time_us_stats['Median'],
                'Max': wait_time_us_stats['Max']
            },
            {
                'Metric': 'Success - Wait Time',
                'Unit': 'milliseconds',
                'Mean': wait_time_ms_stats['Mean'],
                'Median': wait_time_ms_stats['Median'],
                'Max': wait_time_ms_stats['Max']
            },
            {
                'Metric': 'Success - Dwell Time',
                'Unit': 'nanoseconds',
                'Mean': dwell_time_ns_stats['Mean'],
                'Median': dwell_time_ns_stats['Median'],
                'Max': dwell_time_ns_stats['Max']
            },
            {
                'Metric': 'Success - Dwell Time',
                'Unit': 'microseconds',
                'Mean': dwell_time_us_stats['Mean'],
                'Median': dwell_time_us_stats['Median'],
                'Max': dwell_time_us_stats['Max']
            },
            {
                'Metric': 'Success - Dwell Time',
                'Unit': 'milliseconds',
                'Mean': dwell_time_ms_stats['Mean'],
                'Median': dwell_time_ms_stats['Median'],
                'Max': dwell_time_ms_stats['Max']
            }
        ])
    
    df_comparison_stats = pd.DataFrame(comparison_stats_list)
    
    # 11. 데이터 검증
    print("\n  - 데이터 검증:")
    print(f"    전체 요청 수: {total_requests}")
    print(f"    성공: {success_count} ({calculate_rate(success_count, total_requests):.2f}%)")
    print(f"    정원초과 실패: {capacity_failed_count} ({calculate_rate(capacity_failed_count, total_requests):.2f}%)")
    print(f"    진입 실패: {lock_failed_count} ({calculate_rate(lock_failed_count, total_requests):.2f}%)")
    
    # 12. Excel 파일로 저장
    output_dir = 'performance_reports'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_path = os.path.join(output_dir, f"{label}_stats_nano.xlsx")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Overall_Summary', index=False)
            df_success_stats.to_excel(writer, sheet_name='Overall_Success_Stats', index=False)
            df_capacity_failed_stats.to_excel(writer, sheet_name='Overall_Capacity_Failed_Stats', index=False)
            df_per_room_stats.to_excel(writer, sheet_name='Per_Room_Stats', index=False)
            if not df_per_bin_stats.empty:
                df_per_bin_stats.to_excel(writer, sheet_name='Per_Bin_Stats', index=False)
            if not df_comparison_stats.empty:
                df_comparison_stats.to_excel(writer, sheet_name='Time_Unit_Comparison', index=False)
        
        # Excel 파일 후처리로 숫자 포맷 설정
        wb = load_workbook(output_path)
        
        # Overall_Summary의 Percentage 컬럼 포맷
        if 'Overall_Summary' in wb.sheetnames:
            ws = wb['Overall_Summary']
            for row in range(2, ws.max_row + 1):
                cell = ws[f'C{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # Overall_Success_Stats와 Overall_Capacity_Failed_Stats의 Value 컬럼 포맷 (나노초)
        for sheet_name in ['Overall_Success_Stats', 'Overall_Capacity_Failed_Stats']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'C{row}']
                    if cell.value is not None:
                        # Statistic 컬럼 확인
                        stat_cell = ws[f'B{row}']
                        if stat_cell.value and 'Non-Zero %' in str(stat_cell.value):
                            cell.number_format = '0.00"%"'
                        else:
                            cell.number_format = '#,##0'  # 나노초는 정수로 표시
        
        # Per_Room_Stats의 포맷 설정
        if 'Per_Room_Stats' in wb.sheetnames:
            ws = wb['Per_Room_Stats']
            # 비율 컬럼들
            for col in ['F', 'G', 'H']:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col}{row}']
                    if cell.value is not None:
                        cell.number_format = '0.00"%"'
            
            # 시간 컬럼들 (나노초)
            for col in ['I', 'J', 'K', 'L', 'M', 'N']:
                for row in range(2, ws.max_row + 1):
                    cell = ws[f'{col}{row}']
                    if cell.value is not None:
                        cell.number_format = '#,##0'  # 나노초는 정수로 표시
        
        # Time_Unit_Comparison의 포맷 설정
        if 'Time_Unit_Comparison' in wb.sheetnames:
            ws = wb['Time_Unit_Comparison']
            for row in range(2, ws.max_row + 1):
                unit_cell = ws[f'B{row}']
                if unit_cell.value:
                    # 단위에 따라 다른 포맷 적용
                    for col in ['C', 'D', 'E']:
                        cell = ws[f'{col}{row}']
                        if cell.value is not None:
                            if 'nanoseconds' in str(unit_cell.value):
                                cell.number_format = '#,##0'
                            elif 'microseconds' in str(unit_cell.value):
                                cell.number_format = '#,##0.000'
                            elif 'milliseconds' in str(unit_cell.value):
                                cell.number_format = '#,##0.000000'
        
        wb.save(output_path)
        
        print(f"  - Excel 파일 저장 완료: {output_path}")
        return True
    except Exception as e:
        print(f"오류: Excel 파일 저장 실패 - {e}")
        return False

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(
        description='성능 테스트 결과 CSV 파일을 분석하여 나노초 단위 통계 지표를 계산하고 Excel로 저장합니다.'
    )
    parser.add_argument(
        '--inputs',
        type=str,
        required=True,
        help='분석할 CSV 파일 경로들 (콤마로 구분)'
    )
    parser.add_argument(
        '--labels',
        type=str,
        required=True,
        help='각 CSV 파일에 해당하는 출력 레이블 (콤마로 구분)'
    )
    parser.add_argument(
        '--compare',
        type=str,
        help='비교할 참조 Excel 파일 경로 (선택사항)'
    )
    
    args = parser.parse_args()
    
    # 입력 파일과 레이블 파싱
    input_files = [f.strip() for f in args.inputs.split(',')]
    labels = [l.strip() for l in args.labels.split(',')]
    
    if len(input_files) != len(labels):
        print("오류: 입력 파일 수와 레이블 수가 일치하지 않습니다.")
        sys.exit(1)
    
    # 시작 시간 기록
    start_time = datetime.now()
    print(f"성능 분석 시작: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"버전: v6.0 - 나노초 단위 분석")
    
    # 파일 처리
    success_count = 0
    for csv_path, label in zip(input_files, labels):
        if process_performance_data(csv_path, label):
            success_count += 1
    
    # 종료 시간 및 소요 시간 계산
    end_time = datetime.now()
    elapsed_time = end_time - start_time
    
    print(f"\n성능 분석 완료: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"소요 시간: {elapsed_time}")
    print(f"처리 결과: {success_count}/{len(input_files)} 파일 성공")

if __name__ == "__main__":
    main()