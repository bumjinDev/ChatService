#!/usr/bin/env python3
"""
성능 분석 지표 계산 스크립트 (v5.2) - 깔끔한 리팩토링 버전
- 불필요한 로그/디버깅 코드 제거
- 중복 코드 제거 및 최적화
- 핵심 로직 보존
"""

import pandas as pd
import numpy as np
import os
import argparse
from datetime import datetime
import sys
from openpyxl import load_workbook

def calculate_rate(count, total):
    """전체 대비 비율 계산"""
    if isinstance(count, pd.Series) or isinstance(total, pd.Series):
        return np.where(total != 0, (count / total) * 100, 0)
    return (count / total) * 100 if total > 0 else 0

def get_stats_with_precision(series, unit="ms"):
    """시리즈의 통계값 계산 (Mean, Median, Max, NonZero 통계)"""
    if len(series) == 0:
        return {
            'Mean': 0.0, 'Median': 0.0, 'Max': 0.0,
            'Mean_NonZero': 0.0, 'Median_NonZero': 0.0,
            'NonZero_Count': 0, 'NonZero_Percentage': 0.0
        }
    
    valid_series = series.dropna()
    if len(valid_series) == 0:
        return {
            'Mean': 0.0, 'Median': 0.0, 'Max': 0.0,
            'Mean_NonZero': 0.0, 'Median_NonZero': 0.0,
            'NonZero_Count': 0, 'NonZero_Percentage': 0.0
        }
    
    # 단위 변환
    factor = {'ms': 1.0, 'us': 1000.0, 'ns': 1_000_000.0}.get(unit, 1.0)
    
    # 기본 통계
    mean_val = float(valid_series.mean())
    median_val = float(valid_series.median())
    max_val = float(valid_series.max())
    
    # 0이 아닌 값들의 통계
    non_zero = valid_series[valid_series > 0]
    non_zero_count = len(non_zero)
    non_zero_percentage = (non_zero_count / len(valid_series)) * 100
    
    mean_non_zero = float(non_zero.mean()) if non_zero_count > 0 else 0.0
    median_non_zero = float(non_zero.median()) if non_zero_count > 0 else 0.0
    
    return {
        'Mean': mean_val * factor,
        'Median': median_val * factor,
        'Max': max_val * factor,
        'Mean_NonZero': mean_non_zero * factor,
        'Median_NonZero': median_non_zero * factor,
        'NonZero_Count': non_zero_count,
        'NonZero_Percentage': non_zero_percentage
    }

def parse_nano_time_precise(nano_str):
    """나노초 문자열을 정확하게 파싱"""
    if pd.isna(nano_str) or nano_str == 'nan':
        return np.nan
    
    try:
        if isinstance(nano_str, str):
            if 'E' in nano_str or 'e' in nano_str:
                from decimal import Decimal
                return int(Decimal(nano_str))
            else:
                return int(float(nano_str))
        else:
            return int(nano_str)
    except (ValueError, TypeError):
        return np.nan

def calculate_time_diff_nano_precise(start_nano, end_nano):
    """나노초 값 간의 시간 차이를 밀리초로 계산"""
    if pd.isna(start_nano) or pd.isna(end_nano):
        return np.nan
    
    try:
        diff_nano = int(end_nano) - int(start_nano)
        return diff_nano / 1_000_000.0  # 나노초를 밀리초로 변환
    except (ValueError, TypeError):
        return np.nan

def set_join_result_from_events(df):
    """이벤트 타입 기반으로 join_result 컬럼 설정"""
    def determine_join_result(row):
        if pd.notna(row.get('join_result')) and row.get('join_result') != '':
            existing_result = str(row['join_result']).strip()
            if existing_result in ['SUCCESS', 'FAIL_OVER_CAPACITY', 'FAIL_ENTRY']:
                return existing_result
        
        if pd.isna(row.get('critical_leave_event_type')) or row.get('critical_leave_event_type') == '':
            return 'FAIL_ENTRY'
        elif str(row['critical_leave_event_type']).strip() == 'SUCCESS':
            return 'SUCCESS'
        elif str(row['critical_leave_event_type']).strip() == 'FAIL_OVER_CAPACITY':
            return 'FAIL_OVER_CAPACITY'
        else:
            return 'UNKNOWN'
    
    df_result = df.copy()
    df_result['join_result'] = df_result.apply(determine_join_result, axis=1)
    return df_result

def calculate_time_metrics(df_group, time_cols):
    """시간 메트릭 계산"""
    if len(df_group) == 0:
        return {}
    
    df_group[time_cols[0]] = df_group.apply(
        lambda row: calculate_time_diff_nano_precise(
            row['waiting_start_nanoTime'], 
            row['critical_enter_nanoTime']
        ), axis=1
    )
    
    df_group[time_cols[1]] = df_group.apply(
        lambda row: calculate_time_diff_nano_precise(
            row['critical_enter_nanoTime'], 
            row['critical_leave_nanoTime']
        ), axis=1
    )
    
    # 유효한 데이터만 필터링
    valid_mask = (
        df_group[time_cols[0]].notna() & 
        df_group[time_cols[1]].notna() &
        (df_group[time_cols[0]] >= 0) & 
        (df_group[time_cols[1]] >= 0)
    )
    
    return df_group[valid_mask]

def create_stats_dataframe(stats_dict, metric_names):
    """통계 데이터프레임 생성"""
    metrics = []
    statistics = []
    values = []
    
    for i, (metric_name, stats) in enumerate(zip(metric_names, stats_dict)):
        metrics.extend([metric_name] * 6)
        statistics.extend(['Mean', 'Median', 'Max', 'Mean (Non-Zero)', 'Median (Non-Zero)', 'Non-Zero %'])
        values.extend([
            stats['Mean'], stats['Median'], stats['Max'],
            stats['Mean_NonZero'], stats['Median_NonZero'], stats['NonZero_Percentage']
        ])
    
    return pd.DataFrame({
        'Metric': metrics,
        'Statistic': statistics,
        'Value': values,
        'Unit': ['ms'] * len(values)
    })

def calculate_room_stats(df_total, df_success, df_capacity_failed, df_lock_failed):
    """방별 통계 계산"""
    room_stats_list = []
    
    for room in sorted(df_total['roomNumber'].unique()):
        room_data = {
            'total': df_total[df_total['roomNumber'] == room],
            'success': df_success[df_success['roomNumber'] == room] if len(df_success) > 0 else pd.DataFrame(),
            'capacity_failed': df_capacity_failed[df_capacity_failed['roomNumber'] == room] if len(df_capacity_failed) > 0 else pd.DataFrame(),
            'lock_failed': df_lock_failed[df_lock_failed['roomNumber'] == room] if len(df_lock_failed) > 0 else pd.DataFrame()
        }
        
        # 0이 아닌 값들의 중간값 계산
        wait_median_non_zero = 0
        dwell_median_non_zero = 0
        if len(room_data['success']) > 0:
            wait_non_zero = room_data['success'][room_data['success']['wait_time_ms'] > 0]['wait_time_ms']
            dwell_non_zero = room_data['success'][room_data['success']['dwell_time_ms'] > 0]['dwell_time_ms']
            wait_median_non_zero = wait_non_zero.median() if len(wait_non_zero) > 0 else 0
            dwell_median_non_zero = dwell_non_zero.median() if len(dwell_non_zero) > 0 else 0
        
        stats = {
            'roomNumber': room,
            'total_requests': len(room_data['total']),
            'success_count': len(room_data['success']),
            'capacity_failed_count': len(room_data['capacity_failed']),
            'entry_failed_count': len(room_data['lock_failed']),
            'success_rate(%)': calculate_rate(len(room_data['success']), len(room_data['total'])),
            'capacity_failed_rate(%)': calculate_rate(len(room_data['capacity_failed']), len(room_data['total'])),
            'entry_failed_rate(%)': calculate_rate(len(room_data['lock_failed']), len(room_data['total'])),
            'success_avg_wait_time(ms)': room_data['success']['wait_time_ms'].mean() if len(room_data['success']) > 0 else 0,
            'success_median_wait_time(ms)': room_data['success']['wait_time_ms'].median() if len(room_data['success']) > 0 else 0,
            'success_median_wait_time_non_zero(ms)': wait_median_non_zero,
            'success_avg_dwell_time(ms)': room_data['success']['dwell_time_ms'].mean() if len(room_data['success']) > 0 else 0,
            'success_median_dwell_time(ms)': room_data['success']['dwell_time_ms'].median() if len(room_data['success']) > 0 else 0,
            'success_median_dwell_time_non_zero(ms)': dwell_median_non_zero,
            'capacity_failed_avg_wait_time(ms)': room_data['capacity_failed']['wait_time_ms'].mean() if len(room_data['capacity_failed']) > 0 else 0,
            'capacity_failed_avg_fail_processing_time(ms)': room_data['capacity_failed']['fail_processing_time_ms'].mean() if len(room_data['capacity_failed']) > 0 else 0
        }
        
        room_stats_list.append(stats)
    
    return pd.DataFrame(room_stats_list)

def calculate_bin_stats(df_total, df_success, df_capacity_failed, df_lock_failed):
    """구간별 통계 계산"""
    if 'bin' not in df_total.columns:
        return pd.DataFrame()
    
    bin_stats_list = []
    for room, bin_num in df_total.groupby(['roomNumber', 'bin']).size().index.tolist():
        bin_data = {
            'total': df_total[(df_total['roomNumber'] == room) & (df_total['bin'] == bin_num)],
            'success': df_success[(df_success['roomNumber'] == room) & (df_success['bin'] == bin_num)] if len(df_success) > 0 else pd.DataFrame(),
            'capacity_failed': df_capacity_failed[(df_capacity_failed['roomNumber'] == room) & (df_capacity_failed['bin'] == bin_num)] if len(df_capacity_failed) > 0 else pd.DataFrame(),
            'lock_failed': df_lock_failed[(df_lock_failed['roomNumber'] == room) & (df_lock_failed['bin'] == bin_num)] if len(df_lock_failed) > 0 else pd.DataFrame()
        }
        
        stats = {
            'roomNumber': room,
            'bin': bin_num,
            'total_requests': len(bin_data['total']),
            'success_count': len(bin_data['success']),
            'capacity_failed_count': len(bin_data['capacity_failed']),
            'entry_failed_count': len(bin_data['lock_failed']),
            'success_rate(%)': calculate_rate(len(bin_data['success']), len(bin_data['total'])),
            'capacity_failed_rate(%)': calculate_rate(len(bin_data['capacity_failed']), len(bin_data['total'])),
            'entry_failed_rate(%)': calculate_rate(len(bin_data['lock_failed']), len(bin_data['total'])),
            'success_avg_wait_time(ms)': bin_data['success']['wait_time_ms'].mean() if len(bin_data['success']) > 0 else 0,
            'success_median_wait_time(ms)': bin_data['success']['wait_time_ms'].median() if len(bin_data['success']) > 0 else 0,
            'success_avg_dwell_time(ms)': bin_data['success']['dwell_time_ms'].mean() if len(bin_data['success']) > 0 else 0,
            'success_median_dwell_time(ms)': bin_data['success']['dwell_time_ms'].median() if len(bin_data['success']) > 0 else 0,
            'capacity_failed_avg_wait_time(ms)': bin_data['capacity_failed']['wait_time_ms'].mean() if len(bin_data['capacity_failed']) > 0 else 0,
            'capacity_failed_avg_fail_processing_time(ms)': bin_data['capacity_failed']['fail_processing_time_ms'].mean() if len(bin_data['capacity_failed']) > 0 else 0
        }
        
        bin_stats_list.append(stats)
    
    return pd.DataFrame(bin_stats_list)

def apply_excel_formatting(wb, output_path):
    """Excel 파일 포맷 적용"""
    # Overall_Summary 포맷
    if 'Overall_Summary' in wb.sheetnames:
        ws = wb['Overall_Summary']
        for row in range(2, ws.max_row + 1):
            cell = ws[f'C{row}']
            if cell.value is not None:
                cell.number_format = '0.00"%"'
    
    # Overall_Success_Stats와 Overall_Capacity_Failed_Stats 포맷
    for sheet_name in ['Overall_Success_Stats', 'Overall_Capacity_Failed_Stats']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in range(2, ws.max_row + 1):
                cell = ws[f'C{row}']
                if cell.value is not None:
                    stat_cell = ws[f'B{row}']
                    if stat_cell.value and 'Non-Zero %' in str(stat_cell.value):
                        cell.number_format = '0.00"%"'
                    else:
                        cell.number_format = '0.000000'
    
    # Per_Room_Stats 포맷
    if 'Per_Room_Stats' in wb.sheetnames:
        ws = wb['Per_Room_Stats']
        # 비율 컬럼들
        for col in ['F', 'G', 'H']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.00"%"'
        
        # 시간 컬럼들
        for col in ['I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.000006'
    
    # Nanosecond_Precision_Stats 포맷
    if 'Nanosecond_Precision_Stats' in wb.sheetnames:
        ws = wb['Nanosecond_Precision_Stats']
        # 마이크로초 컬럼들
        for col in ['B', 'C', 'D', 'E', 'F']:
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{col}{row}']
                if cell.value is not None:
                    cell.number_format = '0.000'
        
        # 퍼센트 컬럼
        for row in range(2, ws.max_row + 1):
            cell = ws[f'H{row}']
            if cell.value is not None:
                cell.number_format = '0.00"%"'
    
    wb.save(output_path)

def process_performance_data(csv_path, label):
    """CSV 파일 처리 및 성능 통계 계산"""
    print(f"\n처리 중: {csv_path} (레이블: {label})")
    
    # 1. 데이터 로드
    try:
        dtype_spec = {
            'waiting_start_nanoTime': str,
            'critical_enter_nanoTime': str,
            'critical_leave_nanoTime': str,
            'increment_before_nanoTime': str,
            'increment_after_nanoTime': str
        }
        
        df_total = pd.read_csv(csv_path, dtype=dtype_spec)
        
        # 나노초 컬럼 변환
        for col in [k for k in dtype_spec.keys() if k in df_total.columns]:
            df_total[col] = df_total[col].apply(parse_nano_time_precise)
        
    except FileNotFoundError:
        print(f"오류: CSV 파일을 찾을 수 없습니다 - {csv_path}")
        return False
    except Exception as e:
        print(f"오류: CSV 파일 로드 실패 - {e}")
        return False
    
    # 2. join_result 설정
    df_total = set_join_result_from_events(df_total)
    
    # 3. 데이터 그룹 분류
    df_success = df_total[df_total['join_result'] == 'SUCCESS'].copy()
    df_lock_failed = df_total[df_total['join_result'] == 'FAIL_ENTRY'].copy()
    df_capacity_failed = df_total[df_total['join_result'] == 'FAIL_OVER_CAPACITY'].copy()
    
    # 4. 시간 계산
    if len(df_success) > 0:
        df_success = calculate_time_metrics(df_success, ['wait_time_ms', 'dwell_time_ms'])
    
    if len(df_capacity_failed) > 0:
        df_capacity_failed = calculate_time_metrics(df_capacity_failed, ['wait_time_ms', 'fail_processing_time_ms'])
    
    # 5. 통계 계산
    total_requests = len(df_total)
    success_count = len(df_success)
    lock_failed_count = len(df_lock_failed)
    capacity_failed_count = len(df_capacity_failed)
    
    # Overall Summary
    df_summary = pd.DataFrame({
        'Category': ['Total Requests', 'Success', 'Entry Failed (Lock, etc)', 'Capacity Failed'],
        'Count': [total_requests, success_count, lock_failed_count, capacity_failed_count],
        'Percentage (%)': [
            100.0,
            calculate_rate(success_count, total_requests),
            calculate_rate(lock_failed_count, total_requests),
            calculate_rate(capacity_failed_count, total_requests)
        ]
    })
    
    # Success Stats
    if len(df_success) > 0:
        wait_stats = get_stats_with_precision(df_success['wait_time_ms'])
        dwell_stats = get_stats_with_precision(df_success['dwell_time_ms'])
        df_success_stats = create_stats_dataframe(
            [wait_stats, dwell_stats], 
            ['Wait Time', 'Dwell Time (Critical Section)']
        )
    else:
        df_success_stats = pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # Capacity Failed Stats
    if len(df_capacity_failed) > 0:
        capacity_wait_stats = get_stats_with_precision(df_capacity_failed['wait_time_ms'])
        fail_proc_stats = get_stats_with_precision(df_capacity_failed['fail_processing_time_ms'])
        df_capacity_failed_stats = create_stats_dataframe(
            [capacity_wait_stats, fail_proc_stats], 
            ['Wait Time', 'Fail Processing Time']
        )
    else:
        df_capacity_failed_stats = pd.DataFrame({'Metric': [], 'Statistic': [], 'Value': [], 'Unit': []})
    
    # Per Room Stats
    df_per_room_stats = calculate_room_stats(df_total, df_success, df_capacity_failed, df_lock_failed)
    
    # Per Bin Stats
    df_per_bin_stats = calculate_bin_stats(df_total, df_success, df_capacity_failed, df_lock_failed)
    
    # Nanosecond Precision Stats
    df_nano_stats = pd.DataFrame()
    if len(df_success) > 0:
        wait_us_stats = get_stats_with_precision(df_success['wait_time_ms'], "us")
        dwell_us_stats = get_stats_with_precision(df_success['dwell_time_ms'], "us")
        
        df_nano_stats = pd.DataFrame([
            {
                'Category': 'Success - Wait Time',
                'Mean (μs)': wait_us_stats['Mean'],
                'Median (μs)': wait_us_stats['Median'],
                'Max (μs)': wait_us_stats['Max'],
                'Mean Non-Zero (μs)': wait_us_stats['Mean_NonZero'],
                'Median Non-Zero (μs)': wait_us_stats['Median_NonZero'],
                'Non-Zero Count': wait_us_stats['NonZero_Count'],
                'Non-Zero %': wait_us_stats['NonZero_Percentage']
            },
            {
                'Category': 'Success - Dwell Time',
                'Mean (μs)': dwell_us_stats['Mean'],
                'Median (μs)': dwell_us_stats['Median'],
                'Max (μs)': dwell_us_stats['Max'],
                'Mean Non-Zero (μs)': dwell_us_stats['Mean_NonZero'],
                'Median Non-Zero (μs)': dwell_us_stats['Median_NonZero'],
                'Non-Zero Count': dwell_us_stats['NonZero_Count'],
                'Non-Zero %': dwell_us_stats['NonZero_Percentage']
            }
        ])
    
    # 6. Excel 저장
    output_dir = 'performance_reports'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    output_path = os.path.join(output_dir, f"{label}_stats.xlsx")
    
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df_summary.to_excel(writer, sheet_name='Overall_Summary', index=False)
            df_success_stats.to_excel(writer, sheet_name='Overall_Success_Stats', index=False)
            df_capacity_failed_stats.to_excel(writer, sheet_name='Overall_Capacity_Failed_Stats', index=False)
            df_per_room_stats.to_excel(writer, sheet_name='Per_Room_Stats', index=False)
            if not df_per_bin_stats.empty:
                df_per_bin_stats.to_excel(writer, sheet_name='Per_Bin_Stats', index=False)
            if not df_nano_stats.empty:
                df_nano_stats.to_excel(writer, sheet_name='Nanosecond_Precision_Stats', index=False)
        
        wb = load_workbook(output_path)
        apply_excel_formatting(wb, output_path)
        
        print(f"Excel 파일 저장 완료: {output_path}")
        
        # 요약 결과 출력
        print(f"전체 요청: {total_requests}, 성공: {success_count} ({calculate_rate(success_count, total_requests):.1f}%), "
              f"정원초과실패: {capacity_failed_count} ({calculate_rate(capacity_failed_count, total_requests):.1f}%), "
              f"진입실패: {lock_failed_count} ({calculate_rate(lock_failed_count, total_requests):.1f}%)")
        
        return True
    except Exception as e:
        print(f"오류: Excel 파일 저장 실패 - {e}")
        return False

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description='성능 테스트 결과 CSV 파일 분석')
    parser.add_argument('--inputs', type=str, required=True, help='분석할 CSV 파일 경로들 (콤마로 구분)')
    parser.add_argument('--labels', type=str, required=True, help='각 CSV 파일에 해당하는 출력 레이블 (콤마로 구분)')
    
    args = parser.parse_args()
    
    input_files = [f.strip() for f in args.inputs.split(',')]
    labels = [l.strip() for l in args.labels.split(',')]
    
    if len(input_files) != len(labels):
        print("오류: 입력 파일 수와 레이블 수가 일치하지 않습니다.")
        sys.exit(1)
    
    start_time = datetime.now()
    print(f"성능 분석 시작: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    success_count = 0
    for csv_path, label in zip(input_files, labels):
        if process_performance_data(csv_path, label):
            success_count += 1
    
    end_time = datetime.now()
    print(f"\n성능 분석 완료: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"소요 시간: {end_time - start_time}")
    print(f"처리 결과: {success_count}/{len(input_files)} 파일 성공")

if __name__ == "__main__":
    main()