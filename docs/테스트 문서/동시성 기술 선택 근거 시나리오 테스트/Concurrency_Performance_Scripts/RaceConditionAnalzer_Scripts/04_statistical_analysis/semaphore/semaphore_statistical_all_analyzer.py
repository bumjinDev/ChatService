#!/usr/bin/env python3
"""
Semaphore 전용 전체 통합 통계 분석기 - 완성판
- 세마포어의 3가지 핵심 분석 영역에 집중
- 분석 1: 순차적 일관성 관찰 (규칙 1+4 통합) - 현상 관찰
- 분석 2: 동시 실행 패턴 관찰 (규칙 2 재정의) - 현상 관찰  
- 분석 3: 정원 초과 방지 검증 (규칙 3 유지) - 성공/실패 판정
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

def load_and_validate_semaphore_data(preprocessor_file, analysis_file):
    """세마포어 데이터 로드 및 필수 컬럼 검증"""
    print("📂 세마포어 데이터 파일 로드 중...")
    
    # 전처리 데이터 로드
    preprocessor_df = pd.read_csv(preprocessor_file)
    print(f"✅ 세마포어 전처리 데이터 로드 완료: {len(preprocessor_df)}행")
    
    # 분석 결과 데이터 로드
    analysis_df = pd.read_csv(analysis_file)
    print(f"✅ 세마포어 분석 결과 데이터 로드 완료: {len(analysis_df)}행")
    
    # 세마포어 전처리 데이터 필수 컬럼 검증 (10개 컬럼)
    preprocessor_required = ['roomNumber', 'bin', 'user_id', 'prev_people', 'curr_people', 
                           'max_people', 'room_entry_sequence', 'join_result',
                           'true_critical_section_nanoTime_start', 'true_critical_section_nanoTime_end']
    missing_preprocessor = [col for col in preprocessor_required if col not in preprocessor_df.columns]
    if missing_preprocessor:
        raise ValueError(f"세마포어 전처리 데이터에서 필수 컬럼 누락: {missing_preprocessor}")
    
    # 세마포어 분석 결과 데이터 필수 컬럼 검증 (14개 컬럼)
    analysis_required = ['roomNumber', 'bin', 'user_id', 'anomaly_type', 
                        'over_capacity_amount', 'over_capacity_curr', 'over_capacity_max']
    missing_analysis = [col for col in analysis_required if col not in analysis_df.columns]
    if missing_analysis:
        raise ValueError(f"세마포어 분석 결과 데이터에서 필수 컬럼 누락: {missing_analysis}")
    
    # 데이터 타입 정리
    preprocessor_df['roomNumber'] = pd.to_numeric(preprocessor_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    preprocessor_df['bin'] = pd.to_numeric(preprocessor_df['bin'], errors='coerce').fillna(0).astype(int)
    
    analysis_df['roomNumber'] = pd.to_numeric(analysis_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    analysis_df['bin'] = pd.to_numeric(analysis_df['bin'], errors='coerce').fillna(0).astype(int)
    
    print("✅ 세마포어 데이터 검증 및 타입 변환 완료")
    return preprocessor_df, analysis_df

def calculate_semaphore_total_info(preprocessor_df):
    """세마포어 전체 요청 정보 집계"""
    print("📊 세마포어 전체 요청 정보 집계 중...")
    
    total_requests = len(preprocessor_df)
    total_rooms = preprocessor_df['roomNumber'].nunique()
    total_bins = len(preprocessor_df.groupby(['roomNumber', 'bin']))
    
    # 세마포어 특화 정보
    success_requests = len(preprocessor_df[preprocessor_df['join_result'] == 'SUCCESS'])
    fail_requests = len(preprocessor_df[preprocessor_df['join_result'].str.contains('FAIL', na=False)])
    
    print(f"✅ 세마포어 집계 완료:")
    print(f"  - 전체 permit 요청수: {total_requests:,}건")
    print(f"  - 전체 방 수: {total_rooms}개")
    print(f"  - 전체 (방×bin) 조합: {total_bins}개")
    print(f"  - permit 획득 성공: {success_requests:,}건")
    print(f"  - permit 획득 실패: {fail_requests:,}건")
    
    return {
        'total_requests': total_requests,
        'total_rooms': total_rooms,
        'total_bins': total_bins,
        'success_requests': success_requests,
        'fail_requests': fail_requests
    }

def analyze_sequential_consistency_observation(preprocessor_df, analysis_df, total_info):
    """분석 1: 순차적 일관성 관찰 (기존 규칙 1+4 통합) - 현상 관찰"""
    print("🔍 분석 1: 순차적 일관성 관찰 (규칙 1+4 통합) 중...")
    
    # 순차적 일관성 관찰: 이상적 순차 상태 vs 실제 기록값 비교
    sequential_differences = 0
    total_requests = total_info['total_requests']
    total_diff_amount = 0
    
    # room_entry_sequence 기반 이상적 순차 상태 계산
    for _, row in preprocessor_df.iterrows():
        initial_people = 1  # 테스트 시작 시 방에 1명 존재
        ideal_sequential_state = min(initial_people + row['room_entry_sequence'], row['max_people'])
        
        if row['curr_people'] != ideal_sequential_state:
            sequential_differences += 1
            total_diff_amount += abs(row['curr_people'] - ideal_sequential_state)
    
    # 결과 생성
    result = {
        '분석 구분': '순차적 일관성 관찰 (규칙 1+4 통합)',
        '전체 permit 요청수': total_requests,
        '전체 방 수': total_info['total_rooms'],
        '전체 (방×bin) 조합수': total_info['total_bins'],
        '순차적 일관성 차이 발생 건수': sequential_differences,
        '순차적 일관성 차이 발생률 (%)': round((sequential_differences / total_requests * 100), 2) if total_requests > 0 else 0,
        '순차적 처리 일치율 (%)': round(((total_requests - sequential_differences) / total_requests * 100), 2) if total_requests > 0 else 100,
        '평균 차이 크기': round(total_diff_amount / sequential_differences, 2) if sequential_differences > 0 else 0,
        '관찰된 특성': '부분적 보장 특성' if sequential_differences > 0 else '완전 순차적 처리',
        '세마포어 특성': 'CAS 기반 비순차적 처리의 자연스러운 현상' if sequential_differences > 0 else 'permit 기반 순차적 처리'
    }
    
    print(f"  순차적 일관성 차이: {sequential_differences}건 ({result['순차적 일관성 차이 발생률 (%)']}%)")
    print("✅ 순차적 일관성 관찰 완료")
    return result

def analyze_concurrent_execution_observation(preprocessor_df, analysis_df, total_info):
    """분석 2: 동시 실행 패턴 관찰 (기존 규칙 2 재정의) - 기존 방식으로 통일"""
    print("🔍 분석 2: 동시 실행 패턴 관찰 (규칙 2 재정의) 중...")
    
    # 실제 경합이 발생한 레코드만 필터링 (contention_group_size > 1)
    concurrent_records = analysis_df[
        (analysis_df['contention_group_size'].notna()) & 
        (analysis_df['contention_group_size'] > 1)
    ]
    
    total_requests = total_info['total_requests']
    
    if len(concurrent_records) > 0:
        # contention_group_size 기반 통계 계산 (기존 방식과 동일)
        contention_sizes = concurrent_records['contention_group_size'].dropna()
        
        # 통계 계산
        concurrent_executions = len(contention_sizes)
        max_concurrent_level = contention_sizes.max() if len(contention_sizes) > 0 else 1
        avg_concurrent_level = contention_sizes.mean() if len(contention_sizes) > 0 else 1.0
        total_concurrent_threads = contention_sizes.sum() if len(contention_sizes) > 0 else 0
        
        # 영향받은 방과 bin 계산
        affected_rooms = concurrent_records['roomNumber'].nunique()
        affected_bins = len(concurrent_records.groupby(['roomNumber', 'bin']))
    else:
        concurrent_executions = 0
        max_concurrent_level = 1
        avg_concurrent_level = 1.0
        total_concurrent_threads = 0
        affected_rooms = 0
        affected_bins = 0
    
    print(f"  실제 경합 발생: {concurrent_executions}건 (contention_group_size > 1)")
    
    # 결과 생성 (기존 동시성 측정과 유사한 구조)
    result = {
        '분석 구분': '동시 실행 패턴 관찰 (규칙 2 재정의)',
        '전체 permit 요청수': total_requests,
        '전체 방 수': total_info['total_rooms'],
        '전체 (방×bin) 조합수': total_info['total_bins'],
        '발생 건수': concurrent_executions,
        '발생률 (%)': round((concurrent_executions / total_requests * 100), 2) if total_requests > 0 else 0,
        '영향받은 방 수': affected_rooms,
        '영향받은 (방×bin) 조합수': affected_bins,
        '총 경합 스레드 수': int(total_concurrent_threads),
        '평균 경합 스레드 수': round(avg_concurrent_level, 2),
        '최대 경합 스레드 수': int(max_concurrent_level),
        '중간값 경합 그룹 크기': round(contention_sizes.median(), 2) if len(contention_sizes) > 0 else np.nan,
        '경합 강도 표준편차': round(contention_sizes.std(), 4) if len(contention_sizes) > 1 else 0.0,
        '관찰된 패턴': 'CAS 기반 효율적 동시성 구현' if concurrent_executions > 0 else '순차적 실행',
        '해석': '의도된 정상 동작 (오류 아님)'
    }
    
    print(f"  동시 실행 발생: {concurrent_executions}건 (최대 {max_concurrent_level}개 스레드 동시)")
    print("✅ 동시 실행 패턴 관찰 완료")
    return result

def analyze_capacity_prevention_verification(preprocessor_df, analysis_df, total_info):
    """분석 3: 정원 초과 방지 검증 (기존 규칙 3 유지) - 성공/실패 판정"""
    print("🔍 분석 3: 정원 초과 방지 검증 (규칙 3 유지) 중...")
    
    # anomaly_type이 NaN인 경우 빈 문자열로 처리
    analysis_df['anomaly_type'] = analysis_df['anomaly_type'].fillna('')
    
    # 정원 초과 오류 필터링
    capacity_exceeded = analysis_df[analysis_df['anomaly_type'].str.contains('정원 초과 오류', na=False)]
    capacity_exceeded_count = len(capacity_exceeded)
    total_requests = total_info['total_requests']
    
    # 정원 초과 통계
    if capacity_exceeded_count > 0:
        avg_exceeded_amount = capacity_exceeded['over_capacity_amount'].mean()
        max_exceeded_amount = capacity_exceeded['over_capacity_amount'].max()
        affected_rooms = capacity_exceeded['roomNumber'].nunique()
        affected_bins = len(capacity_exceeded.groupby(['roomNumber', 'bin']))
    else:
        avg_exceeded_amount = 0
        max_exceeded_amount = 0
        affected_rooms = 0
        affected_bins = 0
    
    # 결과 생성
    result = {
        '분석 구분': '정원 초과 방지 검증 (규칙 3 유지)',
        '전체 permit 요청수': total_requests,
        '전체 방 수': total_info['total_rooms'],
        '전체 (방×bin) 조합수': total_info['total_bins'],
        '정원 초과 발생 건수': capacity_exceeded_count,
        '정원 초과 발생률 (%)': round((capacity_exceeded_count / total_requests * 100), 2) if total_requests > 0 else 0,
        '정원 초과 방지 성공률 (%)': round(((total_requests - capacity_exceeded_count) / total_requests * 100), 2) if total_requests > 0 else 100,
        '영향받은 방 수': affected_rooms,
        '영향받은 (방×bin) 조합수': affected_bins,
        '평균 초과 인원': round(avg_exceeded_amount, 2),
        '최대 초과 인원': max_exceeded_amount,
        '검증 결과': '완벽한 방지 성공' if capacity_exceeded_count == 0 else '방지 실패',
        '비즈니스 규칙 준수': '✅ 완벽 준수' if capacity_exceeded_count == 0 else '❌ 규칙 위반'
    }
    
    print(f"  정원 초과 발생: {capacity_exceeded_count}건 ({result['정원 초과 발생률 (%)']}%)")
    print("✅ 정원 초과 방지 검증 완료")
    return result

def create_semaphore_dataframes(sequential_result, concurrent_result, capacity_result):
    """세마포어 3개 분석 결과를 각각의 DataFrame으로 생성"""
    print("📊 세마포어 3개 분석별 개별 DataFrame 생성 중...")
    print("  - 분석 1+2: 현상 관찰 (위반/오류 아님)")
    print("  - 분석 3: 효과성 검증 (성공/실패 판정)")
    
    sequential_df = pd.DataFrame([sequential_result])
    concurrent_df = pd.DataFrame([concurrent_result])
    capacity_df = pd.DataFrame([capacity_result])
    
    print("✅ 세마포어 개별 DataFrame 생성 완료")
    return sequential_df, concurrent_df, capacity_df

def add_semaphore_dataframe_to_sheet(ws, df, sheet_title):
    """세마포어 분석 결과를 워크시트에 추가 (단순화된 스타일)"""
    # 시트 제목 추가 (단순화)
    ws['A1'] = sheet_title
    ws['A1'].font = Font(size=14, bold=True)
    
    # 시작 행
    start_row = 3
    
    # 컬럼 헤더 추가 (스타일 단순화)
    for col_idx, column in enumerate(df.columns, 1):
        # 컬럼명 길이 제한 및 특수문자 제거
        safe_column = str(column).replace('%', 'percent').replace('(', '').replace(')', '')[:30]
        cell = ws.cell(row=start_row, column=col_idx, value=safe_column)
        cell.font = Font(bold=True)
    
    # 데이터 추가 (스타일 최소화)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            # 데이터 값 안전성 처리
            safe_value = value
            if isinstance(value, str):
                safe_value = str(value).replace('%', 'percent')[:100]  # 길이 제한
            elif pd.isna(value):
                safe_value = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
    
    # 컬럼 너비 기본 설정
    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

def create_semaphore_excel_output(sequential_df, concurrent_df, capacity_df, output_file):
    """세마포어 3개 시트로 구성된 Excel 파일 생성 (안전한 방식)"""
    print("📊 세마포어 Excel 파일 생성 중...")
    
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 시트명 단순화
    # 시트 1: 순차적 일관성 관찰
    ws1 = wb.create_sheet("Overall_Sequential_Analysis")
    add_semaphore_dataframe_to_sheet(ws1, sequential_df, "Overall Analysis 1: Sequential Consistency")
    
    # 시트 2: 동시 실행 패턴 관찰 (시트명 변경: 기존과 통일)
    ws2 = wb.create_sheet("Overall_Contention_Analysis") 
    add_semaphore_dataframe_to_sheet(ws2, concurrent_df, "Overall Analysis 2: Contention")
    
    # 시트 3: 정원 초과 방지 검증
    ws3 = wb.create_sheet("Overall_Capacity_Analysis")
    add_semaphore_dataframe_to_sheet(ws3, capacity_df, "Overall Analysis 3: Capacity Prevention")
    
    wb.save(output_file)
    print(f"✅ 세마포어 Excel 파일 저장 완료: {output_file}")
    print("  📋 생성된 시트:")
    print("    - Overall_Sequential_Analysis: 순차적 일관성 관찰")
    print("    - Overall_Contention_Analysis: 동시 실행 패턴 관찰 (기존과 통일)") 
    print("    - Overall_Capacity_Analysis: 정원 초과 방지 검증")

def print_semaphore_summary_statistics(sequential_df, concurrent_df, capacity_df):
    """세마포어 분석 결과 요약 통계 출력"""
    print("\n" + "="*90)
    print("📈 세마포어 전체 통합 분석 결과 요약")
    print("="*90)
    
    # 각 DataFrame에서 정보 추출
    dataframes = [
        ("순차적 일관성 관찰 (규칙 1+4 통합)", sequential_df),
        ("동시 실행 패턴 관찰 (규칙 2 재정의)", concurrent_df),
        ("정원 초과 방지 검증 (규칙 3 유지)", capacity_df)
    ]
    
    # 첫 번째 분석에서 전체 정보 추출
    if len(sequential_df) > 0:
        first_row = sequential_df.iloc[0]
        total_requests = first_row['전체 permit 요청수']
        total_rooms = first_row['전체 방 수']
        total_bins = first_row['전체 (방×bin) 조합수']
        
        print(f"전체 세마포어 분석 대상:")
        print(f"  - permit 요청 수: {total_requests:,}건")
        print(f"  - 방 수: {total_rooms}개")
        print(f"  - (방×bin) 조합: {total_bins}개")
    
    print(f"\n--- 세마포어 3가지 분석 결과 ---")
    print(f"📋 분석 1+2: 현상 관찰 (위반/오류 아님)")
    print(f"📋 분석 3: 성공/실패 검증")
    
    for analysis_name, df in dataframes:
        if len(df) > 0:
            row = df.iloc[0]
            print(f"\n🎯 {analysis_name}:")
            
            if '순차적 일관성' in analysis_name:
                differences = row['순차적 일관성 차이 발생 건수']
                difference_rate = row['순차적 일관성 차이 발생률 (%)']
                consistency_rate = row['순차적 처리 일치율 (%)']
                characteristic = row['관찰된 특성']
                
                print(f"  - 순차적 일관성 차이: {differences:,}건 ({difference_rate}%)")
                print(f"  - 순차적 처리 일치율: {consistency_rate}%")
                print(f"  - 관찰된 특성: {characteristic}")
                
            elif '동시 실행' in analysis_name:
                concurrent_count = row['발생 건수']
                concurrent_rate = row['발생률 (%)']
                max_level = row['최대 경합 스레드 수']
                interpretation = row['해석']
                
                print(f"  - 실제 경합 발생: {concurrent_count:,}건 ({concurrent_rate}%)")
                print(f"  - 최대 동시 실행 수준: {max_level}개 스레드")
                print(f"  - 해석: {interpretation}")
                
            elif '정원 초과' in analysis_name:
                exceeded_count = row['정원 초과 발생 건수']
                exceeded_rate = row['정원 초과 발생률 (%)']
                prevention_rate = row['정원 초과 방지 성공률 (%)']
                verification_result = row['검증 결과']
                business_rule = row['비즈니스 규칙 준수']
                
                print(f"  - 정원 초과 발생: {exceeded_count:,}건 ({exceeded_rate}%)")
                print(f"  - 정원 초과 방지 성공률: {prevention_rate}%")
                print(f"  - 검증 결과: {verification_result}")
                print(f"  - 비즈니스 규칙 준수: {business_rule}")
        else:
            print(f"\n{analysis_name}: 데이터 없음")
    
    print(f"\n--- 세마포어 종합 평가 ---")
    if len(capacity_df) > 0:
        capacity_row = capacity_df.iloc[0]
        if capacity_row['정원 초과 발생 건수'] == 0:
            print("🎉 세마포어 핵심 기능 (정원 초과 방지): 완벽 성공")
        else:
            print("⚠️ 세마포어 핵심 기능 (정원 초과 방지): 검증 실패")
    
    if len(sequential_df) > 0:
        sequential_row = sequential_df.iloc[0]
        if sequential_row['순차적 일관성 차이 발생 건수'] > 0:
            print("📝 순차적 일관성: 부분적 보장 특성 관찰 (세마포어 고유 특성)")
        else:
            print("📝 순차적 일관성: 완전 순차적 처리 관찰")
    
    if len(concurrent_df) > 0:
        concurrent_row = concurrent_df.iloc[0]
        if concurrent_row['발생 건수'] > 0:
            print("🚀 동시성 활용: CAS 기반 효율적 동시 실행 관찰 (의도된 정상 동작)")
        else:
            print("🚀 동시성 활용: 순차적 실행 관찰")

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description="세마포어 전용 전체 통합 통계 분석기")
    parser.add_argument('preprocessor_csv', help='세마포어 전처리 결과 CSV 파일 (preprocessor_semaphore.csv)')
    parser.add_argument('analysis_csv', help='세마포어 분석 결과 CSV 파일 (semaphore_analysis_result.csv)')
    parser.add_argument('output_xlsx', help='세마포어 전체 통합 분석 Excel 출력 파일')
    parser.add_argument('--rooms', help='분석할 방 번호 (쉼표로 구분)')
    
    args = parser.parse_args()
    
    try:
        print("🚀 세마포어 전용 전체 통합 분석기 시작...")
        print("🎯 세마포어 3가지 핵심 분석: 순차적 일관성 관찰 + 동시 실행 패턴 관찰 + 정원 초과 방지 검증")
        print(f"입력 파일 1: {args.preprocessor_csv}")
        print(f"입력 파일 2: {args.analysis_csv}")
        print(f"출력 파일: {args.output_xlsx}")
        
        # 1. 세마포어 데이터 로드 및 검증
        preprocessor_df, analysis_df = load_and_validate_semaphore_data(args.preprocessor_csv, args.analysis_csv)
        
        # 2. 방 번호 필터링 (선택사항)
        if args.rooms:
            room_numbers = [int(room.strip()) for room in args.rooms.split(',')]
            preprocessor_df = preprocessor_df[preprocessor_df['roomNumber'].isin(room_numbers)]
            analysis_df = analysis_df[analysis_df['roomNumber'].isin(room_numbers)]
            print(f"🔍 방 번호 {room_numbers}로 필터링 적용")
        
        # 3. 세마포어 전체 요청 정보 집계
        total_info = calculate_semaphore_total_info(preprocessor_df)
        
        # 4. 세마포어 3가지 핵심 분석 실행
        print("\n=== 🎯 세마포어 3가지 분석 실행 ===")
        print("📋 분석 1+2: 현상 관찰 (위반/오류가 아닌 특성 관찰)")
        print("📋 분석 3: 효과성 검증 (유일한 성공/실패 판정)")
        
        # 세마포어 3가지 분석 실행
        sequential_result = analyze_sequential_consistency_observation(preprocessor_df, analysis_df, total_info)
        concurrent_result = analyze_concurrent_execution_observation(preprocessor_df, analysis_df, total_info)
        capacity_result = analyze_capacity_prevention_verification(preprocessor_df, analysis_df, total_info)
        
        # 5. 개별 DataFrame 생성
        sequential_df, concurrent_df, capacity_df = create_semaphore_dataframes(
            sequential_result, concurrent_result, capacity_result)
        
        # 6. Excel 파일 생성
        create_semaphore_excel_output(sequential_df, concurrent_df, capacity_df, args.output_xlsx)
        
        # 7. 요약 통계 출력
        print_semaphore_summary_statistics(sequential_df, concurrent_df, capacity_df)
        
        print("\n🎉 세마포어 전체 통합 분석 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()