#!/usr/bin/env python3
"""
Race Condition 통계 분석기
- 이상현상 탐지 결과를 바탕으로 4가지 규칙별 통계적 집계 분석
- bin 단위 발생률, 심각도, 분포 분석 제공
"""

import pandas as pd
import numpy as np
from datetime import datetime
import argparse
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

def load_and_validate_data(preprocessor_file, analysis_file):
    """데이터 로드 및 필수 컬럼 검증"""
    print("📂 데이터 파일 로드 중...")
    
    # 전처리 데이터 로드
    preprocessor_df = pd.read_csv(preprocessor_file)
    print(f"✅ 전처리 데이터 로드 완료: {len(preprocessor_df)}행")
    
    # 이상현상 분석 데이터 로드
    analysis_df = pd.read_csv(analysis_file)
    print(f"✅ 이상현상 분석 데이터 로드 완료: {len(analysis_df)}행")
    
    # 전처리 데이터 필수 컬럼 검증
    preprocessor_required = ['roomNumber', 'bin', 'user_id']
    missing_preprocessor = [col for col in preprocessor_required if col not in preprocessor_df.columns]
    if missing_preprocessor:
        raise ValueError(f"전처리 데이터에서 필수 컬럼 누락: {missing_preprocessor}")
    
    # 이상현상 분석 데이터 필수 컬럼 검증
    analysis_required = ['roomNumber', 'bin', 'anomaly_type', 'lost_update_diff', 
                        'contention_group_size', 'over_capacity_amount', 'curr_sequence_diff']
    missing_analysis = [col for col in analysis_required if col not in analysis_df.columns]
    if missing_analysis:
        raise ValueError(f"이상현상 분석 데이터에서 필수 컬럼 누락: {missing_analysis}")
    
    # 데이터 타입 정리 (NaN/무한대 값 처리)
    preprocessor_df['roomNumber'] = pd.to_numeric(preprocessor_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    preprocessor_df['bin'] = pd.to_numeric(preprocessor_df['bin'], errors='coerce').fillna(0).astype(int)
    
    analysis_df['roomNumber'] = pd.to_numeric(analysis_df['roomNumber'], errors='coerce').fillna(0).astype(int)
    analysis_df['bin'] = pd.to_numeric(analysis_df['bin'], errors='coerce').fillna(0).astype(int)
    
    print("✅ 데이터 검증 및 타입 변환 완료")
    return preprocessor_df, analysis_df

def calculate_total_requests_per_bin(preprocessor_df):
    """방별 bin 단위 전체 요청수 집계"""
    print("📊 방별 bin 단위 전체 요청수 집계 중...")
    
    total_requests = preprocessor_df.groupby(['roomNumber', 'bin']).size().reset_index(name='total_requests')
    
    print(f"✅ 집계 완료: {len(total_requests)}개 (방×bin) 조합")
    return total_requests

def calculate_statistics(filtered_df, value_column, total_requests_df, use_absolute=False):
    """공통 통계 계산 함수 - 모든 bin 구간 포함"""
    # 전체 (roomNumber, bin) 조합을 기준으로 결과 생성
    result_stats = total_requests_df.copy()
    
    # 통계 컬럼들 초기화 (이상현상 없는 bin은 0 또는 NaN)
    result_stats['occurrence_count'] = 0
    result_stats['occurrence_rate'] = 0.0
    result_stats['sum_value'] = 0.0
    result_stats['avg_value'] = np.nan
    result_stats['min_value'] = np.nan
    result_stats['max_value'] = np.nan
    result_stats['median_value'] = np.nan
    result_stats['std_value'] = 0.0
    
    # 이상현상이 있는 경우에만 실제 통계 계산
    if len(filtered_df) > 0:
        # (roomNumber, bin) 단위로 그룹화
        grouped = filtered_df.groupby(['roomNumber', 'bin'])
        
        for (room_num, bin_num), group in grouped:
            # 해당 값들 추출 (NaN 제거)
            values = group[value_column].dropna()
            
            if len(values) == 0:
                continue
            
            # 결과 DataFrame에서 해당 row 찾기
            mask = (result_stats['roomNumber'] == room_num) & (result_stats['bin'] == bin_num)
            row_idx = result_stats[mask].index
            
            if len(row_idx) == 0:
                continue
                
            row_idx = row_idx[0]
            total_requests = result_stats.loc[row_idx, 'total_requests']
            
            # 통계 계산
            occurrence_count = len(values)
            occurrence_rate = (occurrence_count / total_requests * 100) if total_requests > 0 else 0
            
            # 결과 업데이트 (절댓값 옵션 적용)
            result_stats.loc[row_idx, 'occurrence_count'] = int(occurrence_count)
            result_stats.loc[row_idx, 'occurrence_rate'] = round(occurrence_rate, 2)
            
            if use_absolute:
                result_stats.loc[row_idx, 'sum_value'] = round(values.abs().sum(), 2)
                result_stats.loc[row_idx, 'avg_value'] = round(values.abs().mean(), 2)
            else:
                result_stats.loc[row_idx, 'sum_value'] = round(values.sum(), 2)
                result_stats.loc[row_idx, 'avg_value'] = round(values.mean(), 2)
            
            result_stats.loc[row_idx, 'min_value'] = round(values.min(), 2)
            result_stats.loc[row_idx, 'max_value'] = round(values.max(), 2)
            result_stats.loc[row_idx, 'median_value'] = round(values.median(), 2)
            result_stats.loc[row_idx, 'std_value'] = round(values.std(), 4) if len(values) > 1 else 0.0
    
    # 데이터 타입 정리
    result_stats['roomNumber'] = result_stats['roomNumber'].astype(int)
    result_stats['bin'] = result_stats['bin'].astype(int)
    result_stats['total_requests'] = result_stats['total_requests'].astype(int)
    result_stats['occurrence_count'] = result_stats['occurrence_count'].fillna(0).astype(int)
    
    return result_stats

def analyze_lost_update(analysis_df, total_requests_df):
    """규칙 1: 값 불일치 통계 분석"""
    print("🔍 규칙 1: 값 불일치 (Lost Update) 분석 중...")
    
    # '값 불일치' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].fillna('').str.contains('값 불일치', na=False)]
    print(f"  - 값 불일치 발생 레코드: {len(filtered_df)}건")
    
    # lost_update_diff 기준 통계 계산
    stats_df = calculate_statistics(filtered_df, 'lost_update_diff', total_requests_df)
    
    # 컬럼명 변경
    column_mapping = {
        'sum_value': '오차 누적 총합',
        'avg_value': '평균 오차',
        'min_value': '최소 오차',
        'max_value': '최대 오차',
        'median_value': '중위 값',
        'std_value': '오차 표준편차'
    }
    
    final_columns = {
        'roomNumber': '방 번호',
        'bin': 'Bin',
        'total_requests': '전체 요청수',
        'occurrence_count': '발생 건수',
        'occurrence_rate': '발생률 (%)',
        **column_mapping
    }
    
    stats_df = stats_df.rename(columns=final_columns)
    
    print(f"✅ 값 불일치 분석 완료: {len(stats_df)}개 (방×bin) 조합")
    return stats_df

def analyze_contention(analysis_df, total_requests_df):
    """규칙 2: 경합 발생 통계 분석"""
    print("🔍 규칙 2: 경합 발생 (Contention) 분석 중...")
    
    # '경합 발생' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].fillna('').str.contains('경합 발생 오류', na=False)]
    print(f"  - 경합 발생 레코드: {len(filtered_df)}건")
    
    # contention_group_size 기준 통계 계산
    stats_df = calculate_statistics(filtered_df, 'contention_group_size', total_requests_df)
    
    # 컬럼명 변경
    column_mapping = {
        'sum_value': '총 경합 스레드 수',
        'avg_value': '평균 경합 스레드 수',
        'min_value': '최소 경합 그룹 크기',
        'max_value': '최대 경합 스레드 수',
        'median_value': '중간값 경합 그룹 크기',
        'std_value': '경합 강도 표준편차'
    }
    
    final_columns = {
        'roomNumber': '방 번호',
        'bin': 'Bin',
        'total_requests': '전체 요청수',
        'occurrence_count': '발생 건수',
        'occurrence_rate': '발생률 (%)',
        **column_mapping
    }
    
    stats_df = stats_df.rename(columns=final_columns)
    
    print(f"✅ 경합 발생 분석 완료: {len(stats_df)}개 (방×bin) 조합")
    return stats_df

def analyze_capacity_exceeded(analysis_df, total_requests_df):
    """규칙 3: 정원 초과 통계 분석"""
    print("🔍 규칙 3: 정원 초과 (Capacity Exceeded) 분석 중...")
    
    # '정원 초과' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].fillna('').str.contains('정원 초과 오류', na=False)]
    print(f"  - 정원 초과 발생 레코드: {len(filtered_df)}건")
    
    # over_capacity_amount 기준 통계 계산
    stats_df = calculate_statistics(filtered_df, 'over_capacity_amount', total_requests_df)
    
    # 컬럼명 변경
    column_mapping = {
        'sum_value': '총 초과 인원',
        'avg_value': '평균 초과 인원',
        'min_value': '최소 초과 인원',
        'max_value': '최대 초과 인원',
        'median_value': '중간값 초과 인원',
        'std_value': '초과 규모 표준편차'
    }
    
    final_columns = {
        'roomNumber': '방 번호',
        'bin': 'Bin',
        'total_requests': '전체 요청수',
        'occurrence_count': '발생 건수',
        'occurrence_rate': '발생률 (%)',
        **column_mapping
    }
    
    stats_df = stats_df.rename(columns=final_columns)
    
    print(f"✅ 정원 초과 분석 완료: {len(stats_df)}개 (방×bin) 조합")
    return stats_df

def analyze_state_transition(analysis_df, total_requests_df):
    """규칙 4: 상태 전이 오류 통계 분석"""
    print("🔍 규칙 4: 상태 전이 오류 (State Transition) 분석 중...")
    
    # '상태 전이' 포함된 레코드 필터링
    filtered_df = analysis_df[analysis_df['anomaly_type'].fillna('').str.contains('상태 전이 오류', na=False)]
    print(f"  - 상태 전이 오류 발생 레코드: {len(filtered_df)}건")
    
    # curr_sequence_diff 기준 통계 계산 (절댓값 사용)
    stats_df = calculate_statistics(filtered_df, 'curr_sequence_diff', total_requests_df, use_absolute=True)
    
    # 컬럼명 변경
    column_mapping = {
        'sum_value': '총합 값',
        'avg_value': '평균 값',
        'min_value': '최소 깂',
        'max_value': '최대 깂',
        'median_value': '중간 값',
        'std_value': '표준편차 값'
    }
    
    final_columns = {
        'roomNumber': '방 번호',
        'bin': 'Bin',
        'total_requests': '전체 요청수',
        'occurrence_count': '발생 건수',
        'occurrence_rate': '발생률 (%)',
        **column_mapping
    }
    
    stats_df = stats_df.rename(columns=final_columns)
    
    print(f"✅ 상태 전이 오류 분석 완료: {len(stats_df)}개 (방×bin) 조합")
    return stats_df

def add_dataframe_to_sheet(ws, df, sheet_title):
    """DataFrame을 워크시트에 추가하고 스타일 적용"""
    # 시트 제목 추가
    ws['A1'] = sheet_title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    
    # 빈 행 추가
    start_row = 3
    
    # 컬럼 헤더 추가
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=column)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 추가
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 컬럼 너비 조정
    for col_idx, column in enumerate(df.columns, 1):
        max_length = max(len(str(column)), 15)
        if len(df) > 0:
            values = [len(str(val)) for val in df.iloc[:, col_idx-1] if pd.notna(val)]
            if values:
                max_length = max(max_length, max(values))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)



def create_excel_output(lost_update_df, contention_df, capacity_df, state_transition_df, output_file):
    """4개 시트로 구성된 Excel 파일 생성"""
    print("📊 Excel 파일 생성 중...")
    
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 시트 1: Lost Update 분석
    ws1 = wb.create_sheet("Lost_Update_Analysis")
    add_dataframe_to_sheet(ws1, lost_update_df, "규칙 1: 값 불일치 (Lost Update) 분석")
    
    # 시트 2: Contention 분석
    ws2 = wb.create_sheet("Contention_Analysis")
    add_dataframe_to_sheet(ws2, contention_df, "규칙 2: 경합 발생 (Contention) 분석")
    
    # 시트 3: Capacity Exceeded 분석
    ws3 = wb.create_sheet("Capacity_Exceeded_Analysis")
    add_dataframe_to_sheet(ws3, capacity_df, "규칙 3: 정원 초과 (Capacity Exceeded) 분석")
    
    # 시트 4: State Transition 분석
    ws4 = wb.create_sheet("State_Transition_Analysis")
    add_dataframe_to_sheet(ws4, state_transition_df, "규칙 4: 상태 전이 오류 (State Transition) 분석")
    
    wb.save(output_file)
    print(f"✅ Excel 파일 저장 완료: {output_file}")

def print_summary_statistics(lost_update_df, contention_df, capacity_df, state_transition_df, total_requests_df):
    """분석 결과 요약 통계 출력"""
    print("\n" + "="*80)
    print("📈 RACE CONDITION 통계 분석 결과 요약")
    print("="*80)
    
    total_bins = len(total_requests_df)
    total_requests_sum = total_requests_df['total_requests'].sum()
    
    print(f"전체 분석 대상: {total_bins}개 (방×bin) 조합")
    print(f"전체 요청 수: {total_requests_sum:,}건")
    
    print(f"\n--- 규칙별 분석 결과 ---")
    
    # 규칙 1
    lost_bins = len(lost_update_df)
    lost_requests = lost_update_df['발생 건수'].sum() if len(lost_update_df) > 0 else 0
    print(f"규칙 1 (값 불일치): {lost_bins}개 bin에서 {lost_requests}건 발생")
    
    # 규칙 2  
    contention_bins = len(contention_df)
    contention_requests = contention_df['발생 건수'].sum() if len(contention_df) > 0 else 0
    print(f"규칙 2 (경합 발생): {contention_bins}개 bin에서 {contention_requests}건 발생")
    
    # 규칙 3
    capacity_bins = len(capacity_df)
    capacity_requests = capacity_df['발생 건수'].sum() if len(capacity_df) > 0 else 0
    print(f"규칙 3 (정원 초과): {capacity_bins}개 bin에서 {capacity_requests}건 발생")
    
    # 규칙 4
    state_bins = len(state_transition_df)
    state_requests = state_transition_df['발생 건수'].sum() if len(state_transition_df) > 0 else 0
    print(f"규칙 4 (상태 전이): {state_bins}개 bin에서 {state_requests}건 발생")
    
    total_anomaly_requests = lost_requests + contention_requests + capacity_requests + state_requests
    anomaly_rate = (total_anomaly_requests / total_requests_sum * 100) if total_requests_sum > 0 else 0
    
    print(f"\n전체 이상현상 발생률: {anomaly_rate:.2f}% ({total_anomaly_requests:,}/{total_requests_sum:,})")

def main():
    """메인 함수"""
    parser = argparse.ArgumentParser(description="Race Condition 통계 분석기")
    parser.add_argument('preprocessor_csv', help='전처리 결과 CSV 파일')
    parser.add_argument('analysis_csv', help='이상현상 분석 결과 CSV 파일')
    parser.add_argument('output_xlsx', help='통계 분석 Excel 출력 파일')
    parser.add_argument('--rooms', help='분석할 방 번호 (쉼표로 구분)')
    
    args = parser.parse_args()
    
    try:
        print("🚀 Race Condition 통계 분석기 시작...")
        print(f"입력 파일 1: {args.preprocessor_csv}")
        print(f"입력 파일 2: {args.analysis_csv}")
        print(f"출력 파일: {args.output_xlsx}")
        
        # 1. 데이터 로드 및 검증
        preprocessor_df, analysis_df = load_and_validate_data(args.preprocessor_csv, args.analysis_csv)
        
        # 2. 방 번호 필터링 (선택사항)
        if args.rooms:
            room_numbers = [int(room.strip()) for room in args.rooms.split(',')]
            preprocessor_df = preprocessor_df[preprocessor_df['roomNumber'].isin(room_numbers)]
            analysis_df = analysis_df[analysis_df['roomNumber'].isin(room_numbers)]
            print(f"🔍 방 번호 {room_numbers}로 필터링 적용")
        
        # 3. 전체 요청수 집계
        total_requests_df = calculate_total_requests_per_bin(preprocessor_df)
        
        # 4. 4가지 규칙별 통계 분석
        lost_update_df = analyze_lost_update(analysis_df, total_requests_df)
        contention_df = analyze_contention(analysis_df, total_requests_df)
        capacity_df = analyze_capacity_exceeded(analysis_df, total_requests_df)
        state_transition_df = analyze_state_transition(analysis_df, total_requests_df)
        
        # 5. Excel 파일 생성
        create_excel_output(lost_update_df, contention_df, capacity_df, state_transition_df, args.output_xlsx)
        
        # 6. 요약 통계 출력
        print_summary_statistics(lost_update_df, contention_df, capacity_df, state_transition_df, total_requests_df)
        
        print("\n🎉 통계 분석 완료!")
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()