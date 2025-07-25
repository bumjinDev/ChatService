#!/usr/bin/env python3
"""
임계 영역 성능 측정 로그 전처리 스크립트 (비이중 확인 구조용)

[목적]
멀티스레드 환경에서 비이중 확인 구조의 임계 영역(Critical Section) 접근 로그를 파싱하여
성능 분석용 데이터로 변환합니다.

[주요 기능]
1. 로그 파일에서 5가지 이벤트 추출 (WAITING_START, CRITICAL_ENTER, CRITICAL_LEAVE, INCREMENT_BEFORE, INCREMENT_AFTER)
2. 사용자별, 방별로 이벤트 그룹화
3. 단순화된 정렬: waiting_start_nanoTime → critical_enter_nanoTime
4. CSV 및 Excel 형식으로 결과 저장
5. 사용자 지정 출력 디렉토리 지원

[이벤트 의미]
- WAITING_START: 임계 영역 진입 대기 시작
- CRITICAL_ENTER: 임계 영역 진입 성공
- CRITICAL_LEAVE: 임계 영역 퇴장
- INCREMENT_BEFORE: 카운터 증가 작업 시작
- INCREMENT_AFTER: 카운터 증가 작업 완료

[비이중 확인 구조 특징]
- PRE_CHECK_FAIL_OVER_CAPACITY 이벤트 없음 (락 외부 사전 확인 없음)
- 5개 이벤트만 처리
- SUCCESS/FAIL_OVER_CAPACITY/UNKNOWN 3가지 결과만 분류

[수정 사항]
- PRE_CHECK_FAIL 관련 모든 로직 제거
- 5개 이벤트 기반 단순 처리
- waiting_start_nanoTime → critical_enter_nanoTime 순 단순 정렬
- pandas 호환성 문제 수정 (is_datetime64_tz_dtype → pd.api.types.is_datetime64tz_dtype)
"""

import pandas as pd
import re
import os
import shutil
import argparse
import sys
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook

# ===== 상수 정의 =====
# 파일 경로 상수
LOG_FILE = 'ChatService.log'
NEW_LOG_PATH = r'E:\devSpace\ChatServiceTest\log\ChatService.log'

# 이벤트 타입 상수
EVENT_WAITING_START = 'WAITING_START'
EVENT_CRITICAL_ENTER = 'CRITICAL_ENTER'
EVENT_CRITICAL_LEAVE = 'CRITICAL_LEAVE'
EVENT_INCREMENT_BEFORE = 'INCREMENT_BEFORE'
EVENT_INCREMENT_AFTER = 'INCREMENT_AFTER'

# 결과 타입 상수
RESULT_SUCCESS = 'SUCCESS'
RESULT_FAIL_CAPACITY = 'FAIL_OVER_CAPACITY'
RESULT_UNKNOWN = 'UNKNOWN'

# 분석 구간 수
BINS_COUNT = 10

# 정규 표현식 패턴
CRITICAL_PATTERN = re.compile(
    r'CRITICAL_SECTION_MARK tag=(?P<tag>WAITING_START|CRITICAL_ENTER|CRITICAL_LEAVE)'
    r' timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<event>\w+)'
    r'.* roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r'.* nanoTime=(?P<nanoTime>\d+)'
    r'.* epochNano=(?P<epochNano>\d+)'
)

INCREMENT_PATTERN = re.compile(
    r'timestampIso=(?P<timestamp>[\w\-\:\.TZ]+)'
    r' event=(?P<tag>INCREMENT_BEFORE|INCREMENT_AFTER)'
    r' roomNumber=(?P<roomNumber>\d+)'
    r' userId=(?P<userId>[\w\-\_]+)'
    r'.* epochNano=(?P<epochNano>\d+)'
    r'.* nanoTime=(?P<nanoTime>\d+)'
)


def test_critical_pattern():
    """CRITICAL_PATTERN 정규식 테스트 함수"""
    test_logs = [
        "CRITICAL_SECTION_MARK tag=WAITING_START timestampIso=2025-07-19T02:24:37.452142100Z event=PRE_JOIN_ATTEMPT className=RoomJoinServiceSynchronized methodName=confirmJoinRoom roomNumber=1203 userId=yhjj875 nanoTime=38434598060100 epochNano=1752891877453106700 threadId=247",
        "CRITICAL_SECTION_MARK tag=CRITICAL_ENTER timestampIso=2025-07-19T02:24:37.452142100Z event=CRITICAL_ENTER_EVENT className=RoomJoinServiceSynchronized methodName=confirmJoinRoom roomNumber=1203 userId=yhjj875 nanoTime=38434598060200 epochNano=1752891877453106800 threadId=247",
        "CRITICAL_SECTION_MARK tag=CRITICAL_LEAVE timestampIso=2025-07-19T02:24:37.452142100Z event=SUCCESS className=RoomJoinServiceSynchronized methodName=confirmJoinRoom roomNumber=1203 userId=yhjj875 nanoTime=38434598060300 epochNano=1752891877453106900 threadId=247"
    ]
    
    success_count = 0
    for i, test_log in enumerate(test_logs):
        match = CRITICAL_PATTERN.search(test_log)
        if match:
            result = match.groupdict()
            print(f"✅ 테스트 {i+1} 매칭 성공: tag={result['tag']}, event={result['event']}")
            success_count += 1
        else:
            print(f"❌ 테스트 {i+1} 매칭 실패")
    
    return success_count == len(test_logs)


def replace_log_file() -> None:
    """
    기존 로그 파일을 새 로그 파일로 교체
    """
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)
    shutil.copy(NEW_LOG_PATH, LOG_FILE)
    print(f"로그 파일 교체 완료: {NEW_LOG_PATH} → {LOG_FILE}")


def parse_log_line(line: str) -> Optional[Dict[str, Any]]:
    """
    로그 라인을 파싱하여 이벤트 정보 추출
    """
    # CRITICAL_SECTION_MARK 패턴 매칭
    match = CRITICAL_PATTERN.search(line)
    if match:
        data = match.groupdict()
        data['event_type'] = data['event']
        return data
    
    # INCREMENT 이벤트 패턴 매칭
    match = INCREMENT_PATTERN.search(line)
    if match:
        data = match.groupdict()
        data['event_type'] = None
        return data
    
    return None


def parse_five_events_clean(filepath: str, room_number: Optional[int] = None) -> pd.DataFrame:
    """
    로그 파일에서 5가지 이벤트를 파싱하여 DataFrame으로 변환 (비이중 확인 구조용)
    """
    records = []
    
    try:
        with open(filepath, encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                data = parse_log_line(line)
                
                if data:
                    # 방 번호를 정수로 변환
                    data['roomNumber'] = int(data['roomNumber'])
                    
                    # 나노초 값들을 문자열로 유지
                    data['nanoTime'] = data['nanoTime']
                    data['epochNano'] = data['epochNano']
                    
                    # 방 번호 필터링
                    if room_number is None or data['roomNumber'] == room_number:
                        records.append(data)
    
    except FileNotFoundError:
        print(f"오류: 로그 파일을 찾을 수 없습니다 - {filepath}")
        return pd.DataFrame()
    except Exception as e:
        print(f"오류: 로그 파일 파싱 실패 - {e}")
        return pd.DataFrame()
    
    return pd.DataFrame(records)


def process_user_events(user_id: str, user_data: pd.DataFrame) -> Dict[str, Any]:
    """
    단일 사용자의 이벤트들을 처리하여 성능 프로필 생성
    """
    # nanoTime 기준으로 정렬
    user_data = user_data.sort_values(['nanoTime_num']).reset_index(drop=True)
    
    # 이벤트별로 그룹화
    events = {}
    for _, row in user_data.iterrows():
        event_name = row['tag']
        events[event_name] = {
            'timestamp': row['timestamp'],
            'nanoTime': str(int(float(row['nanoTime']))),
            'epochNano': str(int(float(row['epochNano']))),
            'event_type': row.get('event_type')
        }
    
    # 기본 프로필 정보
    profile = {
        'user_id': user_id,
        'roomNumber': user_data.iloc[0]['roomNumber']
    }
    
    # 각 이벤트별 정보 추가
    event_mappings = {
        EVENT_WAITING_START: 'waiting_start',
        EVENT_CRITICAL_ENTER: 'critical_enter',
        EVENT_CRITICAL_LEAVE: 'critical_leave',
        EVENT_INCREMENT_BEFORE: 'increment_before',
        EVENT_INCREMENT_AFTER: 'increment_after'
    }
    
    for event_tag, prefix in event_mappings.items():
        if event_tag in events:
            event_data = events[event_tag]
            profile.update({
                f'{prefix}_time': event_data['timestamp'],
                f'{prefix}_nanoTime': event_data['nanoTime'],
                f'{prefix}_epochNano': event_data['epochNano']
            })
            
            # event_type이 있는 경우 추가
            if event_data['event_type'] is not None:
                profile[f'{prefix}_event_type'] = event_data['event_type']
    
    # join_result 설정 (CRITICAL_LEAVE의 event_type 기반)
    if EVENT_CRITICAL_LEAVE in events:
        leave_type = events[EVENT_CRITICAL_LEAVE]['event_type']
        if leave_type == RESULT_SUCCESS:
            profile['join_result'] = RESULT_SUCCESS
        elif leave_type == RESULT_FAIL_CAPACITY:
            profile['join_result'] = RESULT_FAIL_CAPACITY
        else:
            profile['join_result'] = RESULT_UNKNOWN
    
    return profile


def assign_bins_and_sequence_after_first_sort(df: pd.DataFrame) -> pd.DataFrame:
    """
    1차 정렬 직후에 정렬된 순서를 기준으로 bin과 room_entry_sequence 부여
    """
    if df.empty:
        return df
    
    print("🔢 1차 정렬 기준으로 bin과 room_entry_sequence 부여 중...")
    
    df = df.copy()
    
    # 방별로 처리
    for room_num in df['roomNumber'].unique():
        room_mask = df['roomNumber'] == room_num
        room_data = df[room_mask]
        room_indices = room_data.index
        total_requests = len(room_data)
        
        # room_entry_sequence 할당 (1부터 시작)
        df.loc[room_indices, 'room_entry_sequence'] = range(1, total_requests + 1)
        
        # bin 할당 (10개 구간으로 균등 분할)
        if total_requests <= BINS_COUNT:
            # 요청이 10개 이하면 각각을 하나의 bin으로
            bins = range(1, total_requests + 1)
        else:
            # 10개 구간으로 균등 분할
            bins = pd.cut(range(total_requests), bins=BINS_COUNT, labels=range(1, BINS_COUNT + 1)).astype(int)
        
        df.loc[room_indices, 'bin'] = bins
        
        print(f"  방 {room_num}: {total_requests}개 레코드에 room_entry_sequence(1~{total_requests})와 bin({len(set(bins))}개 구간) 할당")
    
    print("✅ bin과 room_entry_sequence 부여 완료")
    return df


def sort_final_dataframe_simplified(df: pd.DataFrame) -> pd.DataFrame:
    """
    비이중 확인 구조용 단순화된 정렬 로직
    1차: 모든 이벤트를 waiting_start_nanoTime 기준으로 정렬
    1차 정렬 직후: bin과 room_entry_sequence 부여
    2차: SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime 기준으로 재정렬
    """
    if df.empty:
        return df
    
    print("📊 비이중 확인 구조용 단순화된 정렬 로직 시작...")
    
    # 1차 정렬: 모든 이벤트를 waiting_start_nanoTime 기준으로 정렬
    if 'waiting_start_nanoTime' in df.columns:
        # 나노초 값을 숫자로 변환
        df['waiting_start_nanoTime_num'] = pd.to_numeric(df['waiting_start_nanoTime'], errors='coerce')
        df['waiting_start_nanoTime_num'] = df['waiting_start_nanoTime_num'].fillna(float('inf'))
        
        # 방 번호 → waiting_start_nanoTime 순으로 1차 정렬
        df = df.sort_values(['roomNumber', 'waiting_start_nanoTime_num']).reset_index(drop=True)
        
        print(f"✅ 1차 정렬 완료: 모든 이벤트를 waiting_start_nanoTime 기준으로 정렬")
    else:
        print("⚠️ waiting_start_nanoTime 컬럼이 없어서 1차 정렬을 건너뜁니다.")
        df = df.sort_values(['roomNumber']).reset_index(drop=True)
    
    # 1차 정렬 직후: bin과 room_entry_sequence 부여
    df = assign_bins_and_sequence_after_first_sort(df)
    
    # 2차 정렬: SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime 기준으로 재정렬
    if 'critical_enter_nanoTime' in df.columns and 'join_result' in df.columns:
        
        # 방별로 처리
        final_df_list = []
        
        for room_num in df['roomNumber'].unique():
            room_df = df[df['roomNumber'] == room_num].copy()
            
            print(f"  방 {room_num}: SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime으로 재정렬 중...")
            
            # SUCCESS/FAIL_OVER_CAPACITY와 기타 분리
            paired_mask = (room_df['join_result'] == RESULT_SUCCESS) | (room_df['join_result'] == RESULT_FAIL_CAPACITY)
            room_paired = room_df[paired_mask].copy()
            room_others = room_df[~paired_mask].copy()
            
            if not room_paired.empty:
                # critical_enter_nanoTime을 숫자로 변환
                room_paired['critical_enter_nanoTime_num'] = pd.to_numeric(room_paired['critical_enter_nanoTime'], errors='coerce')
                room_paired['critical_enter_nanoTime_num'] = room_paired['critical_enter_nanoTime_num'].fillna(float('inf'))
                
                # SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime으로 재정렬
                room_paired = room_paired.sort_values(['critical_enter_nanoTime_num']).reset_index(drop=True)
                room_paired = room_paired.drop(columns=['critical_enter_nanoTime_num'])
                
                print(f"    SUCCESS/FAIL_OVER_CAPACITY {len(room_paired)}개 스레드를 critical_enter_nanoTime으로 재정렬")
            
            # 기타는 원래 waiting_start_nanoTime 순서 유지
            if not room_others.empty:
                print(f"    기타 {len(room_others)}개는 waiting_start_nanoTime 순서 유지")
            
            # 재정렬된 데이터 결합
            room_result_list = []
            paired_idx = 0
            others_idx = 0
            
            for _, original_row in room_df.iterrows():
                if original_row['join_result'] == RESULT_SUCCESS or original_row['join_result'] == RESULT_FAIL_CAPACITY:
                    # SUCCESS/FAIL_OVER_CAPACITY는 critical_enter_nanoTime 정렬된 순서대로
                    if paired_idx < len(room_paired):
                        room_result_list.append(room_paired.iloc[paired_idx])
                        paired_idx += 1
                else:
                    # 기타는 원래 waiting_start_nanoTime 순서 유지
                    if others_idx < len(room_others):
                        room_result_list.append(room_others.iloc[others_idx])
                        others_idx += 1
            
            # 혹시 남은 데이터들 추가
            while paired_idx < len(room_paired):
                room_result_list.append(room_paired.iloc[paired_idx])
                paired_idx += 1
            while others_idx < len(room_others):
                room_result_list.append(room_others.iloc[others_idx])
                others_idx += 1
            
            if room_result_list:
                room_final = pd.DataFrame(room_result_list).reset_index(drop=True)
                final_df_list.append(room_final)
        
        if final_df_list:
            df = pd.concat(final_df_list, ignore_index=True)
            print(f"✅ 2차 정렬 완료: SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime 기준으로 재정렬")
        else:
            print("⚠️ 처리할 데이터가 없어서 2차 정렬을 건너뜁니다.")
    else:
        print("⚠️ critical_enter_nanoTime 또는 join_result 컬럼이 없어서 2차 정렬을 건너뜁니다.")
    
    # 정렬용 임시 컬럼 제거
    if 'waiting_start_nanoTime_num' in df.columns:
        df = df.drop(columns=['waiting_start_nanoTime_num'])
    
    return df.reset_index(drop=True)


def build_clean_performance_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    파싱된 로그 데이터를 성능 분석용 데이터로 변환 (비이중 확인 구조용)
    
    처리 과정:
        1. 타임스탬프 변환 및 nanoTime 기준 정렬
        2. 사용자별 이벤트 페어링 처리 (5개 이벤트만)
        3. waiting_start_nanoTime 기준 1차 정렬
        4. 1차 정렬 직후 bin과 room_entry_sequence 부여
        5. critical_enter_nanoTime 기준 2차 정렬 (SUCCESS/FAIL_OVER_CAPACITY만)
    """
    if df.empty:
        return pd.DataFrame()
    
    # 타임스탬프를 datetime 객체로 변환
    df['timestamp'] = pd.to_datetime(df['timestamp'])
    
    # 나노초 값들을 숫자로 변환 (정렬용)
    df['nanoTime_num'] = df['nanoTime'].astype(float)
    df['epochNano_num'] = df['epochNano'].astype(float)
    
    # 방별 + nanoTime 기준 정렬 (기본 정렬은 유지)
    df = df.sort_values(['roomNumber', 'nanoTime_num']).reset_index(drop=True)
    
    performance_results = []
    
    # 각 방별로 처리
    for room_num in df['roomNumber'].unique():
        print(f"🏠 방 {room_num} 처리 중...")
        
        # 방별 5개 이벤트들을 사용자별로 그룹화하여 처리
        room_events = df[df['roomNumber'] == room_num].copy()
        if not room_events.empty:
            for user_id, user_data in room_events.groupby('userId'):
                profile = process_user_events(user_id, user_data)
                performance_results.append(profile)
    
    if not performance_results:
        return pd.DataFrame()
    
    # DataFrame 생성 및 정리
    result_df = pd.DataFrame(performance_results)
    
    # 나노초 컬럼들을 문자열로 확실히 변환
    nano_columns = [col for col in result_df.columns if 'nanoTime' in col or 'epochNano' in col]
    for col in nano_columns:
        if col in result_df.columns:
            result_df[col] = result_df[col].astype(str)
    
    # 최종 정렬: 1차 정렬 → bin/sequence 부여 → 2차 정렬
    result_df = sort_final_dataframe_simplified(result_df)
    
    # 컬럼 순서 정리
    base_columns = ['roomNumber', 'bin', 'user_id', 'room_entry_sequence', 'join_result']
    event_columns = []
    
    # 이벤트별 컬럼들을 순서대로 추가
    for prefix in ['waiting_start', 'critical_enter', 'critical_leave', 'increment_before', 'increment_after']:
        for suffix in ['_time', '_nanoTime', '_epochNano', '_event_type']:
            col_name = prefix + suffix
            if col_name in result_df.columns:
                event_columns.append(col_name)
    
    # 존재하는 컬럼만 선택
    all_columns = base_columns + event_columns
    existing_columns = [col for col in all_columns if col in result_df.columns]
    result_df = result_df[existing_columns]
    
    return result_df


def get_clean_event_desc_table() -> List[List[str]]:
    """
    Excel 파일에 추가할 컬럼 설명 테이블 생성 (비이중 확인 구조용)
    """
    return [
        ["속성명", "측정 목적", "도출 방법"],
        ["roomNumber", "방 번호 식별", "로그 필드: roomNumber"],
        ["bin", "방별 분석 구간", "각 방의 요청을 10개 구간으로 균등 분할"],
        ["user_id", "사용자 식별", "로그 필드: userId"],
        ["room_entry_sequence", "방별 처리 순번", "단순 순차 할당"],
        ["join_result", "입장 결과 구분", "SUCCESS/FAIL_OVER_CAPACITY/UNKNOWN"],
        ["waiting_start_*", "대기 시작 시점", "WAITING_START 이벤트 속성들"],
        ["critical_enter_*", "임계구역 진입 시점", "CRITICAL_ENTER 이벤트 속성들"],
        ["critical_leave_*", "임계구역 진출 시점", "CRITICAL_LEAVE 이벤트 속성들"],
        ["increment_before_*", "증가 작업 시작 시점", "INCREMENT_BEFORE 이벤트 속성들"],
        ["increment_after_*", "증가 작업 완료 시점", "INCREMENT_AFTER 이벤트 속성들"],
        ["*_time", "이벤트 발생 시간", "timestampIso (표시용)"],
        ["*_nanoTime", "나노초 정밀도 시간", "1차: waiting_start_nanoTime, 2차: critical_enter_nanoTime 정렬"],
        ["*_epochNano", "Epoch 나노초 시간", "Epoch 기준 나노초 (참조용)"],
        ["*_event_type", "이벤트 상세 타입", "SUCCESS/FAIL_OVER_CAPACITY 등"],
        ["비이중 확인 구조", "락 외부 사전 확인 없음", "PRE_CHECK_FAIL_OVER_CAPACITY 이벤트 없음"]
    ]


def convert_nano_value(x: Any) -> str:
    """
    나노초 값을 문자열로 안전하게 변환
    """
    if pd.isna(x) or x == '' or str(x).lower() == 'nan':
        return ''
    try:
        return str(int(float(x)))
    except (ValueError, TypeError):
        return str(x)


def save_to_csv(df: pd.DataFrame, filepath: str) -> None:
    """
    DataFrame을 CSV 파일로 저장 (나노초 정밀도 유지)
    """
    # 나노초 컬럼들을 문자열로 변환하여 지수 표기법 방지
    df_copy = df.copy()
    nano_columns = [col for col in df.columns if 'nanoTime' in col or 'epochNano' in col]
    
    for col in nano_columns:
        if col in df_copy.columns:
            df_copy[col] = df_copy[col].apply(convert_nano_value)
    
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # UTF-8 with BOM으로 저장 (Excel에서 한글 깨짐 방지)
    df_copy.to_csv(filepath, index=False, encoding='utf-8-sig')
    print(f"CSV 파일 저장 완료: {filepath}")


def is_datetime_with_tz(column):
    """
    datetime 컬럼의 timezone 여부를 확인하는 호환성 함수
    """
    try:
        # 최신 pandas 버전 시도
        return pd.api.types.is_datetime64tz_dtype(column)
    except AttributeError:
        try:
            # 구버전 pandas 시도
            return pd.api.types.is_datetime64_tz_dtype(column)
        except AttributeError:
            # 모든 방법이 실패하면 직접 확인
            return hasattr(column.dtype, 'tz') and column.dtype.tz is not None


def save_with_side_table(df: pd.DataFrame, filepath: str, desc_table: List[List[str]]) -> str:
    """
    DataFrame을 Excel로 저장하고 설명 테이블 추가 (pandas 호환성 개선)
    """
    # 출력 디렉토리 생성
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # timezone 제거한 복사본 생성 (호환성 개선)
    df_excel = df.copy()
    for col in df_excel.columns:
        if is_datetime_with_tz(df_excel[col]):
            df_excel[col] = df_excel[col].dt.tz_localize(None)
    
    # Excel 파일 생성
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False)
        
        # 나노초 컬럼들을 텍스트 형식으로 설정
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        nano_columns = [col for col in df.columns if 'nanoTime' in col or 'epochNano' in col]
        for col_name in nano_columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '@'  # 텍스트 형식
    
    # 설명 테이블 추가
    wb = load_workbook(filepath)
    ws = wb.active
    
    start_col = len(df.columns) + 2
    
    for i, row in enumerate(desc_table):
        for j, val in enumerate(row):
            ws.cell(row=i + 1, column=start_col + j, value=val)
    
    wb.save(filepath)
    print(f"Excel 파일 저장 완료: {filepath}")
    
    return filepath


def analyze_clean_results(df: pd.DataFrame) -> None:
    """
    처리 결과 분석 및 출력 (비이중 확인 구조용)
    """
    if df.empty:
        print("분석할 데이터가 없습니다.")
        return
    
    total_operations = len(df)
    
    print(f"\n{'='*60}")
    print(f"비이중 확인 구조 분석 결과")
    print(f"{'='*60}")
    print(f"전체 처리 작업 수: {total_operations}")
    
    # 입장 결과별 분석
    if 'join_result' in df.columns:
        print(f"\n[입장 결과별 분석]")
        join_result_counts = df['join_result'].value_counts()
        for result, count in join_result_counts.items():
            print(f"  {result}: {count}건 ({count/total_operations*100:.1f}%)")
    
    # 이벤트 완성도 분석
    events = ['waiting_start_time', 'critical_enter_time', 'critical_leave_time', 
              'increment_before_time', 'increment_after_time']
    
    print(f"\n[이벤트 완성도 분석]")
    for event in events:
        if event in df.columns:
            complete_count = df[event].notna().sum()
            print(f"  {event}: {complete_count}건 ({complete_count/total_operations*100:.1f}%)")
    
    # 완전한 5개 이벤트 세션 분석
    complete_sessions = 0
    success_complete = 0
    fail_complete = 0
    
    for _, row in df.iterrows():
        if all(pd.notna(row.get(event)) and row.get(event) != '' for event in events):
            complete_sessions += 1
            if row.get('join_result') == RESULT_SUCCESS:
                success_complete += 1
            elif row.get('join_result') == RESULT_FAIL_CAPACITY:
                fail_complete += 1
    
    print(f"\n[완전한 5개 이벤트 세션 분석]")
    print(f"  완전한 세션: {complete_sessions}건 ({complete_sessions/total_operations*100:.1f}%)")
    print(f"    - 성공: {success_complete}건")
    print(f"    - 실패: {fail_complete}건")
    
    # 방별 성능 통계
    print(f"\n[방별 성능 통계]")
    room_stats = df.groupby('roomNumber').agg({
        'user_id': 'count',
        'bin': 'nunique'
    })
    room_stats.columns = ['total_operations', 'bin_count']
    
    for room_num, stats in room_stats.iterrows():
        print(f"  방 {room_num}: 총 {stats['total_operations']}회, 구간 수: {stats['bin_count']}")
        
        if 'join_result' in df.columns:
            room_data = df[df['roomNumber'] == room_num]
            room_results = room_data['join_result'].value_counts()
            for result, count in room_results.items():
                print(f"    - {result}: {count}건")
    
    # 비이중 확인 구조 특징 확인
    print(f"\n[비이중 확인 구조 특징 확인]")
    print(f"  - PRE_CHECK_FAIL 이벤트: 0건 (비이중 확인 구조에서는 발생하지 않음)")
    print(f"  - 락 외부 사전 확인 없음")
    print(f"  - 5개 이벤트만 처리 (SUCCESS/FAIL_OVER_CAPACITY/UNKNOWN)")
    
    # 단순화된 정렬 순서 확인
    print(f"\n[단순화된 정렬 순서 확인]")
    if 'waiting_start_nanoTime' in df.columns and 'critical_enter_nanoTime' in df.columns:
        for room_num in sorted(df['roomNumber'].unique())[:3]:  # 첫 3개 방만 확인
            room_data = df[df['roomNumber'] == room_num].head(5)
            print(f"  방 {room_num}:")
            for i, (_, row) in enumerate(room_data.iterrows()):
                seq = row['room_entry_sequence']
                waiting_nano = row.get('waiting_start_nanoTime', 'N/A')
                critical_nano = row.get('critical_enter_nanoTime', 'N/A')
                user_id = row['user_id']
                result = row['join_result']
                print(f"    순번{seq}: waiting={waiting_nano}, critical={critical_nano} (user: {user_id}, result:{result})")
    else:
        print("  필요한 컬럼이 없어서 확인할 수 없습니다.")
    
    # 나노초 정밀도 확인
    print(f"\n[나노초 정밀도 확인 (첫 3개 샘플)]")
    nano_columns = [col for col in df.columns if 'nanoTime' in col]
    if nano_columns:
        for i in range(min(3, len(df))):
            print(f"  샘플 {i+1}:")
            for col in nano_columns[:3]:  # 첫 3개 나노초 컬럼만
                if col in df.columns:
                    print(f"    {col}: {df.iloc[i][col]}")


def main():
    """
    메인 실행 함수 (비이중 확인 구조용)
    """
    parser = argparse.ArgumentParser(
        description="비이중 확인 구조용 전처리기: PRE_CHECK_FAIL 없는 5개 이벤트 처리",
        epilog="예시: py -3 preprocess_logs_single_check.py --output_dir C:\\single_check_analysis\\ --room 1 --xlsx single_check_output.xlsx"
    )
    parser.add_argument('--output_dir', type=str, default='results_single_check', 
                        help='출력 디렉토리 경로 (기본값: results_single_check)')
    parser.add_argument('--room', type=int, help='특정 방 번호만 처리 (옵션)')
    parser.add_argument('--csv', type=str, help='추가 CSV 파일명 (옵션)')
    parser.add_argument('--xlsx', type=str, help='Excel 파일명 (옵션)')
    parser.add_argument('--test', action='store_true', help='정규식 패턴 테스트 실행')
    
    args = parser.parse_args()
    
    # 정규식 테스트 옵션
    if args.test:
        print("🧪 CRITICAL_PATTERN 정규식 테스트 실행 (비이중 확인용)...")
        test_result = test_critical_pattern()
        if test_result:
            print("✅ 정규식 테스트 통과")
        else:
            print("❌ 정규식 테스트 실패 - 패턴을 재검토하세요")
        return
    
    try:
        # 1. 로그 파일 교체
        replace_log_file()
        
        # 2. 로그 파싱
        print(f"\n로그 파일 파싱 중 (비이중 확인 구조)...")
        df = parse_five_events_clean(LOG_FILE, room_number=args.room)
        print(f"파싱 완료: {len(df)}개 이벤트")
        
        # 3. 성능 데이터 구축 (비이중 확인 구조용)
        print(f"\n비이중 확인 구조 데이터 구축 중...")
        result = build_clean_performance_data(df)
        print(f"구축 완료: {len(result)}개 세션")
        
        # 4. 출력 디렉토리 설정 및 생성
        output_dir = args.output_dir
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"출력 디렉토리 생성: {output_dir}")
        
        # 5. 기본 CSV 파일 저장
        if args.room:
            base_filename = f'room{args.room}_single_check.csv'
        else:
            base_filename = 'all_rooms_single_check.csv'
        
        csv_path = os.path.join(output_dir, base_filename)
        save_to_csv(result, csv_path)
        
        # 6. 추가 CSV 파일 저장 (옵션)
        if args.csv:
            additional_csv_path = os.path.join(output_dir, args.csv)
            save_to_csv(result, additional_csv_path)
        
        # 7. Excel 파일 저장 (옵션)
        if args.xlsx:
            xlsx_path = os.path.join(output_dir, args.xlsx)
            desc_table = get_clean_event_desc_table()
            save_with_side_table(result, xlsx_path, desc_table)
        
        # 8. 결과 분석 출력
        analyze_clean_results(result)
        
        print(f"\n{'='*60}")
        print(f"비이중 확인 구조 처리 완료!")
        print(f"특징: PRE_CHECK_FAIL 이벤트 없음, 5개 이벤트만 처리")
        print(f"정렬 순서: 1차(waiting_start_nanoTime) → 2차(SUCCESS/FAIL_OVER_CAPACITY만 critical_enter_nanoTime)")
        print(f"출력 디렉토리: {os.path.abspath(output_dir)}")
        print(f"{'='*60}")
        
    except FileNotFoundError as e:
        print(f"\n오류: 파일을 찾을 수 없습니다 - {e}")
        sys.exit(1)
    except PermissionError as e:
        print(f"\n오류: 파일 접근 권한이 없습니다 - {e}")
        sys.exit(1)
    except Exception as e:
        print(f"\n예상치 못한 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


# 스크립트가 직접 실행될 때만 main() 호출
if __name__ == '__main__':
    main()